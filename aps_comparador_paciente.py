"""
aps_comparador_paciente.py
Cruza planilhas C1-C7 por paciente.
Exporta planilha visualmente rica com 3 abas.
"""
from __future__ import annotations

import os, re, unicodedata
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from aps_utils import infer_indicator_code_from_path

# â”€â”€ Paleta â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
C = {
    "azul":    "1F4E79", "azul_m":  "2E75B6", "azul_h":  "BDD7EE", "azul_c": "EBF3FB",
    "verde":   "C6EFCE", "verde_e": "375623", "verde_c": "E8F5E9",
    "amarelo": "FFEB9C", "amar_t":  "9C5700", "amar_m":  "FFE699",
    "verm":    "FFC7CE", "verm_t":  "9C0006", "verm_e":  "C00000", "verm_c": "FFEBEE",
    "laranja": "FFF2CC", "lar_t":   "843C0C",
    "roxo":    "7030A0", "roxo_c":  "F3E5F5",
    "cinza":   "F5F5F5", "cinza_m": "E0E0E0", "branco":  "FFFFFF",
}
# Mapeia AMBOS os formatos: com emoji (planilha interativa) e sem (sistema original)
def _norm_prio(p: str) -> str:
    """
    Normaliza prioridade para formato padronizado com emoji.
    Cobre dois formatos:
      Sistema APS original: "Alta", "MÃ©dia", "Baixa", "ConcluÃ­do"
      Clonador interativo:  "ðŸ”´ URGENTE", "ðŸŸ  ALTA", "ðŸŸ¡ MONITORAR", "ðŸŸ¢ CONCLUÃDO"
    """
    p = str(p).strip()
    # 1. Verifica formato com emoji direto (do clonador) â€” PRIMEIRO para nÃ£o confundir
    if "URGENTE" in p: return "ðŸ”´ URGENTE"
    if "MONITORAR" in p or "MONITOR" in p: return "ðŸŸ¡ MONITORAR"
    if "CONCLUÃDO" in p or "CONCLUIDO" in p: return "ðŸŸ¢ CONCLUÃDO"
    if "ðŸŸ " in p: return "ðŸŸ  ALTA"
    if "ðŸŸ¡" in p: return "ðŸŸ¡ MONITORAR"
    if "ðŸŸ¢" in p: return "ðŸŸ¢ CONCLUÃDO"
    if "ðŸ”´" in p: return "ðŸ”´ URGENTE"
    # 2. Formato texto sem emoji (sistema APS original)
    pu = p.upper().strip()
    if pu in ("ALTA",):                    return "ðŸ”´ URGENTE"   # Alta = pior no sistema original
    if pu in ("MÃ‰DIA", "MEDIA"):           return "ðŸŸ  ALTA"
    if pu in ("BAIXA",):                   return "ðŸŸ¡ MONITORAR"
    if pu in ("CONCLUÃDO","CONCLUIDO","Ã“TIMO","OTIMO"): return "ðŸŸ¢ CONCLUÃDO"
    # Fallback por substring
    if "URGENT" in pu or "CRITI" in pu:   return "ðŸ”´ URGENTE"
    if "BAIXA" in pu or "BOM" in pu:      return "ðŸŸ¡ MONITORAR"
    if "CONCLU" in pu or "Ã“TIMO" in pu:   return "ðŸŸ¢ CONCLUÃDO"
    return p

PRIO_ORDER = {"ðŸ”´ URGENTE":0,"ðŸŸ  ALTA":1,"ðŸŸ¡ MONITORAR":2,"ðŸŸ¢ CONCLUÃDO":3}
PRIO_THEME = {
    "ðŸ”´ URGENTE":   (C["verm_c"],   C["verm_e"]),
    "ðŸŸ  ALTA":      (C["laranja"],  C["lar_t"]),
    "ðŸŸ¡ MONITORAR": (C["amarelo"],  C["amar_t"]),
    "ðŸŸ¢ CONCLUÃDO": (C["verde_c"],  C["verde_e"]),
}

def _f(h):  return PatternFill("solid", fgColor=h)
def _fn(bold=False,color="000000",size=9,italic=False):
    return Font(name="Calibri",bold=bold,color=color,size=size,italic=italic)
def _al(h="left",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def _bd(style="thin",color="CCCCCC"):
    s=Side(style=style,color=color); return Border(left=s,right=s,top=s,bottom=s)

def _norm(t):
    if t is None: return ""
    t = unicodedata.normalize("NFKD",str(t).strip()).encode("ascii","ignore").decode("ascii")
    return re.sub(r"\s+"," ",t).lower().strip()

def _detect_code(p:Path)->str:
    code = infer_indicator_code_from_path(p)
    return code if code else p.stem[:4].upper()

def _read(path:Path):
    try:
        xls=pd.ExcelFile(path)
        sh=next((s for s in xls.sheet_names if str(s).startswith("ðŸ“‹ Dados")),xls.sheet_names[0])
        for h in (2,1,0):
            try:
                df=pd.read_excel(path,sheet_name=sh,header=h,dtype=str).dropna(how="all")
                if "Nome" in df.columns and len(df):
                    return df[df["Nome"].str.strip().ne("")].reset_index(drop=True)
            except: continue
    except: pass
    return None

# â”€â”€ Motor â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def build_unified(paths:list[Path])->pd.DataFrame:
    def _source_penalty(path: Path) -> int:
        n = path.name.lower()
        penalty = 0
        if "backup" in n:
            penalty += 100
        if "interativa" in n:
            penalty += 40
        if "cruz" in n or "compar" in n or "unificad" in n:
            penalty += 20
        return penalty

    def _pick_best_source(current: Path | None, candidate: Path) -> Path:
        if current is None:
            return candidate
        pc = _source_penalty(current)
        pn = _source_penalty(candidate)
        if pn < pc:
            return candidate
        if pn > pc:
            return current
        try:
            mc = current.stat().st_mtime_ns
        except Exception:
            mc = 0
        try:
            mn = candidate.stat().st_mtime_ns
        except Exception:
            mn = 0
        return candidate if mn >= mc else current
    def _get(row, *names):
        cols = list(row.index)
        for n in names:
            # Se houver colunas duplicadas no pandas (ex.: "Pontuação", "Pontuação.1"),
            # prefere a ocorrência mais recente com valor preenchido.
            target = _norm(n)
            candidates = []
            for c in cols:
                nc = _norm(c)
                if nc == target or nc.startswith(target + " "):
                    candidates.append(c)
            if not candidates and n in row.index:
                candidates = [n]
            if candidates:
                for c in reversed(candidates):
                    v = row.get(c, "")
                    if pd.isna(v):
                        continue
                    txt = str(v).strip()
                    if txt != "":
                        return txt
                v = row.get(candidates[-1], "")
                return "" if pd.isna(v) else str(v).strip()
        return ""

    def _get_contains(row, *tokens):
        norm_tokens = [_norm(t) for t in tokens if str(t).strip()]
        if not norm_tokens:
            return ""
        fallback = ""
        for col in row.index:
            nc = _norm(col)
            if all(tok in nc for tok in norm_tokens):
                v = row.get(col, "")
                if pd.isna(v):
                    continue
                txt = str(v).strip()
                if txt:
                    fallback = txt
        return fallback

    def _compose_full_address(row, bairro_hint=""):
        endereco_base = _get(row, "Endereço", "Endereco", "Logradouro", "Endereço completo", "Endereco completo")
        numero = _get(row, "Número", "Numero", "Num")
        complemento = _get(row, "Complemento")
        bairro = _get(row, "Bairro", "Bairro/Localidade", "Localidade") or str(bairro_hint or "").strip()
        cidade = _get(row, "Cidade", "Município", "Municipio")
        uf = _get(row, "UF", "Estado")
        cep = _get(row, "CEP")

        rua = " ".join([x for x in [endereco_base, numero] if x]).strip()
        if complemento:
            rua = f"{rua} - {complemento}" if rua else complemento

        local = ", ".join([x for x in [bairro, cidade] if x]).strip()
        if uf:
            local = f"{local}/{uf}" if local else uf
        if cep:
            local = f"{local} - CEP {cep}" if local else f"CEP {cep}"

        full = " | ".join([x for x in [rua, local] if x]).strip()
        return full

    def _pick_more_complete(current, candidate):
        cur = str(current or "").strip()
        new = str(candidate or "").strip()
        if not cur:
            return new
        if len(new) > len(cur):
            return new
        return cur

    def _to_num(value):
        txt = str(value or "").strip()
        if not txt:
            return None
        txt = txt.replace(",", ".")
        m = re.search(r"-?\d+(?:\.\d+)?", txt)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None

    by_code: dict[str, Path] = {}
    for p in paths:
        code = _detect_code(p)
        by_code[code] = _pick_best_source(by_code.get(code), p)

    data={}
    for code, p in by_code.items():
        df=_read(p)
        if df is None or "Nome" not in df.columns:
            continue
        df["_nn"]=df["Nome"].apply(_norm)
        data[code]=df
    if not data: return pd.DataFrame()

    patients={}
    for code,df in data.items():
        for _,row in df.iterrows():
            nn=row["_nn"]
            if not nn: continue
            tel = _get(row, "Telefone celular", "Telefone residencial", "Telefone de contato", "Telefone")
            bairro = _get(row, "Bairro", "Bairro/Localidade", "Localidade")
            endereco_full = _compose_full_address(row, bairro)
            if nn not in patients:
                patients[nn]={
                    "Nome":row.get("Nome",""),
                    "Microárea":_get(row, "Microárea", "Microarea"),
                    "Bairro":bairro,
                    "Endereco":endereco_full,
                    "Telefone":tel if tel not in {"","nan"} else "",
                }
            else:
                if bairro:
                    patients[nn]["Bairro"] = _pick_more_complete(patients[nn].get("Bairro"), bairro)
                if endereco_full:
                    patients[nn]["Endereco"] = _pick_more_complete(patients[nn].get("Endereco"), endereco_full)
                if not patients[nn].get("Telefone") and tel and tel not in {"","nan"}:
                    patients[nn]["Telefone"] = tel
                if not patients[nn].get("Microárea"):
                    patients[nn]["Microárea"] = _get(row, "Microárea", "Microarea")

    codes=sorted(data.keys())
    records=[]
    for nn,ident in patients.items():
        rec={
            "Nome":ident.get("Nome",""),
            "Microárea":ident.get("Microárea",""),
            "Bairro":ident.get("Bairro",""),
            "Endereco":ident.get("Endereco",""),
            "Telefone":ident.get("Telefone",""),
        }
        present=[]; pend_parts=[]; prios=[]; sum_pts=0; cnt=0
        for code in codes:
            match = data[code][data[code]["_nn"] == nn]
            if match.empty:
                rec[code] = "â€”"
                continue

            best_pts_num = None
            best_pts_txt = ""
            best_pend = ""

            # Em caso de paciente duplicado no mesmo indicador, consolida:
            # usa a maior pontuacao valida para refletir o estado mais atualizado.
            for _, r in match.iterrows():
                pts_candidates = [
                    _get(r, "Pontuação"),
                    _get(r, "Pontuacao"),
                    _get(r, "PontuaÃ§Ã£o"),
                    _get_contains(r, "pontu"),
                ]
                pts_candidates = [str(x).strip() for x in pts_candidates if str(x).strip()]
                pts_txt = pts_candidates[-1] if pts_candidates else ""
                pts_num = None
                for cand in pts_candidates:
                    num = _to_num(cand)
                    if num is None:
                        continue
                    if pts_num is None or num > pts_num:
                        pts_num = num
                        pts_txt = cand

                pend_candidates = [
                    _get(r, "Pendências"),
                    _get(r, "Pendencias"),
                    _get(r, "PendÃªncias"),
                    _get_contains(r, "pend"),
                ]
                pend_candidates = [str(x).strip() for x in pend_candidates if str(x).strip()]
                pend_txt = pend_candidates[-1] if pend_candidates else ""

                if best_pts_num is None:
                    best_pts_num = pts_num
                    best_pts_txt = pts_txt
                    best_pend = pend_txt
                    continue

                if pts_num is not None and (best_pts_num is None or pts_num > best_pts_num):
                    best_pts_num = pts_num
                    best_pts_txt = pts_txt
                    best_pend = pend_txt
                elif pts_num == best_pts_num:
                    # Em empate de pontuacao, prefere a linha sem pendencia.
                    cur_pend = str(best_pend or "").strip().lower()
                    new_pend = str(pend_txt or "").strip().lower()
                    if cur_pend not in {"", "none", "-", "nan"} and new_pend in {"", "none", "-", "nan"}:
                        best_pts_txt = pts_txt
                        best_pend = pend_txt

            rec[code] = f"{best_pts_txt} pts" if str(best_pts_txt).strip() else "â€”"
            present.append(code)
            if best_pts_num is not None:
                sum_pts += float(best_pts_num)
                cnt += 1
            if best_pend and str(best_pend).strip().lower() not in {"", "nan", "none", "-"}:
                pend_parts.append(f"[{code}] {best_pend}")
        media=int(round(sum_pts/cnt,0)) if cnt else 0
        # Prioridade calculada SEMPRE pela pontuaÃ§Ã£o mÃ©dia â€” consistente independente do formato
        if   media >= 100: best = "ðŸŸ¢ CONCLUÃDO"
        elif media >= 75:  best = "ðŸŸ¡ MONITORAR"
        elif media >= 50:  best = "ðŸŸ  ALTA"
        else:              best = "ðŸ”´ URGENTE"
        rec.update({
            "Indicadores": " Â· ".join(present) if present else "â€”",
            "Qtd":len(present), "PendÃªncias":len(pend_parts),
            "MÃ©dia":media, "Prioridade":best,
            "O que fazer":"\n".join(pend_parts) if pend_parts else "âœ” Em dia",
        })
        records.append(rec)

    df_out=pd.DataFrame(records)
    df_out["_o"]=df_out["Prioridade"].map(lambda p:PRIO_ORDER.get(p,4))
    # Ordena por: prioridade (urgente primeiro) â†’ pontuaÃ§Ã£o mÃ©dia (menor primeiro) â†’ nome
    return df_out.sort_values(["_o","MÃ©dia","Nome"],ascending=[True,True,True]).drop(columns=["_o"]).reset_index(drop=True)

def build_indicator_snapshot(path: Path) -> tuple[str, dict[str, dict]]:
    """
    Lê uma planilha de indicador e devolve:
      (codigo_indicador, {nome_normalizado: {"nome","pts","pend","bairro","endereco","telefone"}})
    """
    def _to_num(value):
        txt = str(value or "").strip().replace(",", ".")
        m = re.search(r"-?\d+(?:\.\d+)?", txt)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None

    def _get(row, *names):
        cols = list(row.index)
        for n in names:
            target = _norm(n)
            candidates = []
            for c in cols:
                nc = _norm(c)
                if nc == target or nc.startswith(target + " "):
                    candidates.append(c)
            if not candidates and n in row.index:
                candidates = [n]
            if candidates:
                for c in reversed(candidates):
                    v = row.get(c, "")
                    if pd.isna(v):
                        continue
                    txt = str(v).strip()
                    if txt:
                        return txt
        return ""

    def _get_contains(row, *tokens):
        ts = [_norm(t) for t in tokens if str(t).strip()]
        out = ""
        for c in row.index:
            nc = _norm(c)
            if all(t in nc for t in ts):
                v = row.get(c, "")
                if pd.isna(v):
                    continue
                txt = str(v).strip()
                if txt:
                    out = txt
        return out

    code = _detect_code(path)
    df = _read(path)
    if df is None or "Nome" not in df.columns:
        return code, {}

    out: dict[str, dict] = {}
    for _, row in df.iterrows():
        nome = str(row.get("Nome", "") or "").strip()
        nn = _norm(nome)
        if not nn:
            continue

        pts_candidates = [
            _get(row, "Pontuação"),
            _get(row, "Pontuacao"),
            _get(row, "PontuaÃ§Ã£o"),
            _get_contains(row, "pontu"),
        ]
        pts_candidates = [str(x).strip() for x in pts_candidates if str(x).strip()]
        pts_txt = pts_candidates[-1] if pts_candidates else ""
        pts_num = None
        for cand in pts_candidates:
            num = _to_num(cand)
            if num is None:
                continue
            if pts_num is None or num > pts_num:
                pts_num = num
                pts_txt = cand
        if pts_num is None:
            pts_num = 0.0

        pend_candidates = [
            _get(row, "Pendências"),
            _get(row, "Pendencias"),
            _get(row, "PendÃªncias"),
            _get_contains(row, "pend"),
        ]
        pend_candidates = [str(x).strip() for x in pend_candidates if str(x).strip()]
        pend_txt = pend_candidates[-1] if pend_candidates else ""
        if pend_txt.lower() in {"none", "nan", "-"}:
            pend_txt = ""

        cur = out.get(nn)
        if cur is None or pts_num >= float(cur.get("pts", 0)):
            out[nn] = {
                "nome": nome,
                "pts": float(pts_num),
                "pend": pend_txt,
                "bairro": _get(row, "Bairro", "Bairro/Localidade", "Localidade"),
                "endereco": _get(row, "Endereco", "Endereço", "EndereÃ§o", "Logradouro", "Rua"),
                "telefone": _get(row, "Telefone celular", "Telefone residencial", "Telefone de contato", "Telefone"),
            }
    return code, out
# â”€â”€ Excel premium â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _title(ws,row,text,bg,fg,size,ncols,h=28):
    ecl=get_column_letter(ncols)
    try: ws.merge_cells(f"A{row}:{ecl}{row}")
    except: pass
    c=ws.cell(row,1,text)
    c.font=_fn(bold=True,color=fg,size=size); c.fill=_f(bg)
    c.alignment=_al("center"); c.border=_bd("medium","595959")
    ws.row_dimensions[row].height=h

def _mw(ws,row,col,value,ec=None,er=None):
    """Escreve com merge opcional (como no clonador)."""
    if ec or er:
        r2=er or row; c2=ec or col
        try: ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(c2)}{r2}")
        except: pass
    cell=ws.cell(row,col)
    cell.value=value
    return cell

def _section(ws,row,text,bg,fg,ncols):
    ecl=get_column_letter(ncols)
    try: ws.merge_cells(f"A{row}:{ecl}{row}")
    except: pass
    c=ws.cell(row,1,text)
    c.font=_fn(bold=True,color=fg,size=9); c.fill=_f(bg)
    c.alignment=_al("left"); c.border=_bd()
    ws.row_dimensions[row].height=18

def export_excel(df:pd.DataFrame, out_path:Path):
    """
    Gera planilha de cruzamento com 4 abas e visual profissional:
      ðŸ“‹ Pacientes  â€” lista completa com separadores por prioridade
      ðŸ”´ Urgentes   â€” top crÃ­ticos
      ðŸ“Š Resumo     â€” mini-dashboard com barras e KPIs
      ðŸŸ¢ Em Dia     â€” concluÃ­dos
    """
    from openpyxl.styles import Font as XFont
    wb=Workbook()
    ind_cols=sorted([c for c in df.columns if re.match(r"^C\d$",c)])
    total_pac=len(df)
    _prio_col = df["Prioridade"].astype(str) if "Prioridade" in df.columns else pd.Series(dtype=str)
    _pend_col = pd.to_numeric(df["PendÃªncias"], errors="coerce").fillna(0) if "PendÃªncias" in df.columns else pd.Series(dtype=float)
    urgentes = int(_prio_col.str.contains("URGENTE", na=False).sum())
    em_dia   = int((_pend_col == 0).sum())
    multi    = int((_pend_col > 1).sum())

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ABA 1: Pacientes â€” lista completa
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    ws=wb.active; ws.title="ðŸ“‹ Pacientes"
    cols=(["Nome","Microárea","Bairro","Endereco","Telefone"]+ind_cols
          +["Indicadores","Qtd","PendÃªncias","MÃ©dia","Prioridade","O que fazer"])
    cols=[c for c in cols if c in df.columns]; n=len(cols)

    # Linha 1 â€” TÃ­tulo principal
    _title(ws,1,"APS  Â·  CRUZAMENTO UNIFICADO POR PACIENTE",C["azul"],C["branco"],15,n,36)

    # Linha 2 â€” Barra de KPIs inline
    kpi=(f"  ðŸ‘¥ {total_pac} pacientes   |   ðŸ”´ {urgentes} urgentes   |   "
         f"âœ” {em_dia} em dia   |   âš  {multi} com 2+ pendÃªncias   |   "
         f"ðŸ“… {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _title(ws,2,kpi,C["azul_c"],C["azul"],9,n,18)

    # Linha 3 â€” Legenda semÃ¢ntica
    _title(ws,3,
           "  ðŸŸ¢ ConcluÃ­do = 100 pts     ðŸŸ¡ Monitorar = 75-99     ðŸŸ  Alta = 50-74     "
           "ðŸ”´ Urgente < 50 pts     â¬œ Cinza = nÃ£o estÃ¡ nesta lista",
           C["roxo_c"],C["roxo"],8,n,13)

    # Linha 4 â€” Banner separador
    c4=_mw(ws,4,1,"  â–¼  LISTA COMPLETA  â€”  ordenada por prioridade e nÃºmero de pendÃªncias",ec=n)
    c4.font=_fn(bold=True,color=C["branco"],size=9); c4.fill=_f(C["azul_m"])
    c4.alignment=_al("left"); c4.border=_bd(); ws.row_dimensions[4].height=16

    # Linha 5 â€” CabeÃ§alhos (duas faixas de cor: azul p/ identificaÃ§Ã£o, amarelo p/ indicadores, roxo p/ sÃ­ntese)
    for ci,col in enumerate(cols,1):
        if   col in ("Nome","Microárea","Bairro","Endereco","Telefone"): bg=C["azul"];  fg=C["branco"]
        elif col in ind_cols:                        bg=C["amar_m"];fg="1F3864"
        else:                                        bg=C["roxo"];  fg=C["branco"]
        cell=ws.cell(5,ci,col)
        cell.fill=_f(bg); cell.font=_fn(bold=True,color=fg,size=9)
        cell.alignment=_al("center",wrap=True); cell.border=_bd("medium","595959")
    ws.row_dimensions[5].height=42

    # Dados com separadores visuais por grupo de prioridade
    prev_prio=None; dr=6
    PRIO_LABEL={"ðŸ”´ URGENTE":"ðŸ”´  URGENTE â€” aÃ§Ã£o imediata",
                "ðŸŸ  ALTA":   "ðŸŸ   ALTA â€” acompanhamento prÃ³ximo",
                "ðŸŸ¡ MONITORAR":"ðŸŸ¡  MONITORAR â€” prÃ³ximo de concluir",
                "ðŸŸ¢ CONCLUÃDO":"ðŸŸ¢  CONCLUÃDO â€” todas as listas completas"}

    for _,row in df[cols].iterrows():
        prio=str(row.get("Prioridade",""))
        npend=int(row.get("PendÃªncias",0) or 0)

        if prio!=prev_prio and prio in PRIO_THEME:
            bg_s,fg_s=PRIO_THEME[prio]
            try: ws.merge_cells(f"A{dr}:{get_column_letter(n)}{dr}")
            except: pass
            sep=ws.cell(dr,1,f"  {PRIO_LABEL.get(prio,prio)}")
            sep.fill=_f(bg_s); sep.font=_fn(bold=True,color=fg_s,size=9)
            sep.alignment=_al("left"); sep.border=_bd("medium","888888")
            ws.row_dimensions[dr].height=15; dr+=1; prev_prio=prio

        row_bg,row_fg=PRIO_THEME.get(prio,(C["cinza"],C["azul"]))
        if npend==0: row_bg=C["verde_c"]; row_fg=C["verde_e"]
        # Zebra suave dentro do grupo
        is_even=(dr%2==0)

        for ci,col in enumerate(cols,1):
            val=row.get(col,""); vs="" if str(val) in {"nan","None"} else str(val)
            cell=ws.cell(dr,ci,vs); cell.border=_bd()

            if col in ("Nome","Microárea","Bairro","Endereco","Telefone"):
                cell.fill=_f("F8FBFF" if is_even else C["branco"])
                cell.font=_fn(bold=(col=="Nome"),size=9,color="1A1A2E")
                cell.alignment=_al("left")

            elif col in ind_cols:
                cell.alignment=_al("center")
                if vs=="â€”":
                    cell.fill=_f("EEEEEE"); cell.font=_fn(color="BBBBBB",size=8)
                    cell.value="â€”"
                else:
                    try:
                        pts=float(vs.replace("pts","").strip())
                        if pts>=100:
                            cell.fill=_f(C["verde"]); cell.font=_fn(bold=True,color=C["verde_e"],size=9)
                            cell.value="âœ” 100"
                        elif pts>=75:
                            cell.fill=_f(C["amar_m"]); cell.font=_fn(bold=True,color=C["amar_t"],size=9)
                            cell.value=f"âš¡{int(pts)}"
                        elif pts>=50:
                            cell.fill=_f(C["amarelo"]); cell.font=_fn(bold=True,color=C["amar_t"],size=9)
                            cell.value=f"âš {int(pts)}"
                        else:
                            cell.fill=_f(C["verm"]); cell.font=_fn(bold=True,color=C["verm_t"],size=9)
                            cell.value=f"ðŸ”´{int(pts)}"
                    except:
                        cell.fill=_f(C["cinza"]); cell.font=_fn(size=9)

            elif col=="Prioridade":
                bg2,fg2=PRIO_THEME.get(vs,(C["cinza"],C["azul"]))
                cell.fill=_f(bg2); cell.font=_fn(bold=True,color=fg2,size=9)
                cell.alignment=_al("center")

            elif col=="O que fazer":
                cell.fill=_f(row_bg)
                cell.font=_fn(size=9,color=C["verm_t"] if npend>0 else C["verde_e"],italic=(npend==0))
                cell.alignment=_al("left",wrap=True)

            elif col in ("Qtd","PendÃªncias","MÃ©dia"):
                cell.fill=_f(row_bg); cell.font=_fn(bold=True,color=row_fg,size=10)
                cell.alignment=_al("center")

            elif col=="Indicadores":
                cell.fill=_f(row_bg); cell.font=_fn(size=8,color=row_fg)
                cell.alignment=_al("left")

            else:
                cell.fill=_f(row_bg); cell.font=_fn(size=9,color=row_fg); cell.alignment=_al("center")

        nlines=str(row.get("O que fazer","")).count("\n")+1
        ws.row_dimensions[dr].height=max(18,14*nlines); dr+=1

    WD={"Nome":32,"Microárea":11,"Bairro":16,"Endereco":34,"Telefone":15,"Indicadores":26,
        "Qtd":6,"PendÃªncias":10,"MÃ©dia":8,"Prioridade":18,"O que fazer":55}
    for ci,col in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(ci)].width=WD.get(col,13)
    ws.freeze_panes="D6"
    ws.auto_filter.ref=f"A5:{get_column_letter(n)}5"

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ABA 2: Urgentes
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    ws2=wb.create_sheet("ðŸ”´ Urgentes")
    df2=(df[df["PendÃªncias"]>0].sort_values(["PendÃªncias","MÃ©dia"],ascending=[False,True]).head(60))
    h2=["Nome","Microárea","Bairro","Endereco","Telefone","Indicadores","Qtd","PendÃªncias","Prioridade","O que fazer"]
    h2=[c for c in h2 if c in df2.columns]; n2=len(h2)

    _title(ws2,1,"ðŸ”´  PACIENTES MAIS URGENTES  â€”  maior pendÃªncia, menor pontuaÃ§Ã£o","C00000",C["branco"],14,n2,30)
    _title(ws2,2,
           f"Top {len(df2)} pacientes crÃ­ticos  â€¢  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
           "FFEBEE","C00000",9,n2,14)
    c3s=_mw(ws2,3,1,"  â–¼  Priorize: vermelho = urgÃªncia imediata  |  cada linha mostra o que precisa ser feito",ec=n2)
    c3s.font=_fn(bold=True,color=C["branco"],size=9); c3s.fill=_f("C00000")
    c3s.alignment=_al("left"); c3s.border=_bd(); ws2.row_dimensions[3].height=15

    for ci,h in enumerate(h2,1):
        bg="2C0F0F" if h in ("Nome","Microárea","Bairro","Endereco","Telefone") else "C00000"
        cell=ws2.cell(4,ci,h)
        cell.fill=_f(bg); cell.font=_fn(bold=True,color=C["branco"],size=9)
        cell.alignment=_al("center",wrap=True); cell.border=_bd()
    ws2.row_dimensions[4].height=32

    prev2=None; dr2=5
    for _,row in df2[h2].iterrows():
        prio=str(row.get("Prioridade",""))
        if prio!=prev2 and prio in PRIO_THEME:
            bg_s,fg_s=PRIO_THEME[prio]
            try: ws2.merge_cells(f"A{dr2}:{get_column_letter(n2)}{dr2}")
            except: pass
            s=ws2.cell(dr2,1,f"  {prio}")
            s.fill=_f(bg_s); s.font=_fn(bold=True,color=fg_s,size=9)
            s.alignment=_al("left"); s.border=_bd("medium","888888"); ws2.row_dimensions[dr2].height=14; dr2+=1; prev2=prio
        bg,fg=PRIO_THEME.get(prio,(C["cinza"],C["azul"]))
        for ci,h in enumerate(h2,1):
            vs=str(row.get(h,"") or ""); vs="" if vs in {"nan","None"} else vs
            cell=ws2.cell(dr2,ci,vs); cell.border=_bd()
            if h=="Prioridade":
                bg2,fg2=PRIO_THEME.get(vs,(C["cinza"],C["azul"]))
                cell.fill=_f(bg2); cell.font=_fn(bold=True,color=fg2,size=9); cell.alignment=_al("center")
            elif h=="O que fazer":
                cell.fill=_f(bg); cell.font=_fn(size=9,color=fg); cell.alignment=_al("left",wrap=True)
            elif h=="Nome":
                cell.fill=_f(bg); cell.font=_fn(bold=True,color=fg,size=9); cell.alignment=_al("left")
            else:
                cell.fill=_f(bg); cell.font=_fn(size=9,color=fg)
                cell.alignment=_al("center" if h not in ("Indicadores",) else "left")
        nlines=str(row.get("O que fazer","")).count("\n")+1
        ws2.row_dimensions[dr2].height=max(18,14*nlines); dr2+=1

    for ci,w in enumerate([32,11,15,26,6,10,18,55],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes="A5"

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ABA 3: Resumo / mini-dashboard
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    ws3=wb.create_sheet("ðŸ“Š Resumo")
    _title(ws3,1,"ðŸ“Š  RESUMO POR INDICADOR  â€”  visÃ£o consolidada",C["azul"],C["branco"],14,8,30)
    _title(ws3,2,
           f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  â€¢  "
           f"{len(ind_cols)} indicadores  â€¢  {total_pac} pacientes Ãºnicos",
           C["azul_c"],C["azul"],9,8,16)

    # KPI cards
    _section(ws3,3,"  â–¼  VISÃƒO GERAL",C["azul_m"],C["branco"],8)
    def _kpi3(col,lbl,val,bg,fg):
        for r in (4,5):
            try: ws3.merge_cells(f"{get_column_letter(col)}{r}:{get_column_letter(col+1)}{r}")
            except: pass
        cl=ws3.cell(4,col,lbl); cl.fill=_f(bg); cl.font=_fn(bold=True,color=fg,size=8)
        cl.alignment=_al("center"); cl.border=_bd()
        cv=ws3.cell(5,col,val); cv.fill=_f(bg); cv.font=_fn(bold=True,color=fg,size=18)
        cv.alignment=_al("center"); cv.border=_bd()
        ws3.row_dimensions[4].height=14; ws3.row_dimensions[5].height=28
    _kpi3(1,"ðŸ‘¥ Pacientes",total_pac,C["azul_h"],C["azul"])
    _kpi3(3,"ðŸ”´ Urgentes",urgentes,C["verm_c"],C["verm_e"])
    _kpi3(5,"âœ” Em dia",em_dia,C["verde_c"],C["verde_e"])
    _kpi3(7,"âš  2+ listas",multi,C["amarelo"],C["amar_t"])

    # Tabela por indicador
    _section(ws3,7,"  â–¼  DETALHE POR INDICADOR  â€”  % de conclusÃ£o, pendÃªncias e barra de progresso",C["roxo"],C["branco"],8)
    hdr3=["Indicador","Total","ConcluÃ­dos","Pendentes","MÃ©dia pts","% ConcluÃ­dos","Progresso","Status"]
    for ci,h in enumerate(hdr3,1):
        cell=ws3.cell(8,ci,h)
        cell.fill=_f(C["azul"]); cell.font=_fn(bold=True,color=C["branco"],size=9)
        cell.alignment=_al("center",wrap=True); cell.border=_bd()
    ws3.row_dimensions[8].height=30

    for ri,code in enumerate(ind_cols,9):
        vals=[str(v) for v in df[code] if str(v) not in ("â€”","nan","None","")]
        total=len(vals)
        conc=sum(1 for v in vals if "100" in v or "âœ”" in v)
        pend=total-conc
        try: media=round(sum(float(re.sub(r"[^\d.]","",v)) for v in vals)/total,1) if total else 0
        except: media=0
        pct=round(conc/total*100,1) if total else 0
        filled=int(pct/10); bar="â–ˆ"*filled+"â–‘"*(10-filled)
        if pct>=75:   bg=C["verde"];   fg=C["verde_e"];  st="ðŸŸ¢ Bom"
        elif pct>=50: bg=C["amarelo"]; fg=C["amar_t"];   st="ðŸŸ¡ AtenÃ§Ã£o"
        else:         bg=C["verm"];    fg=C["verm_t"];    st="ðŸ”´ CrÃ­tico"

        data_row=[code,total,conc,pend,media,f"{pct}%",bar,st]
        for ci,v in enumerate(data_row,1):
            cell=ws3.cell(ri,ci,v); cell.border=_bd()
            if ci==1:
                cell.fill=_f(C["azul_h"]); cell.font=_fn(bold=True,color=C["azul"],size=10)
                cell.alignment=_al("center")
            elif ci==3:
                cell.fill=_f(C["verde_c"]); cell.font=_fn(bold=True,color=C["verde_e"],size=10); cell.alignment=_al("center")
            elif ci==4:
                clr=C["verm_c"] if pend>0 else C["verde_c"]
                fcl=C["verm_e"] if pend>0 else C["verde_e"]
                cell.fill=_f(clr); cell.font=_fn(bold=True,color=fcl,size=10); cell.alignment=_al("center")
            elif ci==7:
                cell.fill=_f(bg); cell.font=XFont(name="Courier New",bold=True,color=fg,size=10)
                cell.alignment=_al("left")
            elif ci==8:
                cell.fill=_f(bg); cell.font=_fn(bold=True,color=fg,size=9); cell.alignment=_al("center")
            else:
                cell.fill=_f(C["cinza"] if ri%2==0 else C["branco"])
                cell.font=_fn(size=9); cell.alignment=_al("center")

    for ci,w in enumerate([12,10,13,12,10,13,22,12],1):
        ws3.column_dimensions[get_column_letter(ci)].width=w
    ws3.freeze_panes="A9"

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ABA 4: Em Dia
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    ws4=wb.create_sheet("ðŸŸ¢ Em Dia")
    df4=df[df["PendÃªncias"]==0].sort_values("Nome")
    h4=["Nome","Microárea","Bairro","Endereco","Telefone","Indicadores","Qtd","MÃ©dia"]
    h4=[c for c in h4 if c in df4.columns]; n4=len(h4)

    _title(ws4,1,"ðŸŸ¢  PACIENTES EM DIA  â€”  sem pendÃªncias em nenhuma lista",C["verde_e"],C["branco"],13,n4,28)
    _title(ws4,2,f"{len(df4)} pacientes concluÃ­dos  â€¢  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
           C["verde_c"],C["verde_e"],9,n4,14)
    c3v=_mw(ws4,3,1,"  â–¼  Estes pacientes atingiram 100 pts em todas as listas em que estÃ£o",ec=n4)
    c3v.font=_fn(bold=True,color=C["branco"],size=9); c3v.fill=_f(C["verde_e"])
    c3v.alignment=_al("left"); c3v.border=_bd(); ws4.row_dimensions[3].height=15

    for ci,h in enumerate(h4,1):
        cell=ws4.cell(4,ci,h)
        cell.fill=_f(C["verde_e"]); cell.font=_fn(bold=True,color=C["branco"],size=9)
        cell.alignment=_al("center",wrap=True); cell.border=_bd()
    ws4.row_dimensions[4].height=28

    for ri,(_,row) in enumerate(df4[h4].iterrows(),5):
        for ci,h in enumerate(h4,1):
            vs=str(row.get(h,"") or ""); vs="" if vs in {"nan","None"} else vs
            cell=ws4.cell(ri,ci,vs); cell.border=_bd()
            cell.fill=_f(C["verde_c"] if ri%2==0 else C["branco"])
            cell.font=_fn(bold=(h=="Nome"),size=9,color=C["verde_e"])
            cell.alignment=_al("left" if h in ("Nome","Indicadores") else "center")
        ws4.row_dimensions[ri].height=16

    for ci,w in enumerate([32,11,15,26,6,8],1):
        ws4.column_dimensions[get_column_letter(ci)].width=w
    ws4.freeze_panes="A5"

    wb.save(out_path)


# â”€â”€ ComparaÃ§Ã£o entre pastas (usado pelo dashboard) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _collect_folder(folder: Path) -> dict:
    data = {}
    by_code: dict[str, list[Path]] = {}
    for p in folder.glob("*.xlsx"):
        if any(x in p.name.lower() for x in ("interativa","unificad","comparad","cruzament")):
            continue
        code = infer_indicator_code_from_path(p) or _detect_code(p)
        by_code.setdefault(code, []).append(p)
    for code, paths in by_code.items():
        latest = max(paths, key=lambda p: p.stat().st_mtime)
        df = _read(latest)
        if df is not None and "Nome" in df.columns:
            df["_src"] = latest.name
            data[code] = df
    return data

def build_folder_comparison(folder_a: Path, folder_b: Path,
                            label_a="Periodo A", label_b="Periodo B") -> pd.DataFrame:
    da = _collect_folder(folder_a)
    db = _collect_folder(folder_b)
    codes = sorted(set(list(da.keys()) + list(db.keys())))
    rows = []
    for code in codes:
        def _s(df):
            if df is None:
                return None

            def _pick_series(*aliases, numeric: bool = False):
                norm_map = {_norm(c): c for c in df.columns}
                for alias in aliases:
                    col = norm_map.get(_norm(alias))
                    if col is not None:
                        out = df[col]
                        break
                else:
                    out = pd.Series(index=df.index, dtype=float if numeric else str)
                if numeric:
                    return pd.to_numeric(out, errors="coerce").fillna(0)
                return out.fillna("").astype(str)

            pts = _pick_series("Pontuação", "Pontuacao", "PontuaÃ§Ã£o", numeric=True)
            cls = _pick_series("Classificação", "Classificacao", "ClassificaÃ§Ã£o", numeric=False)
            return {
                "total": len(df), "media": round(float(pts.mean()),1) if len(df) else 0.0,
                "busca": int((pts<100).sum()),
                "otimo": int((cls=="Ã“timo").sum()), "bom": int((cls=="Bom").sum()),
                "suf":   int((cls=="Suficiente").sum()), "reg": int((cls=="Regular").sum()),
                "arquivo": df["_src"].iloc[0] if "_src" in df.columns and len(df) else "â€”",
            }
        sa = _s(da.get(code)); sb = _s(db.get(code))
        def _d(k):
            if sa and sb: return sb[k]-sa[k]
            return None
        rows.append({
            "Indicador":           code,
            f"Arquivo {label_a}":  sa["arquivo"] if sa else "â€”",
            f"Arquivo {label_b}":  sb["arquivo"] if sb else "â€”",
            f"Total {label_a}":    sa["total"]   if sa else "â€”",
            f"Total {label_b}":    sb["total"]   if sb else "â€”",
            f"Media {label_a}":    sa["media"]   if sa else "â€”",
            f"Media {label_b}":    sb["media"]   if sb else "â€”",
            "Variacao Media":      _d("media"),
            f"Busca {label_a}":    sa["busca"]   if sa else "â€”",
            f"Busca {label_b}":    sb["busca"]   if sb else "â€”",
            "Variacao Busca":      _d("busca"),
            "Variacao Otimo":      _d("otimo"),
            "Variacao Bom":        _d("bom"),
            "Variacao Suficiente": _d("suf"),
            "Variacao Regular":    _d("reg"),
        })
    return pd.DataFrame(rows)

def export_folder_comparison_excel(df: pd.DataFrame, out_path: Path,
                                   label_a="Periodo A", label_b="Periodo B"):
    from datetime import datetime
    wb = Workbook()
    ws = wb.active; ws.title = "Comparacao Pastas"
    n = len(df.columns)
    _title(ws,1,f"APS - COMPARACAO ENTRE PASTAS   {label_a} x {label_b}",
           C["azul"],C["branco"],13,n,28)
    _title(ws,2,
           f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  -  "
           f"Verde = melhora  -  Vermelho = piora  -  Variacao = diferenca B-A",
           C["azul_c"],C["azul"],9,n,14)
    for ci,col in enumerate(df.columns,1):
        if col.startswith("Variacao"): bg=C["roxo"]; fg=C["branco"]
        elif label_a in col: bg=C["azul_h"]; fg=C["azul"]
        elif label_b in col: bg=C["amar_m"]; fg=C["azul"]
        else: bg=C["azul_h"]; fg=C["azul"]
        cell=ws.cell(3,ci,col)
        cell.fill=_f(bg); cell.font=_fn(bold=True,color=fg,size=9)
        cell.alignment=_al("center",wrap=True); cell.border=_bd()
    ws.row_dimensions[3].height=40
    for ri,(_,row) in enumerate(df.iterrows(),4):
        for ci,col in enumerate(df.columns,1):
            val=row[col]
            vs="" if str(val) in {"nan","None"} else str(val)
            cell=ws.cell(ri,ci,val if str(val) not in {"nan","None"} else "")
            cell.border=_bd(); cell.alignment=_al("center"); cell.font=_fn(size=9)
            cell.fill=_f(C["cinza"] if ri%2==0 else C["branco"])
            if col.startswith("Variacao") and val not in (None,"â€”",""):
                try:
                    v=float(str(val).replace(",","."))
                    invert=any(x in col for x in ("Busca","Regular","Suf"))
                    good=(v<0) if invert else (v>0)
                    bad=(v>0)  if invert else (v<0)
                    if good:   cell.fill=_f(C["verde"]);  cell.font=_fn(bold=True,color=C["verde_e"],size=9)
                    elif bad:  cell.fill=_f(C["verm"]);   cell.font=_fn(bold=True,color=C["verm_t"],size=9)
                    else:      cell.fill=_f(C["amarelo"]); cell.font=_fn(bold=True,color=C["amar_t"],size=9)
                    cell.value=f"{v:+.1f}" if isinstance(v,float) else f"{int(v):+d}"
                except: pass
    for ci,col in enumerate(df.columns,1):
        ws.column_dimensions[get_column_letter(ci)].width=(32 if "Arquivo" in col else 10 if col=="Indicador" else 14)
    ws.freeze_panes="B4"
    ws.auto_filter.ref=f"A3:{get_column_letter(n)}3"
    wb.save(out_path)

# â”€â”€ UI â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
class ComparadorPacienteApp(tk.Toplevel):
    def __init__(self,master=None,out_dir=None):
        super().__init__(master=master)
        self.title("APS â€“ Cruzamento por Paciente")
        self.geometry("860x480")
        self.minsize(700,380)
        self.configure(bg="#EAF2F8")
        self.out_dir=out_dir or Path.home()/"Desktop"/"APS_RESULTADOS"
        self.files=[]
        self._build()
        self.lift(); self.focus_force(); self.grab_set()

    def _build(self):
        tk.Label(self,text="APS â€“ Cruzamento por Paciente",
                 bg="#1F4E79",fg="white",font=("Segoe UI",13,"bold"),pady=10).pack(fill="x")
        tk.Label(self,
                 text="Selecione as planilhas resultado (C1â€“C7). O sistema cruza por nome,\n"
                      "mostra em quais listas cada paciente estÃ¡ e o que estÃ¡ pendente para ele.",
                 bg="#EAF2F8",font=("Segoe UI",9),justify="left").pack(fill="x",padx=14,pady=(8,2))

        btn=tk.Frame(self,bg="#EAF2F8"); btn.pack(fill="x",padx=14,pady=6)
        for txt,cmd,bg in [
            ("âž• Adicionar planilhas",        self._add,   "#2E75B6"),
            ("ðŸ“ Auto-detectar na pasta",     self._auto,  "#27AE60"),
            ("ðŸ—‘ Limpar lista",               self._clear, "#C0392B"),
        ]:
            tk.Button(btn,text=txt,command=cmd,bg=bg,fg="white",
                      font=("Segoe UI",9,"bold")).pack(side="left",padx=(0,6))

        box=ttk.LabelFrame(self,text="Planilhas selecionadas")
        box.pack(fill="both",expand=True,padx=14,pady=4)
        box.columnconfigure(0,weight=1); box.rowconfigure(0,weight=1)
        self.tree=ttk.Treeview(box,columns=("cod","arq","mb"),show="headings",height=8)
        self.tree.heading("cod",text="CÃ³d"); self.tree.heading("arq",text="Arquivo"); self.tree.heading("mb",text="MB")
        self.tree.column("cod",width=60,anchor="center"); self.tree.column("arq",width=560); self.tree.column("mb",width=70,anchor="e")
        self.tree.grid(row=0,column=0,sticky="nsew")
        sb=ttk.Scrollbar(box,orient="vertical",command=self.tree.yview)
        sb.grid(row=0,column=1,sticky="ns"); self.tree.configure(yscrollcommand=sb.set)

        bot=tk.Frame(self,bg="#EAF2F8"); bot.pack(fill="x",padx=14,pady=8)
        self.sv=tk.StringVar(value="Nenhuma planilha selecionada.")
        tk.Label(bot,textvariable=self.sv,bg="#EAF2F8",font=("Segoe UI",9),anchor="w").pack(side="left",fill="x",expand=True)
        tk.Button(bot,text="â–¶  Gerar planilha unificada",command=self._run,
                  bg="#1F4E79",fg="white",font=("Segoe UI",10,"bold")).pack(side="right")

    def _add(self):
        chosen=filedialog.askopenfilenames(title="Planilhas resultado APS",filetypes=[("Excel","*.xlsx")])
        for f in chosen:
            p=Path(f)
            if p not in self.files: self.files.append(p)
        self._refresh()

    def _auto(self):
        folder=filedialog.askdirectory(title="Pasta com resultados APS",initialdir=str(self.out_dir))
        if not folder: return
        found=[p for p in sorted(Path(folder).iterdir())
               if p.suffix.lower()==".xlsx"
               and not any(x in p.name.lower() for x in ("interativa","unificad","comparad","cruzament"))
               and re.search(r"c\d",p.name.lower())]
        if not found: messagebox.showwarning("Nada encontrado","Nenhuma planilha C1â€“C7 na pasta."); return
        for p in found:
            if p not in self.files: self.files.append(p)
        self._refresh()

    def _clear(self): self.files.clear(); self._refresh()

    def _refresh(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        for p in self.files:
            self.tree.insert("","end",values=(_detect_code(p),p.name,f"{p.stat().st_size/1048576:.2f}"))
        n=len(self.files)
        self.sv.set(f"{n} planilha(s) selecionada(s)." if n else "Nenhuma planilha selecionada.")

    def _run(self):
        if not self.files: messagebox.showwarning("AtenÃ§Ã£o","Adicione ao menos uma planilha."); return
        self.sv.set("Processandoâ€¦ aguarde."); self.update()
        try:
            df=build_unified(self.files)
            if df.empty:
                messagebox.showwarning("Resultado vazio",
                    "NÃ£o foi possÃ­vel cruzar. Verifique se as planilhas tÃªm a aba 'ðŸ“‹ Dados'."); return
            stamp=datetime.now().strftime("%Y%m%d_%H%M%S")
            out=self.out_dir/f"CRUZAMENTO_{stamp}.xlsx"
            out.parent.mkdir(parents=True,exist_ok=True)
            export_excel(df,out)
            self.sv.set(f"âœ” {out.name}")
            messagebox.showinfo("ConcluÃ­do âœ”",f"Planilha gerada:\n\n{out}\n\nPacientes: {len(df)}",parent=self)
            try: os.startfile(out)
            except: pass
        except Exception as exc:
            import traceback
            self.sv.set(f"Erro: {exc}")
            messagebox.showerror("Erro",f"{exc}\n\n{traceback.format_exc()}",parent=self)

def launch_comparador(master=None,out_dir=None):
    return ComparadorPacienteApp(master=master,out_dir=out_dir)

def main():
    root=tk.Tk(); root.withdraw()
    app=ComparadorPacienteApp(master=root)
    app.protocol("WM_DELETE_WINDOW",root.destroy)
    root.mainloop()

if __name__=="__main__": main()


