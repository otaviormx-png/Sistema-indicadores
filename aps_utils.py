
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import aps_config

ENCODINGS = ["utf-8-sig", "utf-8", "latin1", "cp1252"]
CSV_SEPARATORS = [";", ",", "\t"]

# Cores carregadas do config.toml (com fallback embutido em aps_config.py)
COR = aps_config.cores()

BASE_WIDTHS = {
    1: 35, 2: 16, 3: 22, 4: 12, 5: 12,
    6: 10, 7: 30, 8: 8, 9: 15, 10: 20,
    11: 16, 12: 16, 13: 16, 14: 18,
    15: 12, 16: 12, 17: 12, 18: 12,
    19: 30, 20: 22,
}

BUSCA_WIDTHS = {1: 12, 2: 35, 3: 10, 4: 18, 5: 14, 6: 14, 7: 16, 8: 16, 9: 12, 10: 14, 11: 30, 12: 25}
RESUMO_WIDTHS = {1: 18, 2: 12, 3: 20, 4: 14, 5: 16, 6: 14, 7: 25, 8: 12}
ESTAT_WIDTHS = {1: 18, 2: 35, 3: 10, 4: 14, 5: 14, 6: 16, 7: 14, 8: 14, 9: 12, 10: 14}

BASE_PERSON_COLUMNS = [
    "Nome", "Data de nascimento", "Idade", "Sexo", "Raça/cor",
    "Microárea", "Rua", "Número", "Complemento", "Bairro",
    "Telefone celular", "Telefone residencial", "Telefone de contato",
    "CPF", "CNS",
]

BASE_CLINICAL_COLUMNS = [
    "Meses desde o último atendimento médico",
    "Meses desde o último atendimento de enfermagem",
    "Meses desde o último atendimento odontológico",
    "Meses desde a última visita domiciliar",
    "Última medição de peso",
    "Última medição de altura",
    "Data da ultima medição de peso e altura",
    "Última medição de pressão arterial",
    "Data da última medição de pressão arterial",
    "Últimas visitas domiciliares",
    "Quantidade de visitas domiciliares",
]

# Campos derivados na saída para facilitar leitura operacional:
# convertem recência (dias/meses desde atendimento) em data estimada.
BASE_CLINICAL_COLUMNS += [
    "Data estimada do ultimo atendimento medico",
    "Data estimada do ultimo atendimento de enfermagem",
    "Data estimada do ultimo atendimento odontologico",
    "Data estimada da ultima visita domiciliar",
]

COMMON_CANDIDATES = {
    "Nome": ["Nome", "Nome do cidadão", "Paciente", "Usuário", "Cidadão"],
    "Data de nascimento": ["Data de nascimento", "Nascimento", "Dt nascimento"],
    "Idade": ["Idade"],
    "Sexo": ["Sexo"],
    "Raça/cor": ["Raça/cor", "Raca/cor", "Raça", "Raca"],
    "Microárea": ["Microárea", "Microarea"],
    "Rua": ["Rua", "Logradouro"],
    "Número": ["Número", "Numero"],
    "Complemento": ["Complemento"],
    "Bairro": ["Bairro"],
    "Telefone celular": ["Telefone celular", "Celular", "Telefone"],
    "Telefone residencial": ["Telefone residencial"],
    "Telefone de contato": ["Telefone de contato", "Contato"],
    "CPF": ["CPF"],
    "CNS": ["CNS", "Cartão SUS"],
    "Meses desde o último atendimento médico": ["Meses desde o último atendimento médico"],
    "Meses desde o último atendimento de enfermagem": ["Meses desde o último atendimento de enfermagem"],
    "Meses desde o último atendimento odontológico": ["Meses desde o último atendimento odontológico"],
    "Meses desde a última visita domiciliar": ["Meses desde a última visita domiciliar"],
    "Última medição de peso": ["Última medição de peso"],
    "Última medição de altura": ["Última medição de altura"],
    "Data da ultima medição de peso e altura": ["Data da ultima medição de peso e altura", "Data da última medição de peso e altura"],
    "Última medição de pressão arterial": ["Última medição de pressão arterial"],
    "Data da última medição de pressão arterial": ["Data da última medição de pressão arterial"],
    "Últimas visitas domiciliares": ["Últimas visitas domiciliares"],
    "Quantidade de visitas domiciliares": ["Quantidade de visitas domiciliares"],
}

COMMON_CANDIDATES.update(
    {
        "Data estimada do ultimo atendimento medico": [
            "Data estimada do ultimo atendimento medico",
            "Data estimada do último atendimento médico",
        ],
        "Data estimada do ultimo atendimento de enfermagem": [
            "Data estimada do ultimo atendimento de enfermagem",
            "Data estimada do último atendimento de enfermagem",
        ],
        "Data estimada do ultimo atendimento odontologico": [
            "Data estimada do ultimo atendimento odontologico",
            "Data estimada do último atendimento odontológico",
        ],
        "Data estimada da ultima visita domiciliar": [
            "Data estimada da ultima visita domiciliar",
            "Data estimada da última visita domiciliar",
        ],
    }
)

@dataclass
class IndicatorConfig:
    code: str
    titulo: str
    criterio_bloco: str
    subtitulo: str
    theme_keywords: list[str]
    criteria: list[dict]
    extra_columns: list[str]
    builder: Callable[[pd.DataFrame], pd.DataFrame]
    official_like: bool = True


INDICATOR_OUTPUT_NAMES: dict[str, str] = {
    "C1": "Mais acesso",
    "C2": "Cuidado no desenvolvimento infantil",
    "C3": "Cuidado na gestação e puerpério",
    "C4": "Cuidado da pessoa com diabetes",
    "C5": "Cuidado da pessoa com hipertensão",
    "C6": "Cuidado da pessoa idosa",
    "C7": "Cuidado da mulher na prevenção do câncer",
}

_INDICATOR_CODE_ALIASES: dict[str, tuple[str, ...]] = {
    "C1": ("mais acesso",),
    "C2": ("desenvolvimento infantil", "cuidado no desenvolvimento infantil", "infantil"),
    "C3": ("gestacao", "gesta", "puerperio", "cuidado na gestacao e puerperio"),
    "C4": ("diabetes", "cuidado da pessoa com diabetes"),
    "C5": ("hipertensao", "hiperten", "cuidado da pessoa com hipertensao"),
    "C6": ("idosa", "idoso", "cuidado da pessoa idosa"),
    "C7": ("mulher", "cancer", "prevencao", "prevencao do cancer", "cuidado da mulher na prevencao do cancer"),
}


def _norm_name_key(txt) -> str:
    base = normalize_text(txt)
    base = re.sub(r"[^a-z0-9]+", " ", base).strip()
    return base


def indicator_output_filename(code: str, suffix: str = ".xlsx") -> str:
    code_up = str(code or "").upper().strip()
    label = INDICATOR_OUTPUT_NAMES.get(code_up, code_up or "resultado")
    return f"{label}{suffix}"


def infer_indicator_code_from_path(path: str | Path) -> str | None:
    stem = Path(path).stem
    m = re.search(r"(C\d+)", stem, flags=re.I)
    if m:
        return m.group(1).upper()
    name_key = _norm_name_key(stem)
    if not name_key:
        return None
    for code, aliases in _INDICATOR_CODE_ALIASES.items():
        for alias in aliases:
            alias_key = _norm_name_key(alias)
            if alias_key and alias_key in name_key:
                return code
    return None


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def font(bold: bool = False, color: str = "000000", size: int = 10, italic: bool = False, name: str = "Arial") -> Font:
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)


def border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def border_medium() -> Border:
    s = Side(style="medium", color="595959")
    return Border(left=s, right=s, top=s, bottom=s)


def align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def normalize_text(txt) -> str:
    if txt is None:
        return ""
    txt = str(txt).strip()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    return txt.lower()


def detect_header_row(path: str | Path, encoding: str) -> int:
    with open(path, "r", encoding=encoding, errors="ignore") as f:
        for idx, line in enumerate(f):
            norm = normalize_text(line)
            if norm.startswith("nome;data de nascimento;") or norm.startswith("nome,data de nascimento,"):
                return idx
            if "nome" in norm and "data de nascimento" in norm and "cpf" in norm and "cns" in norm:
                return idx
    return 0


def detect_theme_label(path: str | Path, encoding: str = "latin1") -> str:
    try:
        with open(path, "r", encoding=encoding, errors="ignore") as f:
            for line in f:
                if "Lista temática" in line or "Lista tem" in line:
                    parts = [p.strip() for p in line.split(";")]
                    if len(parts) >= 2:
                        return parts[1]
    except Exception:
        pass
    return ""


def read_esus_table(path: str | Path) -> pd.DataFrame:
    path = str(path)
    if path.lower().endswith((".xlsx", ".xlsm", ".xls")):
        xls = pd.ExcelFile(path)
        for sheet in xls.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str)
            if df.shape[1] >= 4 and any(normalize_text(c) == "nome" for c in df.columns):
                return clean_columns(df)
        return clean_columns(pd.read_excel(path, dtype=str))

    last_error = None
    for enc in ENCODINGS:
        try:
            header = detect_header_row(path, enc)
        except Exception as exc:
            last_error = exc
            continue
        for sep in CSV_SEPARATORS:
            try:
                df = pd.read_csv(
                    path,
                    sep=sep,
                    encoding=enc,
                    skiprows=header,
                    dtype=str,
                    engine="python",
                    on_bad_lines="skip",
                )
                df = clean_columns(df)
                if df.shape[1] >= 4 and any(normalize_text(c) == "nome" for c in df.columns):
                    return df
            except Exception as exc:
                last_error = exc
    if last_error:
        raise last_error
    raise ValueError("Não foi possível ler o arquivo bruto do e-SUS.")


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    keep = []
    for col in df.columns:
        name = str(col).strip()
        if name.startswith("Unnamed:"):
            continue
        keep.append(name)
    df = df[keep]
    df = df.dropna(how="all")
    return df


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    norm_map = {normalize_text(c): c for c in df.columns}
    for cand in candidates:
        c_norm = normalize_text(cand)
        if c_norm in norm_map:
            return norm_map[c_norm]
    for cand in candidates:
        c_norm = normalize_text(cand)
        for col in df.columns:
            col_norm = normalize_text(col)
            if c_norm in col_norm or col_norm in c_norm:
                return col
    return None


def value(row: pd.Series | dict, *candidates: str, default=""):
    for candidate in candidates:
        if candidate in row:
            val = row.get(candidate)
            if pd.notna(val) and str(val).strip() not in {"", "-", "nan", "None"}:
                return val
    return default


def to_numeric(val, default=0.0):
    if val is None:
        return default
    s = str(val).strip()
    if s in {"", "-", "nan", "None"}:
        return default
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return default


def parse_date(val):
    if val is None or str(val).strip() in {"", "-", "nan", "None"}:
        return pd.NaT
    try:
        return pd.to_datetime(val, errors="coerce", dayfirst=True)
    except Exception:
        return pd.NaT


def months_leq(val, limit: int) -> bool:
    return to_numeric(val, 999) <= limit


def count_ge(val, limit: int) -> bool:
    return int(to_numeric(val, 0)) >= limit


def has_any_text(val) -> bool:
    return str(val).strip() not in {"", "-", "nan", "None"}


def age_years(val, default: int = -1) -> int:
    s = str(val).strip()
    if s in {"", "-", "nan", "None"}:
        return default
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else default


def value_norm(row: pd.Series | dict, *candidates: str, default=""):
    norm_map = {normalize_text(k): k for k in row.keys()}
    for candidate in candidates:
        key = norm_map.get(normalize_text(candidate))
        if key is None:
            continue
        val = row.get(key)
        if pd.notna(val) and str(val).strip() not in {"", "-", "nan", "None"}:
            return val
    return default


def is_team_type_76(row: pd.Series | dict) -> bool:
    team_raw = value_norm(
        row,
        "Tipo de equipe",
        "Tipo Equipe",
        "Tipo da equipe",
        "Codigo do tipo da equipe",
        "Codigo tipo equipe",
        "Equipe tipo",
        "Tipo eSF/eAP",
        default="",
    )
    txt = normalize_text(team_raw)
    if not txt:
        return False
    if re.search(r"(^|\D)76(\D|$)", txt):
        return True
    return "tipo 76" in txt


def within_last_months(val, limit_months: int, ref_date=None) -> bool:
    dt = parse_date(val)
    if pd.isna(dt):
        return False
    ref = pd.Timestamp(ref_date) if ref_date is not None else pd.Timestamp.now().normalize()
    diff_days = (ref - dt).days
    return 0 <= diff_days <= (limit_months * 31)


def has_recent_date_or_text(val, limit_months: int) -> bool:
    if within_last_months(val, limit_months):
        return True
    dt = parse_date(val)
    if pd.notna(dt):
        return False
    return has_any_text(val)


def _non_negative_number(val) -> float | None:
    n = to_numeric(val, default=None)
    if n is None:
        return None
    try:
        n = float(n)
    except Exception:
        return None
    if n < 0:
        return None
    return n


def _extract_month_day_from_text(val) -> tuple[float | None, float | None]:
    txt = normalize_text(val)
    if not txt or txt in {"-", "nan", "none"}:
        return None, None
    months = None
    days = None
    month_matches = re.findall(r"(\d+(?:[.,]\d+)?)\s*mes(?:es)?", txt)
    day_matches = re.findall(r"(\d+(?:[.,]\d+)?)\s*dia(?:s)?", txt)
    if month_matches:
        months = sum(float(m.replace(",", ".")) for m in month_matches)
    if day_matches:
        days = sum(float(d.replace(",", ".")) for d in day_matches)
    return months, days


def estimated_last_date_from_recency(days_value, months_value, ref_date=None) -> str:
    ref = pd.Timestamp(ref_date) if ref_date is not None else pd.Timestamp.now().normalize()
    days_n = _non_negative_number(days_value)
    months_n = _non_negative_number(months_value)
    m_txt1, d_txt1 = _extract_month_day_from_text(months_value)
    m_txt2, d_txt2 = _extract_month_day_from_text(days_value)

    has_month = months_n is not None or m_txt1 is not None or m_txt2 is not None
    has_day = days_n is not None or d_txt1 is not None or d_txt2 is not None
    if not has_month and not has_day:
        return ""

    if months_n is None:
        months_n = float((m_txt1 or 0) + (m_txt2 or 0))
    if days_n is None:
        days_n = float((d_txt1 or 0) + (d_txt2 or 0))

    dt = ref
    if has_month:
        whole_months = int(months_n or 0)
        frac_days = int(round(((months_n or 0) - whole_months) * 30))
        if whole_months:
            dt = dt - pd.DateOffset(months=whole_months)
        if frac_days:
            dt = dt - pd.Timedelta(days=frac_days)
    if has_day and int(round(days_n or 0)) > 0:
        dt = dt - pd.Timedelta(days=int(round(days_n or 0)))
    return dt.strftime("%d/%m/%Y")


def build_base_row(row: pd.Series, extra_columns: list[str]) -> dict:
    """
    Monta um dicionário de saída com as colunas padrão + extras,
    copiando valores da Series `row` via busca normalizada nos índices.

    Correção: a versão anterior criava um pd.DataFrame vazio apenas para
    reutilizar find_column(), o que era desnecessário e gerava overhead.
    Agora a busca é feita diretamente sobre row.index.
    """
    # Pré-computa mapa normalizado dos índices reais da linha
    norm_index: dict[str, str] = {normalize_text(c): c for c in row.index}

    def _find_in_row(candidates: list[str]) -> str | None:
        """Retorna o nome real da coluna em row que corresponde a um candidato."""
        for cand in candidates:
            cand_n = normalize_text(cand)
            # Correspondência exata normalizada
            if cand_n in norm_index:
                return norm_index[cand_n]
        for cand in candidates:
            cand_n = normalize_text(cand)
            # Correspondência parcial (substring em qualquer direção)
            for col_n, col_real in norm_index.items():
                if cand_n in col_n or col_n in cand_n:
                    return col_real
        return None

    built: dict = {}
    for out_col in BASE_PERSON_COLUMNS + BASE_CLINICAL_COLUMNS + extra_columns:
        candidates = COMMON_CANDIDATES.get(out_col, [out_col])
        found = _find_in_row(candidates)
        built[out_col] = row.get(found, "") if found else ""

    def _row_value(*candidates: str):
        found = _find_in_row(list(candidates))
        return row.get(found, "") if found else ""

    def _row_value_strict(*candidates: str):
        for cand in candidates:
            found = norm_index.get(normalize_text(cand))
            if found is not None:
                return row.get(found, "")
        return ""

    built["Data estimada do ultimo atendimento medico"] = estimated_last_date_from_recency(
        _row_value_strict("Dias desde o ultimo atendimento medico", "Dias desde o último atendimento médico"),
        _row_value_strict("Meses desde o ultimo atendimento medico", "Meses desde o último atendimento médico"),
    )
    built["Data estimada do ultimo atendimento de enfermagem"] = estimated_last_date_from_recency(
        _row_value_strict("Dias desde o ultimo atendimento de enfermagem", "Dias desde o último atendimento de enfermagem"),
        _row_value_strict("Meses desde o ultimo atendimento de enfermagem", "Meses desde o último atendimento de enfermagem"),
    )
    built["Data estimada do ultimo atendimento odontologico"] = estimated_last_date_from_recency(
        _row_value_strict("Dias desde o ultimo atendimento odontologico", "Dias desde o último atendimento odontológico"),
        _row_value_strict("Meses desde o ultimo atendimento odontologico", "Meses desde o último atendimento odontológico"),
    )
    built["Data estimada da ultima visita domiciliar"] = estimated_last_date_from_recency(
        _row_value_strict("Dias desde a ultima visita domiciliar", "Dias desde a última visita domiciliar"),
        _row_value_strict("Meses desde a ultima visita domiciliar", "Meses desde a última visita domiciliar"),
    )
    return built


def infer_phone(base: dict) -> str:
    vals = [base.get("Telefone celular", ""), base.get("Telefone residencial", ""), base.get("Telefone de contato", "")]
    vals = [str(v).strip() for v in vals if str(v).strip() not in {"", "-", "nan", "None"}]
    return " | ".join(vals) if vals else "-"


def classify_score(score: float) -> tuple[str, str]:
    if score > 75:
        return "Ótimo", "Concluído"
    if score > 50:
        return "Bom", "Baixa"
    if score > 25:
        return "Suficiente", "Média"
    return "Regular", "Alta"


def priority_for_busca(score: float) -> tuple[str, str, str]:
    if score == 0:
        return "🔴 URGENTE", "Regular", "FFE7E7"
    if score <= 25:
        return "🟠 ALTA", "Regular", "FFF2CC"
    if score <= 50:
        return "🟡 MÉDIA", "Suficiente", "FFFDE7"
    return "🟢 BAIXA", "Bom", "F0F7EE"


def criteria_subtitle(criteria: list[dict]) -> str:
    parts = []
    for item in criteria:
        parts.append(f"{item['letter']}={item['label']} ({item['weight']} pts)")
    return "  |  ".join(parts)


def safe_sheet_name(name: str) -> str:
    bad = '[]:*?/\\'
    for ch in bad:
        name = name.replace(ch, " ")
    return name[:31]


def write_merged_title(ws, cell_range: str, value: str, bg: str, fg: str = "FFFFFF", size: int = 12):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = value
    c.font = font(bold=True, color=fg, size=size)
    c.fill = fill(bg)
    c.alignment = align("center", wrap=True)
    c.border = border_thin()


def score_fill_and_font(score: float):
    if score > 75:
        return fill(COR["verde_escuro"]), font(bold=True, color=COR["branco"], size=9)
    if score > 50:
        return fill(COR["verde_ok"]), font(bold=True, color=COR["verde_escuro"], size=9)
    if score > 25:
        return fill(COR["amarelo"]), font(bold=True, color=COR["amarelo_txt"], size=9)
    return fill(COR["vermelho"]), font(bold=True, color=COR["vermelho_txt"], size=9)


def class_fill_and_font(label: str):
    if label == "Ótimo":
        return fill(COR["verde_escuro"]), font(bold=True, color=COR["branco"], size=9)
    if label == "Bom":
        return fill(COR["verde_ok"]), font(bold=True, color=COR["verde_escuro"], size=9)
    if label == "Suficiente":
        return fill(COR["amarelo"]), font(bold=True, color=COR["amarelo_txt"], size=9)
    return fill(COR["vermelho"]), font(bold=True, color=COR["vermelho_txt"], size=9)


def render_workbook(df: pd.DataFrame, cfg: IndicatorConfig, output_path: str | Path):
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    create_data_sheet(wb, df, cfg)
    create_busca_sheet(wb, df, cfg)
    create_summary_sheet(wb, df, cfg)
    create_stats_sheet(wb, df, cfg)

    wb.save(output_path)


def create_data_sheet(wb: Workbook, df: pd.DataFrame, cfg: IndicatorConfig):
    ws = wb.create_sheet(f"📋 Dados {cfg.code}", 0)
    person_cols = [c for c in BASE_PERSON_COLUMNS if c in df.columns]
    clinical_cols = [c for c in BASE_CLINICAL_COLUMNS + cfg.extra_columns if c in df.columns]
    criteria_cols = [f"{c['letter']} - {c['label']}" for c in cfg.criteria]
    final_cols = ["Pontuação", "Classificação", "Prioridade", "Pendências"]

    left_end = get_column_letter(max(1, len(person_cols) + len(clinical_cols)))
    total_end = get_column_letter(len(person_cols) + len(clinical_cols) + len(criteria_cols) + len(final_cols))

    ws.merge_cells(f"A1:{left_end}1")
    c = ws["A1"]
    c.value = cfg.titulo
    c.font = font(bold=True, color=COR["branco"], size=13)
    c.fill = fill(COR["azul_escuro"])
    c.alignment = align("center")
    c.border = border_medium()
    ws.row_dimensions[1].height = 30

    if len(criteria_cols) > 0:
        ws.merge_cells(f"{get_column_letter(len(person_cols)+len(clinical_cols)+1)}1:{total_end}1")
        c2 = ws[f"{get_column_letter(len(person_cols)+len(clinical_cols)+1)}1"]
        c2.value = cfg.criterio_bloco
        c2.font = font(bold=True, color=COR["branco"], size=12)
        c2.fill = fill(COR["roxo"])
        c2.alignment = align("center")
        c2.border = border_medium()

    ws.merge_cells(f"A2:{total_end}2")
    c3 = ws["A2"]
    c3.value = cfg.subtitulo
    c3.font = font(italic=True, color=COR["azul_escuro"], size=9)
    c3.fill = fill(COR["azul_claro"])
    c3.alignment = align("center", wrap=True)
    c3.border = border_thin()
    ws.row_dimensions[2].height = 18

    headers = person_cols + clinical_cols + criteria_cols + final_cols
    for idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=3, column=idx, value=hdr)
        if idx <= len(person_cols):
            bg = COR["azul_header"]
        elif idx <= len(person_cols) + len(clinical_cols):
            bg = COR["azul_clinico"]
        elif idx <= len(person_cols) + len(clinical_cols) + len(criteria_cols):
            bg = COR["amarelo_header"]
        else:
            bg = COR["verde_header"]
        cell.fill = fill(bg)
        cell.font = font(bold=True, size=9, color="1F3864")
        cell.alignment = align("center", wrap=True)
        cell.border = border_thin()
    ws.row_dimensions[3].height = 40

    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        bg_row = COR["branco"] if row_idx % 2 == 0 else COR["cinza_claro"]
        values = [row.get(col, "") for col in headers]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill(bg_row)
            cell.font = font(size=9)
            cell.alignment = align("left", wrap=False)
            cell.border = border_thin()
            if isinstance(val, (pd.Timestamp, datetime)):
                cell.number_format = "DD/MM/YYYY"

        crit_start = len(person_cols) + len(clinical_cols) + 1
        for i, crit in enumerate(criteria_cols, crit_start):
            val = row.get(crit, "")
            cell = ws.cell(row=row_idx, column=i, value=val)
            val_up = str(val).strip().upper()
            if val_up == "SIM":
                cell.fill = fill(COR["verde_ok"])
                cell.font = font(bold=True, size=9, color=COR["verde_texto"])
            elif val_up in {"N/A", "NA"}:
                cell.fill = fill("E7E6E6")
                cell.font = font(bold=True, size=9, color="595959")
            else:
                cell.fill = fill(COR["vermelho"])
                cell.font = font(bold=True, size=9, color=COR["vermelho_txt"])
            cell.alignment = align("center")
            cell.border = border_thin()

        score_col = len(person_cols) + len(clinical_cols) + len(criteria_cols) + 1
        class_col = score_col + 1
        prio_col = class_col + 1

        pts = to_numeric(row.get("Pontuação"), 0)
        fill_pts, font_pts = score_fill_and_font(pts)
        ws.cell(row_idx, score_col).fill = fill_pts
        ws.cell(row_idx, score_col).font = font_pts
        ws.cell(row_idx, score_col).alignment = align("center")

        cls = str(row.get("Classificação", ""))
        fill_cls, font_cls = class_fill_and_font(cls)
        ws.cell(row_idx, class_col).fill = fill_cls
        ws.cell(row_idx, class_col).font = font_cls
        ws.cell(row_idx, class_col).alignment = align("center")

        pr = str(row.get("Prioridade", ""))
        pr_cell = ws.cell(row_idx, prio_col)
        if pr == "Concluído":
            pr_cell.fill = fill(COR["verde_ok"]); pr_cell.font = font(bold=True, size=9, color=COR["verde_escuro"])
        elif pr == "Baixa":
            pr_cell.fill = fill("E2F0D9"); pr_cell.font = font(bold=True, size=9, color=COR["verde_escuro"])
        elif pr == "Média":
            pr_cell.fill = fill(COR["amarelo"]); pr_cell.font = font(bold=True, size=9, color=COR["amarelo_txt"])
        else:
            pr_cell.fill = fill(COR["vermelho"]); pr_cell.font = font(bold=True, size=9, color=COR["vermelho_txt"])
        pr_cell.alignment = align("center")

    apply_widths(ws, headers, criteria_count=len(criteria_cols))
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{total_end}{ws.max_row}"


def create_busca_sheet(wb: Workbook, df: pd.DataFrame, cfg: IndicatorConfig):
    ws = wb.create_sheet("🔍 Busca Ativa", 1)
    criterio_headers = [f"{c['letter']} - {c['label']}" for c in cfg.criteria]
    busca = df[df["Pontuação"] < 100].copy().sort_values(["Pontuação", "Nome"], ascending=[True, True])

    last_col = 8 + len(criterio_headers)
    end_col = get_column_letter(last_col)
    write_merged_title(ws, f"A1:{end_col}1", f"BUSCA ATIVA – {cfg.code} | PACIENTES COM CRITÉRIOS PENDENTES", COR["azul_escuro"], size=13)
    write_merged_title(ws, f"A2:{end_col}2", "Lista priorizada dos pacientes que ainda não atingiram 100 pontos.", COR["azul_claro"], fg=COR["azul_escuro"], size=9)
    write_merged_title(ws, f"A3:{end_col}3", "  " + "    ".join([f"{c['letter']} = {c['label']}" for c in cfg.criteria]), COR["roxo"], size=9)

    headers = ["Prioridade", "Nome", "Microárea", "Telefone"] + criterio_headers + ["Pontuação", "Classificação", "Critérios Faltantes", "Observação"]
    for idx, hdr in enumerate(headers, 1):
        cell = ws.cell(4, idx, hdr)
        cell.fill = fill(COR["azul_header"] if idx <= 4 else COR["amarelo_header"] if idx <= 4 + len(criterio_headers) else COR["verde_header"])
        cell.font = font(bold=True, color="1F3864", size=9)
        cell.alignment = align("center", wrap=True)
        cell.border = border_thin()
    ws.row_dimensions[4].height = 32

    for i, (_, row) in enumerate(busca.iterrows(), 5):
        prioridade, classe_busca, bg = priority_for_busca(to_numeric(row["Pontuação"], 0))
        tel = infer_phone(row)
        vals = [prioridade, row.get("Nome", ""), row.get("Microárea", "Não informada"), tel]
        vals += [row.get(ch, "") for ch in criterio_headers]
        vals += [row.get("Pontuação", 0), classe_busca, row.get("Pendências", ""), ""]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(i, col_idx, val)
            cell.fill = fill(bg)
            cell.font = font(size=9)
            cell.alignment = align("center" if col_idx not in [2, len(headers)-1, len(headers)] else "left", wrap=True)
            cell.border = border_thin()
            if col_idx == 1:
                if "URGENTE" in str(val):
                    cell.font = font(bold=True, color="C00000", size=9)
                elif "ALTA" in str(val):
                    cell.font = font(bold=True, color=COR["laranja_txt"], size=9)
                elif "MÉDIA" in str(val):
                    cell.font = font(bold=True, color="7F6000", size=9)
                else:
                    cell.font = font(bold=True, color=COR["verde_escuro"], size=9)
            elif 5 <= col_idx < 5 + len(criterio_headers):
                val_up = str(val).strip().upper()
                if val_up == "SIM":
                    cell.fill = fill(COR["verde_ok"]); cell.font = font(bold=True, color=COR["verde_escuro"], size=9)
                elif val_up in {"N/A", "NA"}:
                    cell.fill = fill("E7E6E6"); cell.font = font(bold=True, color="595959", size=9)
                else:
                    cell.fill = fill(COR["vermelho"]); cell.font = font(bold=True, color=COR["vermelho_txt"], size=9)
            elif col_idx == 5 + len(criterio_headers):
                f_, ft_ = score_fill_and_font(to_numeric(val, 0)); cell.fill = f_; cell.font = ft_
            elif col_idx == 6 + len(criterio_headers):
                f_, ft_ = class_fill_and_font(str(val)); cell.fill = f_; cell.font = ft_
    for col_idx, width in BUSCA_WIDTHS.items():
        if col_idx <= ws.max_column:
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    extra_start = len(BUSCA_WIDTHS) + 1
    for idx in range(extra_start, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{end_col}{ws.max_row}"


def create_summary_sheet(wb: Workbook, df: pd.DataFrame, cfg: IndicatorConfig):
    ws = wb.create_sheet("📊 Resumo", 2)
    write_merged_title(ws, "A1:H1", f"RESUMO GERENCIAL – {cfg.code} | {cfg.titulo.split('|')[0].strip()}", COR["azul_escuro"], size=13)
    write_merged_title(ws, "A2:H2", f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Fonte: Nota metodológica + bruto e-SUS", COR["azul_claro"], fg=COR["azul_escuro"], size=9)

    write_merged_title(ws, "A4:H4", f"▶  VISÃO GERAL DO INDICADOR {cfg.code}", COR["roxo"], size=11)
    total = len(df)
    media = round(df["Pontuação"].astype(float).mean(), 1) if total else 0
    completos = int((df["Pontuação"] >= 100).sum()) if total else 0
    busca = int((df["Pontuação"] < 100).sum()) if total else 0

    cards = [
        ("TOTAL PESSOAS", total, "A5:B6", COR["azul_claro"], COR["azul_escuro"]),
        ("PONTUAÇÃO MÉDIA", media, "C5:D6", COR["azul_claro"], COR["azul_escuro"]),
        ("COMPLETOS (100pts)", completos, "E5:F6", COR["verde_ok"], COR["verde_escuro"]),
        ("EM BUSCA ATIVA", busca, "G5:H6", COR["vermelho"], COR["vermelho_txt"]),
    ]
    for titulo, valor, rng, bg, fg in cards:
        write_merged_title(ws, rng, f"{titulo}\n{valor}", bg, fg=fg, size=12)

    write_merged_title(ws, "A9:H9", f"▶  DISTRIBUIÇÃO POR CLASSIFICAÇÃO {cfg.code}", COR["azul_escuro"], size=11)
    headers = ["Classificação", "Qtd", "%", "Parâmetro"]
    start_row = 10
    for idx, h in enumerate(headers, 1):
        c = ws.cell(start_row, idx, h)
        c.fill = fill(COR["azul_header"])
        c.font = font(bold=True, color="1F3864", size=9)
        c.alignment = align("center")
        c.border = border_thin()
    classes = ["Ótimo", "Bom", "Suficiente", "Regular"]
    for i, cls in enumerate(classes, 1):
        qtd = int((df["Classificação"] == cls).sum()) if total else 0
        pct = round((qtd / total) * 100, 1) if total else 0
        row = start_row + i
        vals = [cls, qtd, f"{pct}%", cls]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row, col_idx, val)
            c.fill = fill(COR["branco"] if i % 2 else COR["cinza_claro"])
            c.font = font(size=9)
            c.alignment = align("center")
            c.border = border_thin()
        f_, ft_ = class_fill_and_font(cls)
        ws.cell(row, 1).fill = f_; ws.cell(row, 1).font = ft_

    write_merged_title(ws, "A17:H17", f"▶  ADESÃO POR CRITÉRIO {cfg.code}", COR["roxo"], size=11)
    row0 = 18
    for idx, item in enumerate(cfg.criteria, 1):
        col = idx
        hdr = ws.cell(row0, col, f"{item['letter']} - {item['label']}")
        hdr.fill = fill(COR["roxo_claro"]); hdr.font = font(bold=True, color=COR["roxo"], size=9)
        hdr.alignment = align("center", wrap=True); hdr.border = border_thin()
        crit_col = f"{item['letter']} - {item['label']}"
        if total:
            serie = df[crit_col].astype(str).str.strip().str.upper()
            aplicaveis = serie[~serie.isin({"N/A", "NA"})]
            base = int(len(aplicaveis))
            qtd = int((aplicaveis == "SIM").sum())
            pct = round((qtd / base) * 100, 1) if base else 0
            txt = f"{qtd} / {base}\n({pct}%)" if base else "N/A"
        else:
            qtd = 0
            pct = 0
            txt = "N/A"
            base = 0
        v = ws.cell(row0 + 1, col, txt)
        if not base:
            bg = "E7E6E6"
            fg = "595959"
        else:
            bg = COR["verde_ok"] if pct > 75 else COR["amarelo"] if pct > 50 else COR["vermelho"]
            fg = COR["verde_escuro"] if pct > 75 else COR["amarelo_txt"] if pct > 50 else COR["vermelho_txt"]
        v.fill = fill(bg); v.font = font(bold=True, color=fg, size=10); v.alignment = align("center", wrap=True); v.border = border_thin()

    for col, w in RESUMO_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def create_stats_sheet(wb: Workbook, df: pd.DataFrame, cfg: IndicatorConfig):
    ws = wb.create_sheet("📈 Estatísticas", 3)
    write_merged_title(ws, "A1:J1", f"ESTATÍSTICAS DETALHADAS – {cfg.code} | {cfg.titulo.split('|')[0].strip()}", COR["azul_escuro"], size=13)
    write_merged_title(ws, "A2:J2", f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", COR["azul_claro"], fg=COR["azul_escuro"], size=9)
    write_merged_title(ws, "A4:J4", "📍 DESEMPENHO POR MICROÁREA", COR["roxo"], size=11)
    headers = ["Microárea", "Total", "Ótimo", "Bom", "Suficiente", "Regular", "% Completos", "Média Pts", "Pendentes", "Prioridade Alta"]
    for idx, h in enumerate(headers, 1):
        c = ws.cell(5, idx, h)
        c.fill = fill(COR["azul_header"]); c.font = font(bold=True, color="1F3864", size=9)
        c.alignment = align("center", wrap=True); c.border = border_thin()
    grouped = []
    if len(df):
        temp = df.copy()
        temp["Microárea"] = temp["Microárea"].replace("", "Não informada").fillna("Não informada")
        for micro, g in temp.groupby("Microárea", dropna=False):
            grouped.append([
                micro,
                len(g),
                int((g["Classificação"] == "Ótimo").sum()),
                int((g["Classificação"] == "Bom").sum()),
                int((g["Classificação"] == "Suficiente").sum()),
                int((g["Classificação"] == "Regular").sum()),
                round((g["Pontuação"] >= 100).mean() * 100, 1),
                round(g["Pontuação"].astype(float).mean(), 1),
                int((g["Pontuação"] < 100).sum()),
                int((g["Pontuação"] <= 25).sum()),
            ])
    for i, rowv in enumerate(grouped, 6):
        bg = COR["branco"] if i % 2 == 0 else COR["cinza_claro"]
        for j, val in enumerate(rowv, 1):
            c = ws.cell(i, j, val)
            c.fill = fill(bg); c.font = font(size=9); c.border = border_thin()
            c.alignment = align("center" if j != 1 else "left")
    write_merged_title(ws, "A20:F20", "👥 DESEMPENHO POR SEXO", COR["roxo"], size=11)
    sex_headers = ["Sexo", "Total", "Ótimo", "Bom", "Suficiente", "Regular"]
    for idx, h in enumerate(sex_headers, 1):
        c = ws.cell(21, idx, h)
        c.fill = fill(COR["roxo_claro"]); c.font = font(bold=True, color=COR["roxo"], size=9)
        c.alignment = align("center"); c.border = border_thin()
    row_i = 22
    if len(df):
        temp = df.copy()
        temp["Sexo"] = temp["Sexo"].replace("", "Não informado").fillna("Não informado")
        for sexo, g in temp.groupby("Sexo", dropna=False):
            vals = [sexo, len(g), int((g["Classificação"] == "Ótimo").sum()), int((g["Classificação"] == "Bom").sum()), int((g["Classificação"] == "Suficiente").sum()), int((g["Classificação"] == "Regular").sum())]
            for j, val in enumerate(vals, 1):
                c = ws.cell(row_i, j, val)
                c.fill = fill(COR["branco"] if row_i % 2 == 0 else COR["cinza_claro"]); c.font = font(size=9); c.border = border_thin()
                c.alignment = align("center" if j != 1 else "left")
            row_i += 1
    for col, w in ESTAT_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"


def apply_widths(ws, headers: list[str], criteria_count: int = 0):
    for idx in range(1, len(headers) + 1):
        if idx in BASE_WIDTHS:
            ws.column_dimensions[get_column_letter(idx)].width = BASE_WIDTHS[idx]
        else:
            ws.column_dimensions[get_column_letter(idx)].width = 14



def _file_matches_code_name(path: Path, code: str) -> bool:
    stem = normalize_text(path.stem)
    code_n = normalize_text(code)
    return (
        stem == code_n
        or stem.startswith(code_n + "_")
        or stem.startswith(code_n + "-")
        or stem.startswith(code_n + " ")
    )


def candidate_file_for_indicator(files: list[Path], cfg: IndicatorConfig) -> Path | None:
    valid = []
    keywords = [normalize_text(k) for k in cfg.theme_keywords]

    for f in files:
        name = normalize_text(f.name)
        if "resultado" in name or "_interativa" in name:
            continue
        if f.suffix.lower() not in {".csv", ".xlsx", ".xls"}:
            continue

        # 1) Primeiro: aceitar arquivo pelo código no nome (ex.: c1.csv, c1_teste.csv)
        if _file_matches_code_name(f, cfg.code):
            valid.append(f)
            continue

        # 2) Depois: tentar pelas palavras-chave do indicador no nome do arquivo
        if any(k in name for k in keywords):
            valid.append(f)
            continue

        # 3) Por fim: tentar detectar o tema dentro do conteúdo do CSV
        if f.suffix.lower() == ".csv":
            try:
                theme = normalize_text(detect_theme_label(f))
                if any(k in theme for k in keywords):
                    valid.append(f)
                    continue
            except Exception:
                pass

    if not valid:
        return None

    # Prioriza o nome por código do indicador
    valid.sort(
        key=lambda p: (
            0 if _file_matches_code_name(p, cfg.code) else 1,
            -p.stat().st_mtime
        )
    )
    return valid[0]


def process_indicator(cfg: IndicatorConfig, input_path: str | Path, output_path: str | Path):
    df_raw = read_esus_table(input_path)
    df_out = cfg.builder(df_raw)
    render_workbook(df_out, cfg, output_path)
