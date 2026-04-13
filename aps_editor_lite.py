from __future__ import annotations

import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from aps_clonador_interativo import (
    _detect_columns,
    _detect_header,
    _ensure_support_columns,
    _normalize_status,
    _patients_from_data,
    _update_data_sheet,
)


class EditorLiteApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("APS - Editor Lite")
        self.geometry("1220x780")
        self.configure(bg="#EEF4F8")
        try:
            self.state("zoomed")
        except Exception:
            pass

        self.path_var = tk.StringVar()
        self.search_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Abra uma planilha para editar.")
        self.pending_var = tk.StringVar(value="0 alteracoes pendentes")

        self.workbook_path: Path | None = None
        self.base_records: list[dict] = []
        self.records: list[dict] = []
        self.record_map: dict[int, dict] = {}
        self.criteria_info: list[tuple[str, str, int]] = []
        self.criteria_index: dict[str, int] = {}
        self.pending_by_row: dict[int, dict[str, str]] = {}
        self.current_row: int | None = None
        self.criterio_vars: dict[str, tk.StringVar] = {}

        self._build_ui()

    def _build_ui(self):
        top = tk.Frame(self, bg="#EEF4F8")
        top.pack(fill="x", padx=12, pady=12)
        tk.Label(top, text="APS - EDITOR LITE", bg="#1F4E79", fg="white", font=("Segoe UI", 13, "bold"), pady=10).pack(fill="x")

        open_line = tk.Frame(self, bg="#EEF4F8")
        open_line.pack(fill="x", padx=12, pady=(0, 10))
        tk.Entry(open_line, textvariable=self.path_var).pack(side="left", fill="x", expand=True)
        tk.Button(open_line, text="Abrir planilha", command=self.choose_file).pack(side="left", padx=(8, 0))
        tk.Button(open_line, text="Salvar", command=self.save_changes, bg="#1F4E79", fg="white").pack(side="left", padx=(8, 0))

        filters = tk.Frame(self, bg="#EEF4F8")
        filters.pack(fill="x", padx=12, pady=(0, 8))
        tk.Label(filters, text="Buscar:", bg="#EEF4F8").pack(side="left")
        ent = tk.Entry(filters, textvariable=self.search_var, width=30)
        ent.pack(side="left", padx=(6, 0))
        ent.bind("<KeyRelease>", lambda _e: self._apply_filter())
        tk.Label(filters, textvariable=self.pending_var, bg="#EEF4F8", fg="#1F4E79").pack(side="right")

        body = tk.Frame(self, bg="#EEF4F8")
        body.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(body, columns=("prio", "nome", "bairro", "pts"), show="headings")
        for key, title, w, anc in [
            ("prio", "Prioridade", 130, "center"),
            ("nome", "Nome", 290, "w"),
            ("bairro", "Bairro", 150, "center"),
            ("pts", "Pontuacao", 90, "center"),
        ]:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=w, anchor=anc)
        self.tree.tag_configure("urgente", background="#FDECEA")
        self.tree.tag_configure("alta", background="#FFF4E5")
        self.tree.tag_configure("monitorar", background="#FFFBE6")
        self.tree.tag_configure("concluido", background="#EAF7EA")
        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        yscroll = ttk.Scrollbar(body, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        yscroll.grid(row=0, column=0, sticky="nse")

        right = tk.Frame(body, bg="#FFFFFF", bd=1, relief="solid")
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        self.lbl_patient = tk.Label(right, text="Paciente: -", bg="#FFFFFF", anchor="w", font=("Segoe UI", 12, "bold"))
        self.lbl_patient.pack(fill="x", padx=10, pady=(10, 8))
        self.lbl_meta = tk.Label(right, text="", bg="#FFFFFF", anchor="w", justify="left")
        self.lbl_meta.pack(fill="x", padx=10, pady=(0, 8))
        self.frm_criteria = tk.Frame(right, bg="#FFFFFF")
        self.frm_criteria.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        tk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w").pack(fill="x", side="bottom")

    def choose_file(self):
        p = filedialog.askopenfilename(
            title="Selecione a planilha para editar",
            filetypes=[("Excel", "*.xlsx *.xls")],
            initialdir=str(Path.home()),
        )
        if not p:
            return
        self.path_var.set(p)
        self.workbook_path = Path(p)
        self.reload_data()

    def _load_context(self):
        if not self.workbook_path:
            raise RuntimeError("Nenhuma planilha aberta.")
        wb = load_workbook(self.workbook_path)
        ws = next((wb[n] for n in wb.sheetnames if str(n).startswith("📋 Dados") or str(n).startswith("Dados")), wb[wb.sheetnames[0]])
        header = _detect_header(ws)
        cols = _ensure_support_columns(ws, header, _detect_columns(ws, header))
        return wb, ws, header, cols

    def _norm_prio(self, value: str) -> str:
        t = str(value or "").upper()
        if "URGENTE" in t:
            return "URGENTE"
        if "ALTA" in t:
            return "ALTA"
        if "MONITOR" in t:
            return "MONITORAR"
        if "CONCL" in t:
            return "CONCLUIDO"
        return t

    def _effective_record(self, base: dict) -> dict:
        row = base["row"]
        overrides = self.pending_by_row.get(row, {})
        if not overrides:
            return dict(base)
        rec = dict(base)
        sts = list(base["statuses"])
        for title, value in overrides.items():
            idx = self.criteria_index.get(title)
            if idx is not None:
                sts[idx] = _normalize_status(value)
        rec["statuses"] = sts
        sim = sum(1 for s in sts if _normalize_status(s) == "SIM")
        n = max(len(sts), 1)
        pts = int(round((sim / n) * 100, 0))
        rec["pts"] = pts
        if pts >= 100:
            rec["prio"] = "CONCLUIDO"
        elif pts >= 75:
            rec["prio"] = "MONITORAR"
        elif pts >= 50:
            rec["prio"] = "ALTA"
        else:
            rec["prio"] = "URGENTE"
        rec["dirty"] = True
        return rec

    def reload_data(self):
        wb, ws, header, cols = self._load_context()
        self.criteria_info = cols["criterios"]
        self.criteria_index = {title: idx for idx, (_l, title, _c) in enumerate(self.criteria_info)}
        patients = _patients_from_data(ws, header, cols)
        wb.close()
        self.base_records = patients
        self._rebuild_records()
        self._apply_filter()
        self.status_var.set(f"Planilha carregada: {self.workbook_path.name} ({len(self.base_records)} pacientes).")

    def _rebuild_records(self):
        rows = [self._effective_record(p) for p in self.base_records]
        rows.sort(key=lambda x: (self._norm_prio(x.get("prio", "")), x.get("pts", 0), str(x.get("nome", "")).lower()))
        self.records = rows
        self.record_map = {r["row"]: r for r in rows}
        rows_dirty = len(self.pending_by_row)
        cells_dirty = sum(len(v) for v in self.pending_by_row.values())
        self.pending_var.set(f"{rows_dirty} linhas com alteracao | {cells_dirty} campos")

    def _apply_filter(self):
        term = self.search_var.get().strip().lower()
        self.tree.delete(*self.tree.get_children(""))
        for p in self.records:
            nome = str(p.get("nome", "")).lower()
            bairro = str(p.get("bairro", "")).lower()
            if term and term not in nome and term not in bairro:
                continue
            pr = self._norm_prio(p.get("prio", ""))
            tag = {"URGENTE": "urgente", "ALTA": "alta", "MONITORAR": "monitorar", "CONCLUIDO": "concluido"}.get(pr, "monitorar")
            self.tree.insert("", "end", iid=str(p["row"]), values=(pr, p["nome"], p.get("bairro", ""), p.get("pts", 0)), tags=(tag,))

    def _on_select(self, _evt=None):
        sel = self.tree.selection()
        if not sel:
            return
        row = int(sel[0])
        self.current_row = row
        p = self.record_map.get(row)
        if not p:
            return
        self.lbl_patient.config(text=f"Paciente: {p.get('nome', '')}")
        self.lbl_meta.config(text=f"Bairro: {p.get('bairro','')}\nPontuacao: {p.get('pts',0)}\nPrioridade: {self._norm_prio(p.get('prio',''))}")
        for w in self.frm_criteria.winfo_children():
            w.destroy()
        self.criterio_vars = {}
        for i, ((_letter, title, _col), value) in enumerate(zip(self.criteria_info, p["statuses"])):
            txt = title.split("-", 1)[1].strip() if "-" in title else title
            tk.Label(self.frm_criteria, text=txt, bg="#FFFFFF", anchor="w").grid(row=i, column=0, sticky="w", pady=4, padx=(0, 8))
            var = tk.StringVar(value=str(value))
            self.criterio_vars[title] = var
            cb = ttk.Combobox(self.frm_criteria, textvariable=var, values=("", "SIM", "NAO", "PENDENTE"), state="readonly", width=14)
            cb.grid(row=i, column=1, sticky="ew", pady=4)
            var.trace_add("write", lambda *_a, t=title, v=var: self._on_criterion_changed(row, t, v.get()))
        self.frm_criteria.columnconfigure(1, weight=1)

    def _on_criterion_changed(self, row: int, title: str, value: str):
        base = next((r for r in self.base_records if r["row"] == row), None)
        if not base:
            return
        idx = self.criteria_index.get(title)
        if idx is None:
            return
        normalized = _normalize_status(value)
        base_norm = _normalize_status(base["statuses"][idx])
        changes = self.pending_by_row.get(row, {}).copy()
        if normalized == base_norm:
            changes.pop(title, None)
        else:
            changes[title] = normalized
        if changes:
            self.pending_by_row[row] = changes
        else:
            self.pending_by_row.pop(row, None)
        self._rebuild_records()
        self._apply_filter()
        if self.tree.exists(str(row)):
            self.tree.selection_set(str(row))
            self.tree.focus(str(row))

    def save_changes(self):
        if not self.workbook_path:
            return
        if not self.pending_by_row:
            self.status_var.set("Nao ha alteracoes pendentes.")
            return
        try:
            wb, ws, header, cols = self._load_context()
            title_to_col = {title: col for _l, title, col in cols["criterios"]}
            for row, changes in self.pending_by_row.items():
                for title, value in changes.items():
                    col = title_to_col.get(title)
                    if col is None:
                        continue
                    ws.cell(row, col).value = value
            _update_data_sheet(ws, header, cols)
            wb.save(self.workbook_path)
            wb.close()
            self.pending_by_row.clear()
            self.reload_data()
            self.status_var.set("Alteracoes salvas com sucesso.")
        except Exception as exc:
            messagebox.showerror("Erro ao salvar", str(exc), parent=self)
            self.status_var.set(f"Erro ao salvar: {exc}")


def main():
    app = EditorLiteApp()
    app.mainloop()


if __name__ == "__main__":
    main()

