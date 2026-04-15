from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import load_workbook


STORE_FILENAME = "aprazamento_controle.json"
CODES = ("C1", "C2", "C3", "C4", "C5", "C6", "C7")
DEFAULT_INTERVAL_MONTHS = 4
DEFAULT_INTERVAL_DAYS = 0
BASE_MODES = ("ENFERMAGEM", "MEDICO", "MAIS_RECENTE", "AMBOS_MAIS_PROXIMO", "AMBOS_MAIS_LONGE")
INTERVAL_PRESETS: dict[str, tuple[int, int] | None] = {
    "BIMESTRAL": (2, 0),
    "QUADRIMESTRAL": (4, 0),
    "SEMESTRAL": (6, 0),
    "ANUAL": (12, 0),
    "LIVRE": None,
}
DEFAULT_SETTINGS = {
    "interval_preset": "QUADRIMESTRAL",
    "custom_interval_value": 4,
    "custom_interval_unit": "MESES",
    "global_base_mode": "ENFERMAGEM",
}


CODE_ALIASES: dict[str, tuple[str, ...]] = {
    "C1": ("mais acesso",),
    "C2": ("desenvolvimento infantil", "infantil"),
    "C3": ("gestacao", "puerperio"),
    "C4": ("diabetes",),
    "C5": ("hipertensao",),
    "C6": ("idosa", "idoso"),
    "C7": ("mulher", "cancer"),
}


def _norm(text) -> str:
    txt = unicodedata.normalize("NFKD", str(text or "")).encode("ascii", "ignore").decode("ascii")
    return txt.strip().lower()


def _norm_spaces(text) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _norm(text)).strip()


def _clean_id(value) -> str:
    txt = str(value or "").strip()
    if txt in {"", "-", "nan", "None"}:
        return ""
    return txt


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    txt = str(value).strip()
    if txt in {"", "-", "nan", "None"}:
        return None
    txt = txt.replace("\\", "/")
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(txt, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(txt).date()
    except Exception:
        return None


def fmt_date(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def add_months(d: date, months: int) -> date:
    month_idx = (d.month - 1) + int(months)
    year = d.year + (month_idx // 12)
    month = (month_idx % 12) + 1
    last_day = [31, 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]
    day = min(d.day, last_day)
    return date(year, month, day)


def _to_float(value) -> float | None:
    if value is None:
        return None
    txt = str(value).strip()
    if txt in {"", "-", "nan", "None"}:
        return None
    txt = txt.replace(".", "").replace(",", ".")
    try:
        n = float(txt)
    except Exception:
        return None
    if n < 0:
        return None
    return n


def _estimate_date_from_months(months_value, ref_date: date) -> date | None:
    n = _to_float(months_value)
    if n is None:
        return None
    days = int(round(n * 30))
    return ref_date - timedelta(days=days)


def _estimate_date_from_days(days_value, ref_date: date) -> date | None:
    n = _to_float(days_value)
    if n is None:
        return None
    return ref_date - timedelta(days=int(round(n)))


def _estimate_date_from_recency(days_value, months_value, ref_date: date) -> date | None:
    days_n = _to_float(days_value)
    months_n = _to_float(months_value)
    if days_n is None and months_n is None:
        return None

    dt = ref_date
    if months_n is not None:
        whole = int(months_n)
        frac_days = int(round((months_n - whole) * 30))
        if whole:
            dt = add_months(dt, -whole)
        if frac_days:
            dt = dt - timedelta(days=frac_days)
    if days_n is not None:
        dt = dt - timedelta(days=int(round(days_n)))
    return dt


def _semaphore(days_to_due: int | None) -> str:
    if days_to_due is None:
        return "SEM DATA"
    if days_to_due < 0:
        return "VENCIDO"
    if days_to_due <= 7:
        return "VERMELHO"
    if days_to_due <= 30:
        return "AMARELO"
    return "VERDE"


def _priority_order(sem: str) -> int:
    return {"VENCIDO": 0, "VERMELHO": 1, "AMARELO": 2, "VERDE": 3, "SEM DATA": 4}.get(sem, 9)


def _next_from_base(base_date: date | None, interval_months: int, interval_days: int) -> date | None:
    if not base_date:
        return None
    if int(interval_days or 0) > 0:
        return base_date + timedelta(days=int(interval_days))
    return add_months(base_date, int(interval_months))


def _choose_base_date(
    med_date: date | None,
    enf_date: date | None,
    mode: str,
    interval_months: int,
    interval_days: int,
) -> tuple[date | None, str]:
    mode_up = str(mode or "").upper().strip()
    if mode_up == "MEDICO":
        return med_date, "MEDICO"
    if mode_up == "ENFERMAGEM":
        return enf_date, "ENFERMAGEM"
    if mode_up == "AMBOS_MAIS_PROXIMO" or mode_up == "AMBOS_MAIS_LONGE":
        candidates: list[tuple[str, date, date]] = []
        if med_date:
            next_med = _next_from_base(med_date, interval_months, interval_days)
            if next_med:
                candidates.append(("MEDICO", med_date, next_med))
        if enf_date:
            next_enf = _next_from_base(enf_date, interval_months, interval_days)
            if next_enf:
                candidates.append(("ENFERMAGEM", enf_date, next_enf))
        if not candidates:
            return None, "-"
        if len(candidates) == 1:
            source, base, _next = candidates[0]
            return base, source
        if mode_up == "AMBOS_MAIS_PROXIMO":
            source, base, _next = min(candidates, key=lambda t: t[2])
            return base, source
        source, base, _next = max(candidates, key=lambda t: t[2])
        return base, source

    if med_date and enf_date:
        if med_date >= enf_date:
            return med_date, "MEDICO"
        return enf_date, "ENFERMAGEM"
    if med_date:
        return med_date, "MEDICO"
    if enf_date:
        return enf_date, "ENFERMAGEM"
    return None, "-"


def compute_control_fields(record: dict, ref_date: date | None = None) -> dict:
    today = ref_date or date.today()
    med_date = parse_date(record.get("last_medico_date"))
    enf_date = parse_date(record.get("last_enfermagem_date"))
    interval = int(record.get("interval_months") or DEFAULT_INTERVAL_MONTHS)
    interval_days = int(record.get("interval_days") or DEFAULT_INTERVAL_DAYS)
    mode = str(record.get("base_mode") or "ENFERMAGEM").upper().strip()
    if mode not in BASE_MODES:
        mode = "ENFERMAGEM"

    base_date, base_source = _choose_base_date(med_date, enf_date, mode, interval, interval_days)
    manual_next = parse_date(record.get("manual_next_date"))
    if manual_next:
        next_date = manual_next
    elif base_date:
        next_date = _next_from_base(base_date, interval, interval_days)
    else:
        next_date = None

    if next_date is None:
        days_to_due = None
    else:
        days_to_due = (next_date - today).days

    sem = _semaphore(days_to_due)
    return {
        "base_mode": mode,
        "base_date": fmt_date(base_date),
        "base_source": base_source,
        "next_date": fmt_date(next_date),
        "days_to_due": days_to_due if days_to_due is not None else "",
        "semaphore": sem,
    }


def _find_header_row(ws) -> int:
    max_scan = min(12, ws.max_row)
    for r in range(1, max_scan + 1):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, min(40, ws.max_column) + 1)]
        if "nome" in vals and ("cpf" in vals or "cns" in vals):
            return r
    return 3


def _pick_col(cols: dict[str, int], *names: str, contains: tuple[str, ...] = ()) -> int | None:
    for name in names:
        n = _norm_spaces(name)
        if n in cols:
            return cols[n]
    if contains:
        req = [_norm_spaces(p) for p in contains]
        for title, idx in cols.items():
            if all(p in title for p in req):
                return idx
    return None


def _select_data_sheet_name(wb) -> str:
    for name in wb.sheetnames:
        if "dados" in _norm(name):
            return name
    return wb.sheetnames[0]


def _load_rows_from_workbook(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_select_data_sheet_name(wb)]
        header = _find_header_row(ws)
        cols: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            key = _norm_spaces(ws.cell(header, c).value)
            if key:
                cols[key] = c

        col_name = _pick_col(cols, "nome", "paciente")
        col_cpf = _pick_col(cols, "cpf")
        col_cns = _pick_col(cols, "cns")
        col_tel1 = _pick_col(cols, "telefone celular", "telefone")
        col_tel2 = _pick_col(cols, "telefone residencial")
        col_tel3 = _pick_col(cols, "telefone de contato")
        col_med = _pick_col(cols, contains=("meses", "atendimento", "medico"))
        col_enf = _pick_col(cols, contains=("meses", "atendimento", "enfermagem"))
        col_days_med = _pick_col(cols, contains=("dias", "atendimento", "medico"))
        col_days_enf = _pick_col(cols, contains=("dias", "atendimento", "enfermagem"))
        col_antrop = _pick_col(cols, contains=("data", "medicao", "peso", "altura"))
        col_sinais = _pick_col(cols, contains=("data", "medicao", "pressao", "arterial"))

        out: list[dict] = []
        if not col_name:
            return out

        max_col = max(c for c in [col_name, col_cpf, col_cns, col_tel1, col_tel2, col_tel3, col_med, col_enf, col_days_med, col_days_enf, col_antrop, col_sinais] if c)
        blank_streak = 0
        hard_limit = min(ws.max_row, header + 200000)

        def _cell_value(row_vals, col_idx: int | None):
            if not col_idx:
                return None
            i = col_idx - 1
            if i < 0 or i >= len(row_vals):
                return None
            return row_vals[i]

        for row_vals in ws.iter_rows(min_row=header + 1, max_row=hard_limit, min_col=1, max_col=max_col, values_only=True):
            name = str(_cell_value(row_vals, col_name) or "").strip()
            if not name or name == "-":
                blank_streak += 1
                # Em planilhas com estilo aplicado em milhares de linhas vazias,
                # interrompe cedo para evitar leitura muito lenta.
                if blank_streak >= 300:
                    break
                continue
            blank_streak = 0
            phones = []
            for tel_col in (col_tel1, col_tel2, col_tel3):
                v = str(_cell_value(row_vals, tel_col) or "").strip()
                if v and v not in {"-", "nan", "None"}:
                    phones.append(v)
            out.append(
                {
                    "name": name,
                    "cpf": _clean_id(_cell_value(row_vals, col_cpf)),
                    "cns": _clean_id(_cell_value(row_vals, col_cns)),
                    "phone": " | ".join(phones) if phones else "",
                    "months_medico": _cell_value(row_vals, col_med),
                    "months_enfermagem": _cell_value(row_vals, col_enf),
                    "days_medico": _cell_value(row_vals, col_days_med),
                    "days_enfermagem": _cell_value(row_vals, col_days_enf),
                    "date_antropometria": _cell_value(row_vals, col_antrop),
                    "date_sinais_vitais": _cell_value(row_vals, col_sinais),
                }
            )
        return out
    finally:
        wb.close()


def patient_key(name: str, cpf: str, cns: str) -> str:
    cpf_clean = _clean_id(cpf)
    if cpf_clean:
        return f"CPF:{cpf_clean}"
    cns_clean = _clean_id(cns)
    if cns_clean:
        return f"CNS:{cns_clean}"
    return f"NOME:{_norm_spaces(name)}"


def find_existing_record_id(records: dict[str, dict], name: str, cpf: str = "", cns: str = "") -> str | None:
    cpf_clean = _clean_id(cpf)
    cns_clean = _clean_id(cns)
    name_norm = _norm_spaces(name)
    if not name_norm and not cpf_clean and not cns_clean:
        return None

    direct_id = patient_key(name, cpf_clean, cns_clean)
    if direct_id in records:
        return direct_id

    if cpf_clean:
        for rid, rec in records.items():
            if _clean_id(rec.get("cpf")) == cpf_clean:
                return rid
    if cns_clean:
        for rid, rec in records.items():
            if _clean_id(rec.get("cns")) == cns_clean:
                return rid
    if name_norm:
        for rid, rec in records.items():
            if _norm_spaces(rec.get("name")) == name_norm:
                return rid
    return None


def _code_in_name(stem: str, code: str) -> bool:
    return re.search(rf"(?<![0-9a-z]){code.lower()}(?![0-9])", stem.lower()) is not None


def find_latest_indicator_files(folder: Path) -> dict[str, Path]:
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"} and not p.name.startswith("~$")]
    result: dict[str, Path] = {}
    for code in CODES:
        alias_tokens = CODE_ALIASES.get(code, ())
        candidates = []
        for f in files:
            stem = f.stem
            stem_n = _norm(stem)
            if _code_in_name(stem, code):
                candidates.append(f)
                continue
            if any(_norm(token) in stem_n for token in alias_tokens):
                candidates.append(f)
        if candidates:
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            result[code] = candidates[0]
    return result


def build_records_from_folder(
    folder: Path,
    existing_records: dict[str, dict] | None = None,
    ref_date: date | None = None,
    default_base_mode: str = "ENFERMAGEM",
    default_interval_months: int = DEFAULT_INTERVAL_MONTHS,
    default_interval_days: int = DEFAULT_INTERVAL_DAYS,
) -> tuple[dict[str, dict], dict[str, str]]:
    ref = ref_date or date.today()
    existing_records = existing_records or {}
    files = find_latest_indicator_files(folder)
    if "C1" not in files:
        raise FileNotFoundError("Nao foi encontrada planilha C1 na pasta selecionada.")

    c1_rows = _load_rows_from_workbook(files["C1"])
    records: dict[str, dict] = {}

    for row in c1_rows:
        key = patient_key(row["name"], row["cpf"], row["cns"])
        base = dict(existing_records.get(key, {}))
        antrop_date = fmt_date(parse_date(row.get("date_antropometria")))
        sinais_date = fmt_date(parse_date(row.get("date_sinais_vitais")))
        record = {
            "id": key,
            "name": row["name"],
            "cpf": row["cpf"],
            "cns": row["cns"],
            "phone": row["phone"] or base.get("phone", ""),
            "conditions": sorted(set(base.get("conditions", []))),
            "last_medico_date": base.get("last_medico_date", ""),
            "last_enfermagem_date": base.get("last_enfermagem_date", ""),
            "last_antropometria_date": (base.get("last_antropometria_date") or antrop_date),
            "last_sinais_vitais_date": (base.get("last_sinais_vitais_date") or sinais_date),
            "estimated_medico": bool(base.get("estimated_medico", False)),
            "estimated_enfermagem": bool(base.get("estimated_enfermagem", False)),
            "base_mode": str(base.get("base_mode") or default_base_mode).upper(),
            "interval_months": int(base.get("interval_months") or default_interval_months),
            "interval_days": int(base.get("interval_days") or default_interval_days),
            "manual_next_date": base.get("manual_next_date", ""),
            "notes": base.get("notes", ""),
        }

        if not parse_date(record["last_medico_date"]):
            est_med = _estimate_date_from_recency(row.get("days_medico"), row.get("months_medico"), ref)
            if est_med:
                record["last_medico_date"] = fmt_date(est_med)
                record["estimated_medico"] = True

        if not parse_date(record["last_enfermagem_date"]):
            est_enf = _estimate_date_from_recency(row.get("days_enfermagem"), row.get("months_enfermagem"), ref)
            if est_enf:
                record["last_enfermagem_date"] = fmt_date(est_enf)
                record["estimated_enfermagem"] = True

        records[key] = record

    for code in CODES:
        if code == "C1":
            continue
        path = files.get(code)
        if not path:
            continue
        for row in _load_rows_from_workbook(path):
            key = patient_key(row["name"], row["cpf"], row["cns"])
            rec = records.get(key)
            if not rec:
                continue
            conds = set(rec.get("conditions", []))
            conds.add(code)
            rec["conditions"] = sorted(conds)

    for rec in records.values():
        rec.update(compute_control_fields(rec, ref))

    used = {code: str(path) for code, path in files.items()}
    return records, used


def load_store_with_meta(path: Path) -> tuple[dict[str, dict], dict]:
    if not path.exists():
        return {}, {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}, {}
    items = data.get("patients", [])
    out: dict[str, dict] = {}
    for rec in items:
        rid = str(rec.get("id") or "").strip()
        if rid:
            out[rid] = rec
    meta = data.get("meta") if isinstance(data, dict) else {}
    if not isinstance(meta, dict):
        meta = {}
    return out, meta


def load_store(path: Path) -> dict[str, dict]:
    records, _meta = load_store_with_meta(path)
    return records


def save_store(
    path: Path,
    records: dict[str, dict],
    source_files: dict[str, str] | None = None,
    settings: dict | None = None,
) -> None:
    payload = {
        "meta": {
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "source_files": source_files or {},
            "settings": settings or {},
        },
        "patients": sorted(records.values(), key=lambda r: (_priority_order(str(r.get("semaphore", ""))), int(r.get("days_to_due") if str(r.get("days_to_due", "")).strip() not in {"", "None"} else 999999), _norm_spaces(r.get("name", "")))),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


@dataclass
class FilterState:
    term: str = ""
    semaphore: str = "TODOS"
    condition: str = "TODAS"


class AprazamentoApp(tk.Toplevel):
    def __init__(self, master=None, base_dir: Path | None = None, auto_import: bool = False):
        super().__init__(master=master)
        self.title("APS - Controle de Aprazamento")
        self.geometry("1320x820")
        self.configure(bg="#EEF4F8")
        try:
            self.state("zoomed")
        except Exception:
            pass

        self.base_dir_var = tk.StringVar(value=str(base_dir or (Path.home() / "Desktop" / "APS_RESULTADOS")))
        self.search_var = tk.StringVar()
        self.semaphore_filter_var = tk.StringVar(value="TODOS")
        self.condition_filter_var = tk.StringVar(value="TODAS")
        self.status_var = tk.StringVar(value="Selecione a pasta de resultados e importe C1..C7.")
        self.count_var = tk.StringVar(value="0 pacientes")
        self.files_var = tk.StringVar(value="Arquivos: -")
        self.settings_preset_var = tk.StringVar(value=str(DEFAULT_SETTINGS["interval_preset"]))
        self.settings_custom_value_var = tk.StringVar(value=str(DEFAULT_SETTINGS["custom_interval_value"]))
        self.settings_custom_unit_var = tk.StringVar(value=str(DEFAULT_SETTINGS["custom_interval_unit"]))
        self.settings_base_mode_var = tk.StringVar(value=str(DEFAULT_SETTINGS["global_base_mode"]))
        self.global_interval_days = tk.IntVar(value=DEFAULT_INTERVAL_DAYS)

        self.records: dict[str, dict] = {}
        self.filtered_ids: list[str] = []
        self.current_id: str | None = None
        self.source_files: dict[str, str] = {}
        self._settings_window: tk.Toplevel | None = None

        self.name_var = tk.StringVar(value="-")
        self.id_var = tk.StringVar(value="-")
        self.cond_var = tk.StringVar(value="-")
        self.phone_var = tk.StringVar(value="-")
        self.med_var = tk.StringVar()
        self.enf_var = tk.StringVar()
        self.antrop_var = tk.StringVar()
        self.sinais_var = tk.StringVar()
        self.base_mode_var = tk.StringVar(value="ENFERMAGEM")
        self.interval_var = tk.StringVar(value=str(DEFAULT_INTERVAL_MONTHS))
        self.next_var = tk.StringVar()
        self.days_var = tk.StringVar()
        self.sem_var = tk.StringVar()
        self.base_info_var = tk.StringVar()
        self.notes_var = tk.StringVar()

        self._build_ui()
        self._load_existing_store()
        if auto_import:
            self._import_folder(silent_errors=True)
        self._apply_filter()

    def _store_path(self) -> Path:
        return Path(self.base_dir_var.get().strip()) / STORE_FILENAME

    def _current_settings_payload(self) -> dict:
        return {
            "interval_preset": str(self.settings_preset_var.get() or "QUADRIMESTRAL").upper(),
            "custom_interval_value": int(self.settings_custom_value_var.get() or DEFAULT_INTERVAL_MONTHS),
            "custom_interval_unit": str(self.settings_custom_unit_var.get() or "MESES").upper(),
            "global_base_mode": str(self.settings_base_mode_var.get() or "ENFERMAGEM").upper(),
        }

    def _load_settings_from_payload(self, settings: dict | None):
        cfg = dict(DEFAULT_SETTINGS)
        if isinstance(settings, dict):
            cfg.update({k: v for k, v in settings.items() if k in cfg})
        preset = str(cfg.get("interval_preset", "QUADRIMESTRAL")).upper()
        if preset not in INTERVAL_PRESETS:
            preset = "QUADRIMESTRAL"
        base_mode = str(cfg.get("global_base_mode", "ENFERMAGEM")).upper()
        if base_mode not in BASE_MODES:
            base_mode = "ENFERMAGEM"
        custom_unit = str(cfg.get("custom_interval_unit", "MESES")).upper()
        if custom_unit not in {"MESES", "DIAS"}:
            custom_unit = "MESES"
        try:
            custom_value = int(cfg.get("custom_interval_value", DEFAULT_INTERVAL_MONTHS))
        except Exception:
            custom_value = DEFAULT_INTERVAL_MONTHS
        if custom_value <= 0:
            custom_value = DEFAULT_INTERVAL_MONTHS

        self.settings_preset_var.set(preset)
        self.settings_custom_value_var.set(str(custom_value))
        self.settings_custom_unit_var.set(custom_unit)
        self.settings_base_mode_var.set(base_mode)

    def _interval_from_settings(self) -> tuple[int, int]:
        preset = str(self.settings_preset_var.get() or "QUADRIMESTRAL").upper()
        preset_cfg = INTERVAL_PRESETS.get(preset)
        if preset_cfg is not None:
            months, days = preset_cfg
            return int(months), int(days)

        unit = str(self.settings_custom_unit_var.get() or "MESES").upper()
        try:
            value = int(self.settings_custom_value_var.get() or DEFAULT_INTERVAL_MONTHS)
        except Exception:
            value = DEFAULT_INTERVAL_MONTHS
        if value <= 0:
            value = DEFAULT_INTERVAL_MONTHS
        if unit == "DIAS":
            return DEFAULT_INTERVAL_MONTHS, value
        return value, 0

    def _apply_global_settings_to_records(self):
        months, days = self._interval_from_settings()
        base_mode = str(self.settings_base_mode_var.get() or "ENFERMAGEM").upper()
        if base_mode not in BASE_MODES:
            base_mode = "ENFERMAGEM"
        self.global_interval_days.set(days)
        for rec in self.records.values():
            rec["interval_months"] = months
            rec["interval_days"] = days
            rec["base_mode"] = base_mode
            rec.update(compute_control_fields(rec))

    def _refresh_settings_custom_state(self):
        if not self._settings_window or not self._settings_window.winfo_exists():
            return
        state = "normal" if str(self.settings_preset_var.get() or "").upper() == "LIVRE" else "disabled"
        if hasattr(self, "_settings_custom_value_entry"):
            self._settings_custom_value_entry.configure(state=state)
        if hasattr(self, "_settings_custom_unit_combo"):
            self._settings_custom_unit_combo.configure(state="readonly" if state == "normal" else "disabled")

    def _build_ui(self):
        top = tk.Frame(self, bg="#EEF4F8")
        top.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(top, text="APS - CONTROLE DE APRAZAMENTO", bg="#1F4E79", fg="white", font=("Segoe UI", 13, "bold"), pady=10).pack(fill="x")

        line = tk.Frame(self, bg="#EEF4F8")
        line.pack(fill="x", padx=12, pady=(0, 8))
        tk.Label(line, text="Pasta:", bg="#EEF4F8").pack(side="left")
        tk.Entry(line, textvariable=self.base_dir_var).pack(side="left", fill="x", expand=True, padx=(6, 6))
        tk.Button(line, text="Escolher", command=self._choose_folder).pack(side="left")
        tk.Button(line, text="Importar C1..C7", command=self._import_folder, bg="#1F4E79", fg="white").pack(side="left", padx=(6, 0))
        tk.Button(line, text="Configuracao", command=self._open_settings_window).pack(side="left", padx=(6, 0))
        tk.Button(line, text="Salvar base", command=self._save_store_clicked).pack(side="left", padx=(6, 0))

        filters = tk.Frame(self, bg="#EEF4F8")
        filters.pack(fill="x", padx=12, pady=(0, 8))
        tk.Label(filters, text="Buscar:", bg="#EEF4F8").pack(side="left")
        ent = tk.Entry(filters, textvariable=self.search_var, width=34)
        ent.pack(side="left", padx=(6, 10))
        ent.bind("<KeyRelease>", lambda _e: self._apply_filter())
        tk.Button(filters, text="Adicionar paciente", command=self._add_patient_dialog).pack(side="left", padx=(0, 12))

        tk.Label(filters, text="Semaforo:", bg="#EEF4F8").pack(side="left")
        cb_sem = ttk.Combobox(filters, textvariable=self.semaphore_filter_var, values=("TODOS", "VENCIDO", "VERMELHO", "AMARELO", "VERDE", "SEM DATA"), state="readonly", width=12)
        cb_sem.pack(side="left", padx=(6, 10))
        cb_sem.bind("<<ComboboxSelected>>", lambda _e: self._apply_filter())

        tk.Label(filters, text="Condicao:", bg="#EEF4F8").pack(side="left")
        cb_cond = ttk.Combobox(filters, textvariable=self.condition_filter_var, values=("TODAS", "C2", "C3", "C4", "C5", "C6", "C7"), state="readonly", width=10)
        cb_cond.pack(side="left", padx=(6, 10))
        cb_cond.bind("<<ComboboxSelected>>", lambda _e: self._apply_filter())

        tk.Label(filters, textvariable=self.count_var, bg="#EEF4F8", fg="#1F4E79").pack(side="right")

        body = tk.Frame(self, bg="#EEF4F8")
        body.pack(fill="both", expand=True, padx=12, pady=(0, 10))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(body, columns=("sem", "dias", "nome", "tel", "cond", "prox"), show="headings")
        for key, title, width, anc in [
            ("sem", "Semaforo", 110, "center"),
            ("dias", "Dias", 70, "center"),
            ("nome", "Nome", 290, "w"),
            ("tel", "Telefone", 180, "w"),
            ("cond", "Condicoes", 120, "center"),
            ("prox", "Proximo", 110, "center"),
        ]:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=width, anchor=anc)
        self.tree.tag_configure("vencido", background="#FDECEA")
        self.tree.tag_configure("vermelho", background="#FCE4D6")
        self.tree.tag_configure("amarelo", background="#FFF4CC")
        self.tree.tag_configure("verde", background="#EAF7EA")
        self.tree.tag_configure("sem_data", background="#F2F2F2")
        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        y_scroll = ttk.Scrollbar(body, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=y_scroll.set)
        y_scroll.grid(row=0, column=0, sticky="nse")

        right = tk.Frame(body, bg="#FFFFFF", bd=1, relief="solid")
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(1, weight=1)

        row = 0
        for label, var in [
            ("Nome", self.name_var),
            ("ID", self.id_var),
            ("Telefone", self.phone_var),
            ("Condicoes", self.cond_var),
        ]:
            tk.Label(right, text=f"{label}:", bg="#FFFFFF", anchor="w", font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="w", padx=10, pady=(10 if row == 0 else 6, 0))
            tk.Label(right, textvariable=var, bg="#FFFFFF", anchor="w").grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(10 if row == 0 else 6, 0))
            row += 1

        tk.Label(right, text="Ultimo atendimento medico (dd/mm/aaaa):", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Entry(right, textvariable=self.med_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(10, 0))
        row += 1

        tk.Label(right, text="Ultimo atendimento enfermagem (dd/mm/aaaa):", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        tk.Entry(right, textvariable=self.enf_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Ultima medicao antropometrica (dd/mm/aaaa):", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        tk.Entry(right, textvariable=self.antrop_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Ultima afericao de sinais vitais (dd/mm/aaaa):", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        tk.Entry(right, textvariable=self.sinais_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Aprazar com base em:", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        base_cb = ttk.Combobox(right, textvariable=self.base_mode_var, values=BASE_MODES, state="readonly")
        base_cb.grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Intervalo (meses):", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        tk.Entry(right, textvariable=self.interval_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Proximo aprazamento (dd/mm/aaaa):", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        tk.Entry(right, textvariable=self.next_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Observacao:", bg="#FFFFFF", anchor="w").grid(row=row, column=0, sticky="w", padx=10, pady=(6, 0))
        tk.Entry(right, textvariable=self.notes_var).grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(6, 0))
        row += 1

        tk.Label(right, text="Base usada:", bg="#FFFFFF", anchor="w", font=("Segoe UI", 9, "bold")).grid(row=row, column=0, sticky="w", padx=10, pady=(10, 0))
        tk.Label(right, textvariable=self.base_info_var, bg="#FFFFFF", anchor="w").grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(10, 0))
        row += 1
        tk.Label(right, text="Dias para vencer:", bg="#FFFFFF", anchor="w", font=("Segoe UI", 9, "bold")).grid(row=row, column=0, sticky="w", padx=10, pady=(4, 0))
        tk.Label(right, textvariable=self.days_var, bg="#FFFFFF", anchor="w").grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(4, 0))
        row += 1
        tk.Label(right, text="Semaforo:", bg="#FFFFFF", anchor="w", font=("Segoe UI", 9, "bold")).grid(row=row, column=0, sticky="w", padx=10, pady=(4, 0))
        tk.Label(right, textvariable=self.sem_var, bg="#FFFFFF", anchor="w").grid(row=row, column=1, sticky="we", padx=(0, 10), pady=(4, 0))
        row += 1

        actions = tk.Frame(right, bg="#FFFFFF")
        actions.grid(row=row, column=0, columnspan=2, sticky="we", padx=10, pady=(12, 10))
        tk.Button(actions, text="Hoje medico", command=self._set_today_medico).pack(side="left")
        tk.Button(actions, text="Hoje enfermagem", command=self._set_today_enfermagem).pack(side="left", padx=(6, 0))
        tk.Button(actions, text="Recalcular automatico", command=self._recalc_selected).pack(side="left", padx=(6, 0))
        tk.Button(actions, text="Reaprazar manual", command=self._manual_reaprazar).pack(side="left", padx=(6, 0))
        tk.Button(actions, text="Salvar paciente", command=self._save_selected).pack(side="right")

        bottom = tk.Frame(self, bg="#EEF4F8")
        bottom.pack(fill="x", padx=12, pady=(0, 10))
        tk.Label(bottom, textvariable=self.files_var, bg="#EEF4F8", fg="#1F4E79").pack(side="left")
        tk.Label(bottom, textvariable=self.status_var, bg="#EEF4F8", anchor="e").pack(side="right")

    def _choose_folder(self):
        d = filedialog.askdirectory(initialdir=self.base_dir_var.get().strip() or str(Path.home()))
        if d:
            self.base_dir_var.set(d)
            self._load_existing_store()
            self._apply_filter()

    def _load_existing_store(self):
        path = self._store_path()
        self.records, meta = load_store_with_meta(path)
        self.source_files = meta.get("source_files", {}) if isinstance(meta, dict) else {}
        self._load_settings_from_payload(meta.get("settings") if isinstance(meta, dict) else None)
        _months, days = self._interval_from_settings()
        self.global_interval_days.set(days)
        for rec in self.records.values():
            rec.update(compute_control_fields(rec))
        if self.source_files:
            self.files_var.set("Arquivos: " + " | ".join(f"{k}:{Path(v).name}" for k, v in sorted(self.source_files.items())))
        else:
            self.files_var.set("Arquivos: -")
        self.status_var.set(f"Base carregada: {len(self.records)} pacientes.")

    def _open_settings_window(self):
        if self._settings_window and self._settings_window.winfo_exists():
            self._settings_window.lift()
            self._settings_window.focus_force()
            return
        win = tk.Toplevel(self)
        self._settings_window = win
        win.title("Configuracao do aprazamento")
        win.geometry("560x320")
        win.transient(self)
        win.grab_set()

        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=10, pady=10)
        tab = tk.Frame(nb, bg="#F7FAFD")
        nb.add(tab, text="Regras")

        row = 0
        tk.Label(tab, text="Prazo:", bg="#F7FAFD").grid(row=row, column=0, sticky="w", padx=12, pady=(14, 6))
        preset_cb = ttk.Combobox(
            tab,
            textvariable=self.settings_preset_var,
            values=tuple(INTERVAL_PRESETS.keys()),
            state="readonly",
            width=24,
        )
        preset_cb.grid(row=row, column=1, sticky="w", padx=(0, 12), pady=(14, 6))
        preset_cb.bind("<<ComboboxSelected>>", lambda _e: self._refresh_settings_custom_state())
        row += 1

        tk.Label(tab, text="Tempo livre:", bg="#F7FAFD").grid(row=row, column=0, sticky="w", padx=12, pady=6)
        self._settings_custom_value_entry = tk.Entry(tab, textvariable=self.settings_custom_value_var, width=10)
        self._settings_custom_value_entry.grid(row=row, column=1, sticky="w", padx=(0, 6), pady=6)
        self._settings_custom_unit_combo = ttk.Combobox(
            tab,
            textvariable=self.settings_custom_unit_var,
            values=("MESES", "DIAS"),
            state="readonly",
            width=10,
        )
        self._settings_custom_unit_combo.grid(row=row, column=1, sticky="w", padx=(84, 12), pady=6)
        row += 1

        tk.Label(tab, text="Base da proxima consulta:", bg="#F7FAFD").grid(row=row, column=0, sticky="w", padx=12, pady=6)
        base_cb = ttk.Combobox(tab, textvariable=self.settings_base_mode_var, values=BASE_MODES, state="readonly", width=24)
        base_cb.grid(row=row, column=1, sticky="w", padx=(0, 12), pady=6)
        row += 1

        help_text = (
            "AMBOS_MAIS_PROXIMO: usa a data que faz vencer primeiro.\n"
            "AMBOS_MAIS_LONGE: usa a data que joga o vencimento mais para frente."
        )
        tk.Label(tab, text=help_text, bg="#F7FAFD", fg="#355C7D", justify="left").grid(row=row, column=0, columnspan=2, sticky="w", padx=12, pady=(8, 12))
        row += 1

        actions = tk.Frame(tab, bg="#F7FAFD")
        actions.grid(row=row, column=0, columnspan=2, sticky="e", padx=12, pady=(0, 12))
        tk.Button(actions, text="Aplicar e fechar", command=self._save_settings_and_apply).pack(side="right")
        tk.Button(actions, text="Cancelar", command=win.destroy).pack(side="right", padx=(0, 8))

        tab.columnconfigure(1, weight=1)
        self._refresh_settings_custom_state()

    def _save_settings_and_apply(self):
        months, days = self._interval_from_settings()
        base_mode = str(self.settings_base_mode_var.get() or "ENFERMAGEM").upper()
        if base_mode not in BASE_MODES:
            messagebox.showerror("Configuracao", "Base invalida para calculo.")
            return
        if months <= 0 and days <= 0:
            messagebox.showerror("Configuracao", "Defina um prazo valido.")
            return
        self._apply_global_settings_to_records()
        self._save_store_clicked(silent=True)
        self._apply_filter()
        self.status_var.set(f"Configuracao aplicada | Base: {base_mode} | Prazo: {days} dia(s)" if days > 0 else f"Configuracao aplicada | Base: {base_mode} | Prazo: {months} mes(es)")
        if self._settings_window and self._settings_window.winfo_exists():
            self._settings_window.destroy()
            self._settings_window = None

    def _import_folder(self, silent_errors: bool = False):
        folder = Path(self.base_dir_var.get().strip())
        if not folder.exists():
            if not silent_errors:
                messagebox.showerror("Erro", f"Pasta nao encontrada: {folder}")
            return
        default_months, default_days = self._interval_from_settings()
        default_base_mode = str(self.settings_base_mode_var.get() or "ENFERMAGEM").upper()
        if default_base_mode not in BASE_MODES:
            default_base_mode = "ENFERMAGEM"
        try:
            records, source_files = build_records_from_folder(
                folder,
                existing_records=self.records,
                default_base_mode=default_base_mode,
                default_interval_months=default_months,
                default_interval_days=default_days,
            )
        except Exception as exc:
            if not silent_errors:
                messagebox.showerror("Erro ao importar", str(exc))
            else:
                self.status_var.set("Autoimportacao nao executada: C1..C7 nao encontrados.")
            return
        self.records = records
        self.source_files = source_files
        self.files_var.set("Arquivos: " + " | ".join(f"{k}:{Path(v).name}" for k, v in sorted(source_files.items())))
        self._apply_filter()
        if silent_errors:
            self.status_var.set(f"Autoimportacao concluida: {len(self.records)} pacientes.")
        else:
            self.status_var.set(f"Importacao concluida: {len(self.records)} pacientes.")
        self._save_store_clicked(silent=True)

    def _save_store_clicked(self, silent: bool = False):
        folder = Path(self.base_dir_var.get().strip())
        folder.mkdir(parents=True, exist_ok=True)
        path = self._store_path()
        try:
            save_store(path, self.records, source_files=self.source_files, settings=self._current_settings_payload())
        except Exception as exc:
            messagebox.showerror("Erro ao salvar", str(exc))
            return
        if not silent:
            self.status_var.set(f"Base salva em {path.name}.")

    def _record_matches_filter(self, rec: dict, filt: FilterState) -> bool:
        if filt.semaphore != "TODOS" and str(rec.get("semaphore")) != filt.semaphore:
            return False
        if filt.condition != "TODAS" and filt.condition not in set(rec.get("conditions", [])):
            return False
        term = filt.term
        if not term:
            return True
        hay = " ".join([
            str(rec.get("name", "")),
            str(rec.get("cpf", "")),
            str(rec.get("cns", "")),
            str(rec.get("phone", "")),
            " ".join(rec.get("conditions", [])),
        ]).lower()
        return term in hay

    def _apply_filter(self):
        filt = FilterState(
            term=self.search_var.get().strip().lower(),
            semaphore=self.semaphore_filter_var.get().strip() or "TODOS",
            condition=self.condition_filter_var.get().strip() or "TODAS",
        )
        for iid in self.tree.get_children(""):
            self.tree.delete(iid)

        rows = list(self.records.values())
        rows.sort(key=lambda r: (_priority_order(str(r.get("semaphore"))), int(r.get("days_to_due")) if str(r.get("days_to_due", "")).strip() not in {"", "None"} else 999999, _norm_spaces(r.get("name"))))
        self.filtered_ids = []
        for rec in rows:
            if not self._record_matches_filter(rec, filt):
                continue
            rid = rec["id"]
            self.filtered_ids.append(rid)
            sem = str(rec.get("semaphore", ""))
            tag = {
                "VENCIDO": "vencido",
                "VERMELHO": "vermelho",
                "AMARELO": "amarelo",
                "VERDE": "verde",
                "SEM DATA": "sem_data",
            }.get(sem, "sem_data")
            conds = " ".join(rec.get("conditions", []))
            self.tree.insert(
                "",
                "end",
                iid=rid,
                values=(
                    sem,
                    rec.get("days_to_due", ""),
                    rec.get("name", ""),
                    rec.get("phone", ""),
                    conds,
                    rec.get("next_date", ""),
                ),
                tags=(tag,),
            )
        self.count_var.set(f"{len(self.filtered_ids)} pacientes")

        if self.current_id and self.tree.exists(self.current_id):
            self.tree.selection_set(self.current_id)
            self.tree.focus(self.current_id)
        elif self.filtered_ids:
            self.current_id = self.filtered_ids[0]
            self.tree.selection_set(self.current_id)
            self.tree.focus(self.current_id)
            self._load_selected_into_form()
        else:
            self.current_id = None
            self._clear_form()

    def _clear_form(self):
        self.name_var.set("-")
        self.id_var.set("-")
        self.phone_var.set("-")
        self.cond_var.set("-")
        self.med_var.set("")
        self.enf_var.set("")
        self.antrop_var.set("")
        self.sinais_var.set("")
        self.base_mode_var.set("ENFERMAGEM")
        self.interval_var.set(str(DEFAULT_INTERVAL_MONTHS))
        self.next_var.set("")
        self.days_var.set("")
        self.sem_var.set("")
        self.base_info_var.set("")
        self.notes_var.set("")

    def _focus_record_in_list(self, record_id: str, search_term: str = ""):
        self.semaphore_filter_var.set("TODOS")
        self.condition_filter_var.set("TODAS")
        self.search_var.set(search_term)
        self.current_id = record_id
        self._apply_filter()
        if self.tree.exists(record_id):
            self.tree.selection_set(record_id)
            self.tree.focus(record_id)
            self.tree.see(record_id)
            self.current_id = record_id
            self._load_selected_into_form()

    def _add_patient_dialog(self):
        cfg_months, cfg_days = self._interval_from_settings()
        cfg_base_mode = str(self.settings_base_mode_var.get() or "ENFERMAGEM").upper()
        if cfg_base_mode not in BASE_MODES:
            cfg_base_mode = "ENFERMAGEM"

        win = tk.Toplevel(self)
        win.title("Adicionar paciente")
        win.geometry("560x430")
        win.transient(self)
        win.grab_set()

        name_var = tk.StringVar()
        cpf_var = tk.StringVar()
        cns_var = tk.StringVar()
        phone_var = tk.StringVar()
        med_var = tk.StringVar()
        enf_var = tk.StringVar()
        antrop_var = tk.StringVar()
        sinais_var = tk.StringVar()
        base_mode_var = tk.StringVar(value=cfg_base_mode)
        interval_var = tk.StringVar(value=str(cfg_months))
        notes_var = tk.StringVar()

        form = tk.Frame(win, bg="#F7FAFD")
        form.pack(fill="both", expand=True, padx=12, pady=12)
        form.columnconfigure(1, weight=1)

        fields = [
            ("Nome *", name_var),
            ("CPF", cpf_var),
            ("CNS", cns_var),
            ("Telefone", phone_var),
            ("Ultimo atendimento medico (dd/mm/aaaa)", med_var),
            ("Ultimo atendimento enfermagem (dd/mm/aaaa)", enf_var),
            ("Ultima medicao antropometrica (dd/mm/aaaa)", antrop_var),
            ("Ultima afericao de sinais vitais (dd/mm/aaaa)", sinais_var),
        ]
        row = 0
        for label, var in fields:
            tk.Label(form, text=label, bg="#F7FAFD", anchor="w").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=6)
            tk.Entry(form, textvariable=var).grid(row=row, column=1, sticky="we", pady=6)
            row += 1

        tk.Label(form, text="Aprazar com base em", bg="#F7FAFD", anchor="w").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Combobox(form, textvariable=base_mode_var, values=BASE_MODES, state="readonly").grid(row=row, column=1, sticky="we", pady=6)
        row += 1

        tk.Label(form, text="Intervalo (meses)", bg="#F7FAFD", anchor="w").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=6)
        tk.Entry(form, textvariable=interval_var).grid(row=row, column=1, sticky="we", pady=6)
        row += 1

        tk.Label(form, text="Observacao", bg="#F7FAFD", anchor="w").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=6)
        tk.Entry(form, textvariable=notes_var).grid(row=row, column=1, sticky="we", pady=6)

        def _submit_new_patient():
            name = re.sub(r"\s+", " ", str(name_var.get())).strip()
            cpf = _clean_id(cpf_var.get())
            cns = _clean_id(cns_var.get())
            phone = str(phone_var.get() or "").strip()
            med_txt = str(med_var.get() or "").strip()
            enf_txt = str(enf_var.get() or "").strip()
            antrop_txt = str(antrop_var.get() or "").strip()
            sinais_txt = str(sinais_var.get() or "").strip()
            notes = str(notes_var.get() or "").strip()

            if not name:
                messagebox.showwarning("Campo obrigatorio", "Informe o nome do paciente.", parent=win)
                return

            med_date = parse_date(med_txt)
            enf_date = parse_date(enf_txt)
            antrop_date = parse_date(antrop_txt)
            sinais_date = parse_date(sinais_txt)
            if med_txt and med_date is None:
                messagebox.showerror("Data invalida", "Ultimo atendimento medico invalido. Use dd/mm/aaaa.", parent=win)
                return
            if enf_txt and enf_date is None:
                messagebox.showerror("Data invalida", "Ultimo atendimento enfermagem invalido. Use dd/mm/aaaa.", parent=win)
                return
            if antrop_txt and antrop_date is None:
                messagebox.showerror("Data invalida", "Ultima medicao antropometrica invalida. Use dd/mm/aaaa.", parent=win)
                return
            if sinais_txt and sinais_date is None:
                messagebox.showerror("Data invalida", "Ultima afericao de sinais vitais invalida. Use dd/mm/aaaa.", parent=win)
                return

            base_mode = str(base_mode_var.get() or cfg_base_mode).upper().strip()
            if base_mode not in BASE_MODES:
                base_mode = cfg_base_mode
            try:
                interval_months = int(str(interval_var.get() or "").strip() or cfg_months)
            except Exception:
                messagebox.showerror("Intervalo invalido", "Informe um numero inteiro de meses.", parent=win)
                return
            if interval_months <= 0:
                messagebox.showerror("Intervalo invalido", "Intervalo deve ser maior que zero.", parent=win)
                return

            existing_id = find_existing_record_id(self.records, name=name, cpf=cpf, cns=cns)
            if existing_id:
                self._focus_record_in_list(existing_id, search_term=name)
                self.status_var.set("Paciente ja existe. Registro localizado na busca.")
                win.destroy()
                return

            rec = {
                "id": patient_key(name, cpf, cns),
                "name": name,
                "cpf": cpf,
                "cns": cns,
                "phone": phone,
                "conditions": [],
                "last_medico_date": fmt_date(med_date),
                "last_enfermagem_date": fmt_date(enf_date),
                "last_antropometria_date": fmt_date(antrop_date),
                "last_sinais_vitais_date": fmt_date(sinais_date),
                "estimated_medico": False,
                "estimated_enfermagem": False,
                "base_mode": base_mode,
                "interval_months": interval_months,
                "interval_days": cfg_days,
                "manual_next_date": "",
                "notes": notes,
            }
            rec.update(compute_control_fields(rec))
            self.records[rec["id"]] = rec
            self._save_store_clicked(silent=True)
            self._focus_record_in_list(rec["id"], search_term=name)
            self.status_var.set(f"Paciente adicionado: {name}")
            win.destroy()

        actions = tk.Frame(form, bg="#F7FAFD")
        actions.grid(row=row + 1, column=0, columnspan=2, sticky="e", pady=(14, 0))
        tk.Button(actions, text="Cancelar", command=win.destroy).pack(side="right")
        tk.Button(actions, text="Salvar paciente", command=_submit_new_patient).pack(side="right", padx=(0, 8))

    def _on_select(self, _evt=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.current_id = sel[0]
        self._load_selected_into_form()

    def _load_selected_into_form(self):
        if not self.current_id:
            self._clear_form()
            return
        rec = self.records.get(self.current_id)
        if not rec:
            self._clear_form()
            return
        self.name_var.set(rec.get("name", ""))
        pid = rec.get("cpf") or rec.get("cns") or rec.get("id", "")
        self.id_var.set(pid)
        self.phone_var.set(rec.get("phone", ""))
        self.cond_var.set(" ".join(rec.get("conditions", [])) or "-")
        self.med_var.set(rec.get("last_medico_date", ""))
        self.enf_var.set(rec.get("last_enfermagem_date", ""))
        self.antrop_var.set(rec.get("last_antropometria_date", ""))
        self.sinais_var.set(rec.get("last_sinais_vitais_date", ""))
        mode_val = str(rec.get("base_mode") or "ENFERMAGEM").upper()
        if mode_val not in BASE_MODES:
            mode_val = "ENFERMAGEM"
        self.base_mode_var.set(mode_val)
        self.interval_var.set(str(rec.get("interval_months") or DEFAULT_INTERVAL_MONTHS))
        self.next_var.set(rec.get("next_date", ""))
        self.days_var.set(str(rec.get("days_to_due", "")))
        self.sem_var.set(rec.get("semaphore", ""))
        self.base_info_var.set(f"{rec.get('base_source', '-')} | {rec.get('base_date', '')}")
        self.notes_var.set(rec.get("notes", ""))

    def _selected_record(self) -> dict | None:
        if not self.current_id:
            return None
        return self.records.get(self.current_id)

    def _set_today_medico(self):
        self.med_var.set(date.today().strftime("%d/%m/%Y"))
        self._recalc_selected()

    def _set_today_enfermagem(self):
        self.enf_var.set(date.today().strftime("%d/%m/%Y"))
        self._recalc_selected()

    def _apply_form_to_record(self, rec: dict, clear_manual: bool) -> None:
        med_txt = self.med_var.get().strip()
        enf_txt = self.enf_var.get().strip()
        antrop_txt = self.antrop_var.get().strip()
        sinais_txt = self.sinais_var.get().strip()

        med = parse_date(med_txt)
        enf = parse_date(enf_txt)
        antrop = parse_date(antrop_txt)
        sinais = parse_date(sinais_txt)

        if med_txt and med is None:
            raise ValueError("Data invalida em ultimo atendimento medico. Use dd/mm/aaaa.")
        if enf_txt and enf is None:
            raise ValueError("Data invalida em ultimo atendimento enfermagem. Use dd/mm/aaaa.")
        if antrop_txt and antrop is None:
            raise ValueError("Data invalida em ultima medicao antropometrica. Use dd/mm/aaaa.")
        if sinais_txt and sinais is None:
            raise ValueError("Data invalida em ultima afericao de sinais vitais. Use dd/mm/aaaa.")

        interval_txt = self.interval_var.get().strip()
        if not interval_txt:
            interval = DEFAULT_INTERVAL_MONTHS
        else:
            try:
                interval = int(interval_txt)
            except Exception:
                raise ValueError("Intervalo invalido. Informe numero inteiro de meses.")
            if interval <= 0:
                raise ValueError("Intervalo invalido. Use valor maior que zero.")

        rec["last_medico_date"] = fmt_date(med)
        rec["last_enfermagem_date"] = fmt_date(enf)
        rec["last_antropometria_date"] = fmt_date(antrop)
        rec["last_sinais_vitais_date"] = fmt_date(sinais)
        rec["estimated_medico"] = False if med else rec.get("estimated_medico", False)
        rec["estimated_enfermagem"] = False if enf else rec.get("estimated_enfermagem", False)
        rec["base_mode"] = str(self.base_mode_var.get() or "ENFERMAGEM").upper().strip()
        rec["interval_months"] = interval
        _cfg_months, cfg_days = self._interval_from_settings()
        rec["interval_days"] = int(cfg_days or 0)
        rec["notes"] = self.notes_var.get().strip()
        if clear_manual:
            rec["manual_next_date"] = ""
        rec.update(compute_control_fields(rec))

    def _recalc_selected(self):
        rec = self._selected_record()
        if not rec:
            return
        try:
            self._apply_form_to_record(rec, clear_manual=True)
        except Exception as exc:
            messagebox.showerror("Erro de validacao", str(exc))
            return
        self._apply_filter()
        self.current_id = rec["id"]
        self._load_selected_into_form()
        self.status_var.set(f"Recalculado: {rec.get('name', '')}")

    def _manual_reaprazar(self):
        rec = self._selected_record()
        if not rec:
            return
        manual = parse_date(self.next_var.get())
        if not manual:
            messagebox.showerror("Data invalida", "Informe a data de proximo aprazamento em dd/mm/aaaa.")
            return
        try:
            self._apply_form_to_record(rec, clear_manual=False)
        except Exception as exc:
            messagebox.showerror("Erro de validacao", str(exc))
            return
        rec["manual_next_date"] = fmt_date(manual)
        rec.update(compute_control_fields(rec))
        self._apply_filter()
        self.current_id = rec["id"]
        self._load_selected_into_form()
        self.status_var.set(f"Reaprazamento manual aplicado: {rec.get('name', '')}")

    def _save_selected(self):
        rec = self._selected_record()
        if not rec:
            return
        try:
            self._apply_form_to_record(rec, clear_manual=False)
        except Exception as exc:
            messagebox.showerror("Erro de validacao", str(exc))
            return
        self._save_store_clicked(silent=True)
        self._apply_filter()
        self.current_id = rec["id"]
        self._load_selected_into_form()
        self.status_var.set(f"Paciente salvo: {rec.get('name', '')}")


def launch_aprazamento(master=None, base_dir: Path | None = None, auto_import: bool = False):
    return AprazamentoApp(master=master, base_dir=base_dir, auto_import=auto_import)


def main():
    root = tk.Tk()
    root.withdraw()
    win = AprazamentoApp(master=root)
    win.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()


if __name__ == "__main__":
    main()
