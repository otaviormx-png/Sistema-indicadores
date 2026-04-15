from __future__ import annotations

import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aps_utils import read_esus_table, to_numeric


@dataclass(frozen=True)
class Quadrimester:
    year: int
    code: str
    start: date
    end: date


def _norm(text: str) -> str:
    return unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii").lower()


def _pick_col(df: pd.DataFrame, token: str, required_tokens: list[str] | None = None) -> str | None:
    required_tokens = required_tokens or []
    token_n = _norm(token)
    req_n = [_norm(t) for t in required_tokens]
    for col in df.columns:
        c_n = _norm(col)
        if token_n in c_n and all(t in c_n for t in req_n):
            return col
    return None


def _extract_report_date(path: Path) -> date:
    date_re = re.compile(r"(\d{2}/\d{2}/\d{4})")
    with path.open("r", encoding="utf-8-sig", errors="ignore") as f:
        for line in f:
            if "Gerado em" in line:
                m = date_re.search(line)
                if m:
                    return datetime.strptime(m.group(1), "%d/%m/%Y").date()
    stem_re = re.search(r"(\d{2})[._-](\d{2})[._-](\d{4})", path.stem)
    if stem_re:
        try:
            dd, mm, yyyy = [int(x) for x in stem_re.groups()]
            return date(yyyy, mm, dd)
        except Exception:
            pass
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).date()
    except Exception:
        return date.today()


def quadrimester_of(d: date) -> Quadrimester:
    if d.month <= 4:
        return Quadrimester(d.year, f"{d.year}-Q1", date(d.year, 1, 1), date(d.year, 4, 30))
    if d.month <= 8:
        return Quadrimester(d.year, f"{d.year}-Q2", date(d.year, 5, 1), date(d.year, 8, 31))
    return Quadrimester(d.year, f"{d.year}-Q3", date(d.year, 9, 1), date(d.year, 12, 31))


def next_quadrimester(q: Quadrimester) -> Quadrimester:
    if q.code.endswith("Q1"):
        return Quadrimester(q.year, f"{q.year}-Q2", date(q.year, 5, 1), date(q.year, 8, 31))
    if q.code.endswith("Q2"):
        return Quadrimester(q.year, f"{q.year}-Q3", date(q.year, 9, 1), date(q.year, 12, 31))
    ny = q.year + 1
    return Quadrimester(ny, f"{ny}-Q1", date(ny, 1, 1), date(ny, 4, 30))


def _valid_days(val) -> float | None:
    n = to_numeric(val, float("nan"))
    if pd.isna(n) or n < 0 or n > 5000:
        return None
    return float(n)


def _build_phone(row: pd.Series, cols: list[str]) -> str:
    vals = []
    for c in cols:
        v = str(row.get(c, "")).strip()
        if v and v not in {"-", "nan", "None"}:
            vals.append(v)
    return " | ".join(vals) if vals else "-"


def _semaphore(days_to_deadline: int) -> str:
    if days_to_deadline < 0:
        return "VENCIDO"
    if days_to_deadline <= 15:
        return "VERMELHO"
    if days_to_deadline <= 45:
        return "AMARELO"
    return "VERDE"


def _priority_key(sem: str) -> int:
    order = {"VENCIDO": 0, "VERMELHO": 1, "AMARELO": 2, "VERDE": 3}
    return order.get(sem, 9)


def build_control_dataframe(input_path: str | Path, ref_date: date | None = None) -> pd.DataFrame:
    path = Path(input_path)
    df = read_esus_table(path)

    col_name = _pick_col(df, "nome")
    col_cns = _pick_col(df, "cns")
    col_cpf = _pick_col(df, "cpf")
    col_tel1 = _pick_col(df, "telefone celular")
    col_tel2 = _pick_col(df, "telefone residencial")
    col_tel3 = _pick_col(df, "telefone de contato")
    col_med = _pick_col(df, "dias desde o ultimo atendimento", ["medico"])
    col_enf = _pick_col(df, "dias desde o ultimo atendimento", ["enfermagem"])

    if not col_name or not col_med or not col_enf:
        raise ValueError("Arquivo nao possui colunas esperadas de Nome e dias de atendimento medico/enfermagem.")

    report_date = ref_date or _extract_report_date(path)
    current_q = quadrimester_of(report_date)
    rows: list[dict] = []

    for _, row in df.iterrows():
        name = str(row.get(col_name, "")).strip()
        if not name or name == "-":
            continue

        med_days = _valid_days(row.get(col_med))
        enf_days = _valid_days(row.get(col_enf))

        ref_used = "-"
        best_days: float | None = None
        if med_days is not None and enf_days is not None:
            if med_days <= enf_days:
                best_days = med_days
                ref_used = "MEDICO"
            else:
                best_days = enf_days
                ref_used = "ENFERMAGEM"
        elif med_days is not None:
            best_days = med_days
            ref_used = "MEDICO"
        elif enf_days is not None:
            best_days = enf_days
            ref_used = "ENFERMAGEM"

        if best_days is None:
            last_dt = None
            covered_q = "-"
            next_required = current_q.code
            deadline = current_q.end
            status = "SEM DADO DE ATENDIMENTO"
        else:
            last_dt = report_date - timedelta(days=int(best_days))
            covered_q_obj = quadrimester_of(last_dt)
            covered_q = covered_q_obj.code
            if covered_q_obj.code == current_q.code:
                nq = next_quadrimester(current_q)
                next_required = nq.code
                deadline = nq.end
                status = "COBERTO NO QUADRIMESTRE ATUAL"
            else:
                next_required = current_q.code
                deadline = current_q.end
                status = "PRECISA ATENDER NO QUADRIMESTRE ATUAL"

        days_to_deadline = (deadline - report_date).days
        sem = _semaphore(days_to_deadline)

        rows.append(
            {
                "Nome": name,
                "CNS": str(row.get(col_cns, "")).strip() if col_cns else "",
                "CPF": str(row.get(col_cpf, "")).strip() if col_cpf else "",
                "Telefone": _build_phone(row, [c for c in [col_tel1, col_tel2, col_tel3] if c]),
                "Dias desde atendimento medico": int(med_days) if med_days is not None else "",
                "Dias desde atendimento enfermagem": int(enf_days) if enf_days is not None else "",
                "Referencia usada": ref_used,
                "Dias desde ultimo atendimento (controle)": int(best_days) if best_days is not None else "",
                "Data estimada ultimo atendimento": last_dt.strftime("%d/%m/%Y") if last_dt else "",
                "Quadrimestre coberto": covered_q,
                "Proximo quadrimestre obrigatorio": next_required,
                "Data limite": deadline.strftime("%d/%m/%Y"),
                "Dias para limite": days_to_deadline,
                "Semaforo": sem,
                "Situacao": status,
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["_prio"] = out["Semaforo"].map(_priority_key)
    out = out.sort_values(["_prio", "Dias para limite", "Nome"], ascending=[True, True, True]).drop(columns=["_prio"])
    return out


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


def _style_sheet(ws, header_row: int = 1):
    header_fill = _fill("D9E1F2")
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True, color="1F4E79")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(str(cell.value)) + 2, 12), 38)

    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _apply_semaphore_color(ws, sem_col_name: str):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        sem_col = headers.index(sem_col_name) + 1
    except ValueError:
        return

    colors = {
        "VERDE": ("E2F0D9", "1B5E20"),
        "AMARELO": ("FFF2CC", "7F6000"),
        "VERMELHO": ("FCE4D6", "9C0006"),
        "VENCIDO": ("F8CBAD", "9C0006"),
    }
    for r in range(2, ws.max_row + 1):
        val = str(ws.cell(r, sem_col).value or "").upper().strip()
        if val not in colors:
            continue
        bg, fg = colors[val]
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = _fill(bg)
            ws.cell(r, c).font = Font(color=fg, bold=(c == sem_col))


def export_control_workbook(df: pd.DataFrame, output_path: str | Path, report_date: date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle Geral"

    if df.empty:
        ws["A1"] = "Sem dados para gerar controle."
        wb.save(output_path)
        return

    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(1, c_idx, col)
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, col in enumerate(df.columns, 1):
            ws.cell(r_idx, c_idx, row[col])
    _style_sheet(ws)
    _apply_semaphore_color(ws, "Semaforo")

    ws2 = wb.create_sheet("Busca Ativa")
    busca = df[
        (df["Situacao"] != "COBERTO NO QUADRIMESTRE ATUAL")
        | (df["Semaforo"].isin(["VENCIDO", "VERMELHO", "AMARELO"]))
    ].copy()
    for c_idx, col in enumerate(busca.columns, 1):
        ws2.cell(1, c_idx, col)
    for r_idx, (_, row) in enumerate(busca.iterrows(), 2):
        for c_idx, col in enumerate(busca.columns, 1):
            ws2.cell(r_idx, c_idx, row[col])
    _style_sheet(ws2)
    _apply_semaphore_color(ws2, "Semaforo")

    ws3 = wb.create_sheet("Resumo")
    ws3["A1"] = "Data de referencia"
    ws3["B1"] = report_date.strftime("%d/%m/%Y")
    ws3["A3"] = "Semaforo"
    ws3["B3"] = "Quantidade"
    sem_counts = df["Semaforo"].value_counts()
    row = 4
    for sem in ["VENCIDO", "VERMELHO", "AMARELO", "VERDE"]:
        ws3.cell(row, 1, sem)
        ws3.cell(row, 2, int(sem_counts.get(sem, 0)))
        row += 1
    ws3["D3"] = "Situacao"
    ws3["E3"] = "Quantidade"
    sit_counts = df["Situacao"].value_counts()
    row = 4
    for sit in ["PRECISA ATENDER NO QUADRIMESTRE ATUAL", "COBERTO NO QUADRIMESTRE ATUAL", "SEM DADO DE ATENDIMENTO"]:
        ws3.cell(row, 4, sit)
        ws3.cell(row, 5, int(sit_counts.get(sit, 0)))
        row += 1
    _style_sheet(ws3, header_row=3)
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["D"].width = 46
    ws3.column_dimensions["E"].width = 14

    wb.save(output_path)


def build_output_name(input_path: Path) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"controle_quadrimestral_{input_path.stem}_{stamp}.xlsx"


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Uso: python controle_quadrimestral.py <entrada.csv/xlsx> [saida.xlsx]")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else input_path.parent / build_output_name(input_path)
    report_date = _extract_report_date(input_path)
    df = build_control_dataframe(input_path, ref_date=report_date)
    export_control_workbook(df, output_path, report_date)
    print(f"Controle gerado: {output_path}")


if __name__ == "__main__":
    main()
