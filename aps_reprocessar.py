"""
aps_reprocessar.py — Reprocessamento parcial.

Carrega um xlsx já gerado e identifica pacientes ainda em busca ativa
(pontuação < 100). Cruza com um novo arquivo bruto e reprocessa apenas
esses pacientes, preservando os já concluídos.

Uso via interface: botão "Reprocessar pendentes" no histórico.
Uso via CLI: python aps_cli.py --reprocessar <xlsx_anterior> --entrada <pasta_brutos>
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Callable

import openpyxl
import pandas as pd

from aps_utils import (
    normalize_text, read_esus_table, render_workbook,
    candidate_file_for_indicator,
)
from sistema_aps import INDICADORES, desktop_files


def _read_completos(xlsx_path: Path) -> set[str]:
    """
    Lê o xlsx anterior e retorna o conjunto de CNS/CPF
    dos pacientes que já atingiram 100 pontos (não precisam ser reprocessados).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]

    def _idx(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_pts = _idx("Pontuação")
    idx_cns = _idx("CNS")
    idx_cpf = _idx("CPF")
    idx_nome = _idx("Nome")

    completos: set[str] = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        try:
            pts = float(row[idx_pts]) if idx_pts is not None and row[idx_pts] else 0
        except (TypeError, ValueError):
            pts = 0
        if pts >= 100:
            chave = None
            if idx_cns is not None and row[idx_cns]:
                chave = str(row[idx_cns]).strip()
            elif idx_cpf is not None and row[idx_cpf]:
                chave = str(row[idx_cpf]).strip()
            elif idx_nome is not None and row[idx_nome]:
                chave = normalize_text(str(row[idx_nome]).strip())
            if chave:
                completos.add(chave)
    wb.close()
    return completos


def _chave_paciente(row: pd.Series) -> str | None:
    """Extrai a chave de identificação única do paciente (CNS > CPF > nome)."""
    for col in ["CNS", "Cartão SUS"]:
        v = str(row.get(col, "")).strip()
        if v and v not in {"-", "nan", "None", ""}:
            return v
    for col in ["CPF"]:
        v = str(row.get(col, "")).strip()
        if v and v not in {"-", "nan", "None", ""}:
            return v
    for col in ["Nome", "Nome do cidadão", "Paciente"]:
        v = str(row.get(col, "")).strip()
        if v and v not in {"-", "nan", "None", ""}:
            return normalize_text(v)
    return None


def reprocessar_pendentes(
    xlsx_anterior: Path,
    in_dir: Path,
    out_dir: Path,
    log: Callable[[str], None] | None = None,
) -> dict:
    """
    Reprocessa apenas os pacientes pendentes (pontuação < 100) de um xlsx anterior.

    Retorna dict com:
      - saida: Path do novo xlsx
      - completos_preservados: int
      - reprocessados: int
      - novos_completos: int
    """
    _log = log or print
    _log(f"Reprocessamento parcial: {xlsx_anterior.name}")

    # Descobre o indicador pelo nome do arquivo
    stem = xlsx_anterior.stem
    code = stem.split("_")[0].upper()
    cfg_func = next(((c, f) for c, f in INDICADORES if c.code == code), None)
    if cfg_func is None:
        raise ValueError(
            f"Indicador '{code}' não encontrado. "
            f"Verifique se o arquivo é um resultado gerado pela APS Suite."
        )
    cfg, func = cfg_func

    # CNS/CPF dos pacientes já completos
    completos_anteriores = _read_completos(xlsx_anterior)
    _log(f"  Pacientes já completos (preservados): {len(completos_anteriores)}")

    # Acha o arquivo bruto correspondente
    files = desktop_files(in_dir)
    entrada = candidate_file_for_indicator(files, cfg)
    if not entrada:
        raise FileNotFoundError(
            f"Nenhum arquivo bruto encontrado para {code} em: {in_dir}"
        )
    _log(f"  Bruto encontrado: {entrada.name}")

    # Lê o bruto completo
    df_raw = read_esus_table(entrada)

    # Filtra: mantém apenas os que NÃO estão na lista de completos
    def _is_pendente(row):
        chave = _chave_paciente(row)
        if chave is None:
            return True
        return chave not in completos_anteriores

    mascara = df_raw.apply(_is_pendente, axis=1)
    df_pendentes = df_raw[mascara].copy()
    df_completos_raw = df_raw[~mascara].copy()

    _log(f"  Pendentes para reprocessar: {len(df_pendentes)}")
    _log(f"  Já completos (não reprocessados): {len(df_completos_raw)}")

    if df_pendentes.empty:
        _log("  Todos os pacientes já estão completos. Nada a reprocessar.")
        return {
            "saida": xlsx_anterior,
            "completos_preservados": len(completos_anteriores),
            "reprocessados": 0,
            "novos_completos": 0,
        }

    # Processa apenas os pendentes
    df_resultado_pendentes = cfg.builder(df_pendentes)
    novos_completos = int((df_resultado_pendentes["Pontuação"] >= 100).sum())

    # Salva novo xlsx (só com os pendentes reprocessados)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saida = out_dir / f"{code}_repro_{stamp}.xlsx"
    render_workbook(df_resultado_pendentes, cfg, saida)

    _log(f"  Novos completos após reprocessamento: {novos_completos}")
    _log(f"  Resultado salvo em: {saida.name}")

    return {
        "saida": saida,
        "completos_preservados": len(completos_anteriores),
        "reprocessados": len(df_pendentes),
        "novos_completos": novos_completos,
    }
