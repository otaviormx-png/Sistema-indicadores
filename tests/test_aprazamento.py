from __future__ import annotations

import shutil
from contextlib import contextmanager
from datetime import date
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

from aps_aprazamento import add_months, build_records_from_folder, compute_control_fields, find_existing_record_id, patient_key


TMP_ROOT = Path(__file__).resolve().parent / "_tmp_runtime"
TMP_ROOT.mkdir(parents=True, exist_ok=True)


@contextmanager
def local_tmpdir():
    path = TMP_ROOT / uuid4().hex
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def test_add_months_edge_cases():
    assert add_months(date(2026, 4, 13), 4) == date(2026, 8, 13)
    assert add_months(date(2026, 1, 31), 1) == date(2026, 2, 28)
    assert add_months(date(2024, 1, 31), 1) == date(2024, 2, 29)


def test_compute_control_fields_base_selection():
    rec = {
        "last_medico_date": "10/04/2026",
        "last_enfermagem_date": "13/04/2026",
        "base_mode": "ENFERMAGEM",
        "interval_months": 4,
        "manual_next_date": "",
    }
    out = compute_control_fields(rec, ref_date=date(2026, 4, 13))
    assert out["base_source"] == "ENFERMAGEM"
    assert out["base_date"] == "13/04/2026"
    assert out["next_date"] == "13/08/2026"
    assert out["days_to_due"] == 122
    assert out["semaphore"] == "VERDE"

    rec2 = {**rec, "base_mode": "MEDICO"}
    out2 = compute_control_fields(rec2, ref_date=date(2026, 4, 13))
    assert out2["base_source"] == "MEDICO"
    assert out2["next_date"] == "10/08/2026"


def test_compute_control_fields_manual_next():
    rec = {
        "last_medico_date": "13/04/2026",
        "last_enfermagem_date": "",
        "base_mode": "MEDICO",
        "interval_months": 4,
        "manual_next_date": "20/04/2026",
    }
    out = compute_control_fields(rec, ref_date=date(2026, 4, 13))
    assert out["next_date"] == "20/04/2026"
    assert out["days_to_due"] == 7
    assert out["semaphore"] == "VERMELHO"


def test_compute_control_fields_global_days_interval():
    rec = {
        "last_medico_date": "",
        "last_enfermagem_date": "13/04/2026",
        "base_mode": "ENFERMAGEM",
        "interval_months": 4,
        "interval_days": 180,
        "manual_next_date": "",
    }
    out = compute_control_fields(rec, ref_date=date(2026, 4, 13))
    assert out["base_source"] == "ENFERMAGEM"
    assert out["next_date"] == "10/10/2026"
    assert out["days_to_due"] == 180
    assert out["semaphore"] == "VERDE"


def test_compute_control_fields_ambos_prioridade():
    rec = {
        "last_medico_date": "01/01/2026",
        "last_enfermagem_date": "01/04/2026",
        "interval_months": 4,
        "interval_days": 0,
        "manual_next_date": "",
    }
    out_prox = compute_control_fields({**rec, "base_mode": "AMBOS_MAIS_PROXIMO"}, ref_date=date(2026, 4, 13))
    assert out_prox["base_source"] == "MEDICO"
    assert out_prox["next_date"] == "01/05/2026"

    out_longe = compute_control_fields({**rec, "base_mode": "AMBOS_MAIS_LONGE"}, ref_date=date(2026, 4, 13))
    assert out_longe["base_source"] == "ENFERMAGEM"
    assert out_longe["next_date"] == "01/08/2026"


def test_patient_key_priority():
    assert patient_key("A", "123", "999").startswith("CPF:")
    assert patient_key("A", "", "999").startswith("CNS:")
    assert patient_key("Maria da Silva", "", "").startswith("NOME:")


def test_find_existing_record_id_dedup_by_id_or_name():
    rid_cpf = patient_key("Ana Clara", "123", "")
    rid_nome = patient_key("Joaquim Martins", "", "")
    records = {
        rid_cpf: {"id": rid_cpf, "name": "Ana Clara", "cpf": "123", "cns": ""},
        rid_nome: {"id": rid_nome, "name": "Joaquim Martins", "cpf": "", "cns": ""},
    }
    assert find_existing_record_id(records, name="Outro Nome", cpf="123") == rid_cpf
    assert find_existing_record_id(records, name="joaquim    martins") == rid_nome
    assert find_existing_record_id(records, name="Paciente Novo") is None


def _write_indicator(path: Path, rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Cx"
    headers = [
        "Nome",
        "Telefone celular",
        "CPF",
        "CNS",
        "Dias desde o ultimo atendimento medico",
        "Dias desde o ultimo atendimento de enfermagem",
        "Meses desde o ultimo atendimento medico",
        "Meses desde o ultimo atendimento de enfermagem",
        "Data da ultima medicao de peso e altura",
        "Data da ultima medicao de pressao arterial",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    r = 4
    for row in rows:
        ws.cell(r, 1, row.get("Nome", ""))
        ws.cell(r, 2, row.get("Telefone celular", ""))
        ws.cell(r, 3, row.get("CPF", ""))
        ws.cell(r, 4, row.get("CNS", ""))
        ws.cell(r, 5, row.get("Dias desde o ultimo atendimento medico", ""))
        ws.cell(r, 6, row.get("Dias desde o ultimo atendimento de enfermagem", ""))
        ws.cell(r, 7, row.get("Meses desde o ultimo atendimento medico", ""))
        ws.cell(r, 8, row.get("Meses desde o ultimo atendimento de enfermagem", ""))
        ws.cell(r, 9, row.get("Data da ultima medicao de peso e altura", ""))
        ws.cell(r, 10, row.get("Data da ultima medicao de pressao arterial", ""))
        r += 1
    wb.save(path)


def test_build_records_from_folder_conditions():
    with local_tmpdir() as td:
        c1 = td / f"C1_{uuid4().hex}.xlsx"
        c5 = td / f"C5_{uuid4().hex}.xlsx"
        _write_indicator(
            c1,
            [
                {
                    "Nome": "Joaquim Martins",
                    "Telefone celular": "21999990000",
                    "CPF": "111",
                    "CNS": "999",
                    "Meses desde o ultimo atendimento medico": "1",
                    "Meses desde o ultimo atendimento de enfermagem": "0",
                    "Data da ultima medicao de peso e altura": "20/03/2026",
                    "Data da ultima medicao de pressao arterial": "20/03/2026",
                }
            ],
        )
        _write_indicator(
            c5,
            [
                {
                    "Nome": "Joaquim Martins",
                    "Telefone celular": "21999990000",
                    "CPF": "111",
                    "CNS": "999",
                    "Meses desde o ultimo atendimento medico": "1",
                    "Meses desde o ultimo atendimento de enfermagem": "0",
                }
            ],
        )

        records, files = build_records_from_folder(td, ref_date=date(2026, 4, 13))
        rec = next(iter(records.values()))
        assert "C5" in rec["conditions"]
        assert rec["base_mode"] == "ENFERMAGEM"
        assert rec["last_antropometria_date"] == "20/03/2026"
        assert rec["last_sinais_vitais_date"] == "20/03/2026"
        assert rec["next_date"] != ""
        assert "C1" in files and "C5" in files


def test_build_records_combines_months_and_days():
    with local_tmpdir() as td:
        c1 = td / f"C1_{uuid4().hex}.xlsx"
        _write_indicator(
            c1,
            [
                {
                    "Nome": "Paciente Dias",
                    "Telefone celular": "21900000000",
                    "CPF": "222",
                    "CNS": "888",
                    "Dias desde o ultimo atendimento medico": "15",
                    "Meses desde o ultimo atendimento medico": "3",
                    "Dias desde o ultimo atendimento de enfermagem": "40",
                    "Meses desde o ultimo atendimento de enfermagem": "6",
                }
            ],
        )
        records, _files = build_records_from_folder(td, ref_date=date(2026, 4, 13))
        rec = next(iter(records.values()))
        assert rec["last_medico_date"] == "29/12/2025"
        assert rec["last_enfermagem_date"] == "03/09/2025"


def test_build_records_does_not_use_pa_as_enfermagem_attendance():
    with local_tmpdir() as td:
        c1 = td / f"C1_{uuid4().hex}.xlsx"
        _write_indicator(
            c1,
            [
                {
                    "Nome": "Paciente PA",
                    "Telefone celular": "21911110000",
                    "CPF": "333",
                    "CNS": "777",
                    "Meses desde o ultimo atendimento medico": "2",
                    "Meses desde o ultimo atendimento de enfermagem": "",
                    "Data da ultima medicao de pressao arterial": "01/04/2026",
                }
            ],
        )
        records, _files = build_records_from_folder(td, ref_date=date(2026, 4, 13))
        rec = next(iter(records.values()))
        assert rec["last_medico_date"] != ""
        assert rec["last_enfermagem_date"] == ""
