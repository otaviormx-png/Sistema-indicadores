from __future__ import annotations

import shutil
from contextlib import contextmanager
from pathlib import Path
from uuid import uuid4

import openpyxl

from aps_clonador_interativo import clone_interactive, refresh_interactive_workbook


TMP_ROOT = Path(__file__).resolve().parent / "_tmp_runtime"
TMP_ROOT.mkdir(parents=True, exist_ok=True)


@contextmanager
def local_tmpdir():
    path = TMP_ROOT / uuid4().hex
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def _find_sheet(wb, starts):
    for name in wb.sheetnames:
        if any(prefix.lower() in name.lower() for prefix in starts):
            return wb[name]
    raise AssertionError(f"Nao encontrou aba com prefixos {starts}")


def _make_base_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "📋 Dados C1"

    headers = [
        "Nome",
        "Microárea",
        "Telefone celular",
        "Telefone residencial",
        "A - Consulta em dia",
        "B - Visita em dia",
    ]
    for c, value in enumerate(headers, 1):
        ws.cell(3, c, value)

    rows = [
        ["Carlos", "003", "333", "", "SIM", "SIM"],
        ["Bruno", "002", "222", "", "", ""],
        ["Ana", "001", "111", "", "SIM", ""],
    ]
    for r, row in enumerate(rows, 4):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)

    wb.save(path)
    return path


def test_clone_interactive_cria_busca_e_resumo_estaticos():
    with local_tmpdir() as tmp:
        entrada = _make_base_workbook(tmp / "entrada.xlsx")
        saida = clone_interactive(entrada)

        wb = openpyxl.load_workbook(saida)
        ws_busca = _find_sheet(wb, ("busca",))
        ws_resumo = _find_sheet(wb, ("resumo",))

        assert ws_busca["B5"].value == "Bruno"
        assert "URGENTE" in str(ws_busca["A5"].value).upper()
        assert ws_busca["B6"].value == "Ana"
        assert "ALTA" in str(ws_busca["A6"].value).upper()
        assert ws_busca["B7"].value == "Carlos"
        assert "CONCL" in str(ws_busca["A7"].value).upper()

        assert "RESUMO" in str(ws_resumo["A1"].value).upper()
        assert "TIMO" in str(ws_resumo["A10"].value).upper()


def test_refresh_interactive_workbook_recalcula_pontuacao():
    with local_tmpdir() as tmp:
        path = _make_base_workbook(tmp / "base.xlsx")
        clone = clone_interactive(path)

        wb = openpyxl.load_workbook(clone)
        ws = _find_sheet(wb, ("dados",))
        ws["F5"] = "SIM"
        wb.save(clone)
        wb.close()

        refresh_interactive_workbook(clone)

        wb2 = openpyxl.load_workbook(clone)
        ws2 = _find_sheet(wb2, ("dados",))
        assert ws2[5][6].value == 50
        assert "SUFICIENTE" in str(ws2[5][7].value).upper()
        busca = _find_sheet(wb2, ("busca",))
        nomes = [busca[f"B{r}"].value for r in (5, 6, 7)]
        assert "Bruno" in nomes
