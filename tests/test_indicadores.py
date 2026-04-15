from __future__ import annotations

import shutil
from contextlib import contextmanager
from pathlib import Path
from uuid import uuid4

import openpyxl
import pandas as pd


TMP_ROOT = Path(__file__).resolve().parent / "_tmp_runtime"
TMP_ROOT.mkdir(parents=True, exist_ok=True)


@contextmanager
def local_tmpdir():
    path = TMP_ROOT / uuid4().hex
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


BASE = {
    "Nome": "Paciente Teste",
    "Data de nascimento": "01/01/1980",
    "Idade": "44 anos",
    "Sexo": "Feminino",
    "Raca/cor": "Branca",
    "Microarea": "001",
    "Rua": "Rua das Flores",
    "Numero": "100",
    "Complemento": "",
    "Bairro": "Centro",
    "Telefone celular": "19999990000",
    "Telefone residencial": "",
    "Telefone de contato": "",
    "CPF": "000.000.000-00",
    "CNS": "123456789012345",
    "Meses desde o ultimo atendimento medico": "3",
    "Meses desde o ultimo atendimento de enfermagem": "3",
    "Meses desde o ultimo atendimento odontologico": "10",
    "Meses desde a ultima visita domiciliar": "6",
    "Ultima medicao de peso": "65",
    "Ultima medicao de altura": "165",
    "Data da ultima medicao de peso e altura": "01/01/2024",
    "Ultima medicao de pressao arterial": "120/80",
    "Data da ultima medicao de pressao arterial": "01/01/2024",
    "Ultimas visitas domiciliares": "01/01/2024",
    "Quantidade de visitas domiciliares": "3",
}


def make_csv(rows, base_dir: Path) -> Path:
    if isinstance(rows, dict):
        rows = [rows]
    out = base_dir / f"{uuid4().hex}.csv"
    pd.DataFrame(rows).to_csv(out, index=False, sep=";", encoding="utf-8-sig")
    return out


def read_result(saida: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(saida)
    ws = wb.worksheets[0]
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True)]
    return pd.DataFrame(data)


def run(processar_fn, rows) -> pd.DataFrame:
    with local_tmpdir() as td:
        entrada = make_csv(rows, td)
        saida = td / "saida.xlsx"
        processar_fn(entrada, saida)
        return read_result(saida)


def pick_value(row, *options):
    for name in options:
        if name in row.index:
            return row[name]
    raise KeyError(options[0])


class TestC1:
    def setup_method(self):
        from c1_mais_acesso import processar

        self.p = processar

    def test_pontuacao_maxima(self):
        assert int(run(self.p, BASE)["Pontuação"].iloc[0]) == 100

    def test_pontuacao_zero(self):
        row = {
            **BASE,
            "Meses desde o ultimo atendimento medico": "99",
            "Meses desde o ultimo atendimento de enfermagem": "99",
            "Meses desde o ultimo atendimento odontologico": "99",
            "Meses desde a ultima visita domiciliar": "99",
            "Microarea": "",
        }
        assert int(run(self.p, row)["Pontuação"].iloc[0]) == 0

    def test_linha_sem_nome_ignorada(self):
        assert len(run(self.p, {**BASE, "Nome": ""})) == 0

    def test_classificacao_otimo(self):
        assert run(self.p, BASE)["Classificação"].iloc[0] == "Ótimo"

    def test_multiplos_pacientes(self):
        rows = [{**BASE, "Nome": f"Paciente {i}"} for i in range(5)]
        assert len(run(self.p, rows)) == 5

    def test_converte_dias_ou_meses_em_data_estimada(self):
        row = {
            **BASE,
            "Dias desde o ultimo atendimento medico": "15",
            "Meses desde o ultimo atendimento medico": "3",
        }
        out = run(self.p, row).iloc[0]
        ref = pd.Timestamp.now().normalize()
        esperado = (ref - pd.DateOffset(months=3) - pd.Timedelta(days=15)).strftime("%d/%m/%Y")
        assert out["Data estimada do ultimo atendimento medico"] == esperado


class TestC1O:
    def setup_method(self):
        from c1_oficial import processar

        self.p = processar

    def test_formula_oficial_por_colunas_agregadas(self):
        row = {
            **BASE,
            "Nome": "Linha agregada",
            "Atendimentos por demanda programada": "30",
            "Atendimentos por todos os tipos de demandas": "100",
        }
        out = run(self.p, row).iloc[0]
        assert out["Nome"] == "AGREGADO C1 OFICIAL"
        assert float(out["Percentual C1 oficial (%)"]) == 30.0
        assert int(float(pick_value(out, "PontuaÃ§Ã£o", "Pontuação"))) == 30

    def test_formula_oficial_por_tipo_demanda(self):
        rows = [
            {**BASE, "Nome": "A", "Tipo de demanda": "consulta agendada programada"},
            {**BASE, "Nome": "B", "Tipo de demanda": "cuidado continuado"},
            {**BASE, "Nome": "C", "Tipo de demanda": "consulta no dia"},
        ]
        out = run(self.p, rows).iloc[0]
        assert float(out["Numerador C1 (demanda programada)"]) == 2.0
        assert float(out["Denominador C1 (todas as demandas)"]) == 3.0
        assert float(out["Percentual C1 oficial (%)"]) == 66.67


class TestC4:
    def setup_method(self):
        from c4_diabetes import processar

        self.p = processar

    def _completo(self):
        return {
            **BASE,
            "Meses desde o ultimo atendimento medico": "3",
            "Ultima medicao de pressao arterial": "120/80",
            "Ultima medicao de peso": "70",
            "Ultima medicao de altura": "170",
            "Quantidade de visitas domiciliares": "3",
            "Hemoglobina glicada": "5.8",
            "Data da avaliacao dos pes": "01/01/2024",
        }

    def test_pontuacao_maxima(self):
        assert int(run(self.p, self._completo())["Pontuação"].iloc[0]) == 100

    def test_sem_hb_glicada(self):
        assert int(run(self.p, {**self._completo(), "Hemoglobina glicada": ""})["Pontuação"].iloc[0]) == 85

    def test_sem_pes(self):
        assert int(run(self.p, {**self._completo(), "Data da avaliacao dos pes": ""})["Pontuação"].iloc[0]) == 85


class TestC5:
    def setup_method(self):
        from c5_hipertensao import processar

        self.p = processar

    def _completo(self):
        return {
            **BASE,
            "Meses desde o ultimo atendimento medico": "4",
            "Ultima medicao de pressao arterial": "130/85",
            "Ultima medicao de peso": "72",
            "Ultima medicao de altura": "168",
            "Quantidade de visitas domiciliares": "2",
        }

    def test_pontuacao_maxima(self):
        assert int(run(self.p, self._completo())["Pontuação"].iloc[0]) == 100

    def test_sem_pa(self):
        assert int(run(self.p, {**self._completo(), "Ultima medicao de pressao arterial": ""})["Pontuação"].iloc[0]) == 75

    def test_eap76_isenta_visita_domiciliar(self):
        row = {
            **self._completo(),
            "Tipo de equipe": "76",
            "Quantidade de visitas domiciliares": "0",
        }
        assert int(run(self.p, row)["Pontuação"].iloc[0]) == 100


class TestC7:
    def setup_method(self):
        from c7_mulher import processar

        self.p = processar

    def test_fora_faixa_nao_vira_sim_automatico(self):
        row = {
            **BASE,
            "Idade": "10 anos",
            "HPV": "",
            "Data da ultima consulta de saude sexual e reprodutiva": "",
        }
        out = run(self.p, row).iloc[0]
        assert int(out["Pontuação"]) == 0
        assert out["A - Colo uterino (25-64a / 36m)"] == "N/A"
        assert out["D - Mamografia (50-69a / 24m)"] == "N/A"

    def test_criterio_unico_aplicavel_normaliza_100(self):
        row = {
            **BASE,
            "Idade": "10 anos",
            "HPV": "2026-01-01",
            "Data da ultima consulta de saude sexual e reprodutiva": "",
        }
        out = run(self.p, row).iloc[0]
        assert int(out["Pontuação"]) == 100


class TestSistema:
    def test_csv_encoding_latin1(self):
        from c1_mais_acesso import processar

        with local_tmpdir() as td:
            entrada = td / "c1.csv"
            pd.DataFrame([BASE]).to_csv(entrada, index=False, sep=";", encoding="latin1")
            saida = td / "saida.xlsx"
            processar(entrada, saida)
            assert saida.exists()

    def test_linhas_vazias_ignoradas(self):
        from c1_mais_acesso import processar

        with local_tmpdir() as td:
            entrada = td / "c1.csv"
            pd.DataFrame([BASE, {k: "" for k in BASE}]).to_csv(
                entrada, index=False, sep=";", encoding="utf-8-sig"
            )
            saida = td / "saida.xlsx"
            processar(entrada, saida)
            assert len(read_result(saida)) == 1

    def test_erro_isolado_entre_indicadores(self):
        import sistema_aps

        with local_tmpdir() as td:
            row = {
                **BASE,
                "Meses desde o ultimo atendimento medico": "4",
                "Ultima medicao de pressao arterial": "130/85",
                "Ultima medicao de peso": "72",
                "Ultima medicao de altura": "168",
                "Quantidade de visitas domiciliares": "2",
            }
            pd.DataFrame([row]).to_csv(
                td / "c5.csv", index=False, sep=";", encoding="utf-8-sig"
            )
            (td / "c1.csv").write_bytes(b"\xff\xfe" + b"\x00" * 100)

            original_markers = list(sistema_aps.IGNORE_MARKERS)
            sistema_aps.IGNORE_MARKERS = []
            try:
                res = sistema_aps.process_selected(["C1", "C5"], in_dir=td, out_dir=td / "out")
            finally:
                sistema_aps.IGNORE_MARKERS = original_markers
            status = {r["code"]: r["status"] for r in res}

            assert status.get("C5") == "ok"
            assert status.get("C1") in ("erro", "não encontrado")

    def test_build_base_row_sem_dataframe_fantasma(self):
        import inspect
        from aps_utils import build_base_row

        assert "pd.DataFrame(columns=row.index)" not in inspect.getsource(build_base_row)

    def test_config_carrega(self):
        import aps_config

        assert aps_config.cores()["azul_escuro"] == "1F4E79"
        assert "resultado" in aps_config.ignorar_marcadores()
