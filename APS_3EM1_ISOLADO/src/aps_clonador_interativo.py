from __future__ import annotations

import os
import re
import shutil
import getpass
import json
import threading
import unicodedata
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from aps_utils import infer_indicator_code_from_path

# -----------------------------------------------------------------------------
# Estilos
# -----------------------------------------------------------------------------

VERDE = PatternFill("solid", fgColor="C6EFCE")
VERMELHO = PatternFill("solid", fgColor="FFC7CE")
AMARELO = PatternFill("solid", fgColor="FFEB9C")
AZUL = PatternFill("solid", fgColor="D9EAF7")


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _font(*, bold: bool = False, color: str = "000000", size: int = 9) -> Font:
    return Font(name="Segoe UI", bold=bold, color=color, size=size)


def _align(h: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _border(style: str = "thin", color: str = "D0D0D0") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _norm_header_text(value) -> str:
    txt = str(value or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"[^a-z0-9]+", " ", txt).strip()
    return txt


# -----------------------------------------------------------------------------
# Helpers da planilha
# -----------------------------------------------------------------------------


def _detect_header(ws) -> int:
    for r in range(1, min(10, ws.max_row) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, min(ws.max_column, 20) + 1)]
        if "nome" in vals and any(v in vals for v in ["microÃ¡rea", "microarea", "telefone celular", "telefone", "bairro", "endereco", "endereço"]):
            return r
    return 3



def _detect_columns(ws, header_row: int) -> dict:
    headers = {}
    for c in range(1, ws.max_column + 1):
        value = ws.cell(header_row, c).value
        if value:
            headers[str(value).strip()] = c
    headers_norm = {_norm_header_text(k): v for k, v in headers.items()}

    def pick(*names):
        for name in names:
            if name in headers:
                return headers[name]
        for name in names:
            n = _norm_header_text(name)
            if n in headers_norm:
                return headers_norm[n]
        return None

    def pick_contains(*patterns):
        pats = [_norm_header_text(p) for p in patterns]
        for title, col in headers.items():
            normalized = _norm_header_text(title)
            if any(p in normalized for p in pats):
                return col
        return None

    criterios = []
    for title, col in headers.items():
        if re.match(r"^[A-Za-z]\s*[-â€“]\s+\S", title):
            criterios.append((title.split("-", 1)[0].strip().upper()[:1], title, col))
    criterios.sort(key=lambda x: x[0])

    return {
        "headers": headers,
        "criterios": criterios,
        "nome": pick("Nome", "Paciente") or 1,
        "microarea": pick("MicroÃ¡rea", "Microarea") or 6,
        "bairro": pick("Bairro", "Bairro/Localidade", "Localidade", "Comunidade") or pick_contains("bairro", "localidade", "comunidade"),
        "endereco": pick("EndereÃ§o", "Endereco", "Logradouro", "Endereço", "Rua", "Domicilio") or pick_contains("endereco", "ender", "logradouro", "domicilio", "rua", "avenida"),
        "numero": pick("Numero", "NÃºmero", "NÃºm", "Num") or pick_contains("numero", "num"),
        "complemento": pick("Complemento", "Comp") or pick_contains("complemento", "comp"),
        "cidade": pick("Cidade", "MunicÃ­pio", "Municipio") or pick_contains("cidade", "municipio"),
        "uf": pick("UF", "Estado") or pick_contains("uf", "estado"),
        "cep": pick("CEP", "Codigo Postal") or pick_contains("cep", "codigo postal"),
        "tel1": pick("Telefone celular", "Telefone") or 11,
        "tel2": pick("Telefone residencial", "Telefone de contato") or 12,
        "pontuacao": pick("Pontuação", "Pontuacao", "PontuaÃ§Ã£o") or pick_contains("pontu"),
        "classif": pick("Classificação", "Classificacao", "ClassificaÃ§Ã£o") or pick_contains("classif"),
        "prioridade": pick("Prioridade"),
        "pendencias": pick("Pendências", "Pendencias", "PendÃªncias") or pick_contains("pend"),
    }



def _ensure_support_columns(ws, header_row: int, cols: dict) -> dict:
    last_col = ws.max_column
    for key, title in [
        ("pontuacao", "Pontuacao"),
        ("classif", "Classificacao"),
        ("prioridade", "Prioridade"),
        ("pendencias", "Pendencias"),
    ]:
        if cols.get(key):
            continue
        last_col += 1
        ws.cell(header_row, last_col, title)
        cols[key] = last_col
        cols["headers"][title] = last_col
        c = ws.cell(header_row, last_col)
        c.fill = _fill("BDD7EE")
        c.font = _font(bold=True, color="1F4E79")
        c.alignment = _align("center", wrap=True)
        c.border = _border()
    return cols



def _normalize_status(value) -> str:
    txt = str(value or "").strip().upper()
    if txt in {"NAO", "NÃO", "NÃƒO"}:
        return "NAO"
    return txt



def _weights(criterios) -> dict[str, int]:
    n = len(criterios)
    if not n:
        return {}
    base = 100 // n
    extra = 100 - (base * n)
    out = {}
    for i, (letter, _title, _col) in enumerate(criterios):
        out[letter] = base + (1 if i < extra else 0)
    return out



def _score_row(ws, row: int, criterios, pesos: dict[str, int]) -> int:
    total = 0
    for letter, _title, col in criterios:
        if _normalize_status(ws.cell(row, col).value) == "SIM":
            total += pesos[letter]
    return total



def _classify(score: float) -> str:
    if score >= 100:
        return "Ã“timo"
    if score >= 75:
        return "Bom"
    if score >= 50:
        return "Suficiente"
    return "Regular"



def _priority(score: float) -> str:
    if score >= 100:
        return "ðŸŸ¢ CONCLUÃDO"
    if score >= 75:
        return "ðŸŸ¡ MONITORAR"
    if score >= 50:
        return "ðŸŸ  ALTA"
    return "ðŸ”´ URGENTE"



def _pending_text(ws, row: int, criterios) -> str:
    parts = []
    for _letter, title, col in criterios:
        if _normalize_status(ws.cell(row, col).value) != "SIM":
            parts.append(title.split("-", 1)[1].strip() if "-" in title else title)
    return "; ".join(parts)


def _apply_status_style(cell, value):
    status = _normalize_status(value)
    if status == "SIM":
        cell.fill = _fill("EAF7EA")
        cell.font = _font(bold=True, color="2E7D32")
    elif status == "NAO":
        cell.fill = _fill("FDECEA")
        cell.font = _font(bold=True, color="C62828")
    elif status == "PENDENTE":
        cell.fill = _fill("FFFBE6")
        cell.font = _font(bold=True, color="9E7D00")
    else:
        cell.fill = PatternFill(fill_type=None)
        cell.font = _font()
    cell.alignment = _align("center")
    cell.border = _border()

def _compose_endereco(parts: dict) -> str:
    main = str(parts.get("endereco", "") or "").strip()
    numero = str(parts.get("numero", "") or "").strip()
    comp = str(parts.get("complemento", "") or "").strip()
    bairro = str(parts.get("bairro", "") or "").strip()
    cidade = str(parts.get("cidade", "") or "").strip()
    uf = str(parts.get("uf", "") or "").strip()
    cep = str(parts.get("cep", "") or "").strip()
    pieces = []
    if main:
        pieces.append(main)
    if numero:
        pieces.append(f"n {numero}")
    if comp:
        pieces.append(comp)
    if bairro:
        pieces.append(bairro)
    loc = " - ".join([p for p in [cidade, uf] if p])
    if loc:
        pieces.append(loc)
    if cep:
        pieces.append(f"CEP {cep}")
    return ", ".join(pieces).strip()

def _patients_from_data(ws, header_row: int, cols: dict) -> list[dict]:
    pesos = _weights(cols["criterios"])
    start = header_row + 1
    out = []
    for row in range(start, ws.max_row + 1):
        nome = ws.cell(row, cols["nome"]).value
        if not nome:
            continue
        pts = _score_row(ws, row, cols["criterios"], pesos)
        rec = {
            "row": row,
            "nome": str(nome),
            "bairro": (ws.cell(row, cols["bairro"]).value or "") if cols.get("bairro") else "",
            "endereco": (ws.cell(row, cols["endereco"]).value or "") if cols.get("endereco") else "",
            "tel": ws.cell(row, cols["tel1"]).value or ws.cell(row, cols["tel2"]).value or "",
            "pts": pts,
            "classif": _classify(pts),
            "prio": _priority(pts),
            "pend": _pending_text(ws, row, cols["criterios"]),
            "statuses": [ws.cell(row, col).value or "" for _l, _t, col in cols["criterios"]],
        }
        rec["endereco_full"] = _compose_endereco({
            "endereco": ws.cell(row, cols["endereco"]).value if cols.get("endereco") else "",
            "numero": ws.cell(row, cols["numero"]).value if cols.get("numero") else "",
            "complemento": ws.cell(row, cols["complemento"]).value if cols.get("complemento") else "",
            "bairro": rec["bairro"],
            "cidade": ws.cell(row, cols["cidade"]).value if cols.get("cidade") else "",
            "uf": ws.cell(row, cols["uf"]).value if cols.get("uf") else "",
            "cep": ws.cell(row, cols["cep"]).value if cols.get("cep") else "",
        })
        out.append(rec)
    out.sort(key=lambda x: (x["pts"], x["nome"].lower()))
    return out



def _update_data_sheet(ws, header_row: int, cols: dict) -> list[dict]:
    patients = _patients_from_data(ws, header_row, cols)
    by_row = {p["row"]: p for p in patients}
    for row, rec in by_row.items():
        for _letter, _title, col in cols["criterios"]:
            _apply_status_style(ws.cell(row, col), ws.cell(row, col).value)
        ws.cell(row, cols["pontuacao"]).value = rec["pts"]
        ws.cell(row, cols["classif"]).value = rec["classif"]
        ws.cell(row, cols["prioridade"]).value = rec["prio"]
        ws.cell(row, cols["pendencias"]).value = rec["pend"]
    return patients



def _clear_sheet(ws):
    for rng in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(rng))
        except Exception:
            pass
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = _font()
            cell.alignment = _align()
            cell.border = Border()



def _merge_title(ws, cell_range: str, text: str, bg: str, fg: str = "FFFFFF", size: int = 12):
    try:
        ws.merge_cells(cell_range)
    except Exception:
        pass
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = _fill(bg)
    c.font = _font(bold=True, color=fg, size=size)
    c.alignment = _align("center", wrap=True)
    c.border = _border("medium", "808080")



def _build_search_sheet(ws, cols: dict, patients: list[dict]):
    _clear_sheet(ws)
    criterios = cols["criterios"]
    visible_cols = 4 + len(criterios) + 3
    hidden_row_col = visible_cols + 1
    ws.sheet_view.showGridLines = False

    _merge_title(ws, f"A1:{get_column_letter(visible_cols)}1", "ðŸ” BUSCA ATIVA â€” Editor APS", "1F4E79", size=13)
    ws.merge_cells(f"A2:{get_column_letter(visible_cols)}2")
    ws["A2"] = f"Total: {len(patients)} | Atualizada em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].fill = AZUL
    ws["A2"].font = _font(color="1F4E79")
    ws["A2"].alignment = _align("center")

    headers = ["Prioridade", "Nome", "Bairro", "Telefone"] + [t for _l, t, _c in criterios] + ["PontuaÃ§Ã£o", "ClassificaÃ§Ã£o", "PendÃªncias"]
    for c_idx, title in enumerate(headers, 1):
        cell = ws.cell(4, c_idx, title)
        if c_idx <= 4:
            cell.fill = _fill("1F4E79")
            cell.font = _font(bold=True, color="FFFFFF")
        elif c_idx <= 4 + len(criterios):
            cell.fill = _fill("FFF4CC")
            cell.font = _font(bold=True, color="7F6000")
        else:
            cell.fill = _fill("DFF0D8")
            cell.font = _font(bold=True, color="2E7D32")
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()

    palette = {
        "ðŸ”´ URGENTE": ("FDECEA", "C62828"),
        "ðŸŸ  ALTA": ("FFF4E5", "EF6C00"),
        "ðŸŸ¡ MONITORAR": ("FFFBE6", "9E7D00"),
        "ðŸŸ¢ CONCLUÃDO": ("EAF7EA", "2E7D32"),
    }

    for out_row, p in enumerate(patients, 5):
        bg, fg = palette[p["prio"]]
        values = [p["prio"], p["nome"], p.get("bairro", ""), p["tel"], *p["statuses"], p["pts"], p["classif"], p["pend"]]
        for c_idx, value in enumerate(values, 1):
            cell = ws.cell(out_row, c_idx, value)
            cell.border = _border()
            cell.alignment = _align("left" if c_idx in (2, visible_cols) else "center", wrap=(c_idx == visible_cols))
            cell.fill = _fill(bg)
            cell.font = _font(color=fg, bold=(c_idx == 1))
        ws.cell(out_row, hidden_row_col, p["row"])  # linha real na aba Dados
        ws.row_dimensions[out_row].height = 20

    widths = [18, 34, 12, 16] + [14] * len(criterios) + [11, 14, 46]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.column_dimensions[get_column_letter(hidden_row_col)].hidden = True
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(visible_cols)}4"



def _build_summary_sheet(ws, cols: dict, patients: list[dict], code: str):
    _clear_sheet(ws)
    ws.sheet_view.showGridLines = False

    total = len(patients)
    media = sum(p["pts"] for p in patients) / total if total else 0
    concluidos = sum(1 for p in patients if p["prio"] == "ðŸŸ¢ CONCLUÃDO")
    pendentes = total - concluidos
    class_counts = {"Ã“timo": 0, "Bom": 0, "Suficiente": 0, "Regular": 0}
    for p in patients:
        class_counts[p["classif"]] = class_counts.get(p["classif"], 0) + 1

    _merge_title(ws, "A1:H1", f"ðŸ“Š RESUMO â€” {code}", "1F4E79", size=13)
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Atualizado automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].fill = AZUL
    ws["A2"].alignment = _align("center")
    ws["A2"].font = _font(color="1F4E79")

    cards = [
        ("A4:B6", f"ðŸ‘¥ TOTAL\n{total}", "D9EAF7", "1F4E79"),
        ("C4:D6", f"ðŸ“Š MÃ‰DIA\n{media:.1f}", "EFE3FF", "6A1B9A"),
        ("E4:F6", f"âœ… CONCLUÃDOS\n{concluidos}", "EAF7EA", "2E7D32"),
        ("G4:H6", f"ðŸ” PENDENTES\n{pendentes}", "FDECEA", "C62828"),
    ]
    for rng, text, bg, fg in cards:
        _merge_title(ws, rng, text, bg, fg, 14)

    _merge_title(ws, "A8:H8", "DISTRIBUIÃ‡ÃƒO POR CLASSIFICAÃ‡ÃƒO", "7030A0")
    for idx, title in enumerate(["ClassificaÃ§Ã£o", "Qtd", "%", "Progresso"], 1):
        c = ws.cell(9, idx, title)
        c.fill = _fill("1F4E79")
        c.font = _font(bold=True, color="FFFFFF")
        c.alignment = _align("center")
        c.border = _border()

    class_palette = {
        "Ã“timo": ("EAF7EA", "2E7D32"),
        "Bom": ("FFFBE6", "9E7D00"),
        "Suficiente": ("FFF4E5", "EF6C00"),
        "Regular": ("FDECEA", "C62828"),
    }
    for row, cls in enumerate(["Ã“timo", "Bom", "Suficiente", "Regular"], 10):
        qty = class_counts.get(cls, 0)
        pct = qty / total if total else 0
        bar = "â–ˆ" * round(pct * 20) + "â–‘" * (20 - round(pct * 20))
        bg, fg = class_palette[cls]
        vals = [cls, qty, pct, bar]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, val)
            c.fill = _fill(bg)
            c.border = _border()
            if col == 4:
                c.font = Font(name="Courier New", bold=True, color=fg, size=9)
                c.alignment = _align("left")
            else:
                c.font = _font(bold=(col in (1, 2)), color=fg)
                c.alignment = _align("center")
        ws.cell(row, 3).number_format = "0.0%"

    start_row = 16
    _merge_title(ws, f"A{start_row}:H{start_row}", "ADESÃƒO POR CRITÃ‰RIO", "375623")
    hdr_row = start_row + 1
    for idx, title in enumerate(["CritÃ©rio", "SIM", "NÃƒO/PEND", "%", "Progresso"], 1):
        c = ws.cell(hdr_row, idx, title)
        c.fill = _fill("375623")
        c.font = _font(bold=True, color="FFFFFF")
        c.alignment = _align("center")
        c.border = _border()

    for row, (_letter, title, crit_col) in enumerate(cols["criterios"], hdr_row + 1):
        descr = title.split("-", 1)[1].strip() if "-" in title else title
        sim = sum(1 for p in patients if _normalize_status(load_workbook if False else None) is None)
        sim = sum(1 for p in patients if _normalize_status(p["statuses"][cols["criterios"].index((_letter, title, crit_col))]) == "SIM")
        nao = total - sim
        pct = sim / total if total else 0
        bar = "â–ˆ" * round(pct * 20) + "â–‘" * (20 - round(pct * 20))
        vals = [descr, sim, nao, pct, bar]
        fills = ["E8F5E9", "C6EFCE", "FFC7CE", "F2F2F2", "F2F2F2"]
        colors = ["000000", "2E7D32", "C62828", "000000", "2E7D32"]
        aligns = ["left", "center", "center", "center", "left"]
        for col, (val, bg, fg, align) in enumerate(zip(vals, fills, colors, aligns), 1):
            c = ws.cell(row, col, val)
            c.fill = _fill(bg)
            c.border = _border()
            if col == 5:
                c.font = Font(name="Courier New", bold=True, color=fg, size=9)
            else:
                c.font = _font(bold=(col in (1, 2, 3)), color=fg)
            c.alignment = _align(align)
        ws.cell(row, 4).number_format = "0.0%"

    for idx, width in enumerate([30, 10, 10, 10, 28, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A9"


# -----------------------------------------------------------------------------
# NÃºcleo pÃºblico
# -----------------------------------------------------------------------------


def refresh_interactive_workbook(path) -> Path:
    path = Path(path)
    wb = load_workbook(path)
    ws_data = next((wb[n] for n in wb.sheetnames if n.startswith("ðŸ“‹ Dados") or n.startswith("Dados")), wb[wb.sheetnames[0]])
    ws_search = next((wb[n] for n in wb.sheetnames if n.startswith("ðŸ”") or n.startswith("Busca")), None) or wb.create_sheet("ðŸ” Busca Ativa")
    ws_summary = next((wb[n] for n in wb.sheetnames if n.startswith("ðŸ“Š") or n.startswith("Resumo")), None) or wb.create_sheet("ðŸ“Š Resumo")

    header_row = _detect_header(ws_data)
    cols = _ensure_support_columns(ws_data, header_row, _detect_columns(ws_data, header_row))
    patients = _update_data_sheet(ws_data, header_row, cols)
    code_m = re.search(r"C\d+", ws_data.title, re.I)
    code = code_m.group(0).upper() if code_m else "APS"

    _build_search_sheet(ws_search, cols, patients)
    _build_summary_sheet(ws_summary, cols, patients, code)

    wb.save(path)
    wb.close()
    return path



def clone_interactive(input_path) -> Path:
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(src)
    out = src.with_name(f"{src.stem}_INTERATIVA{src.suffix}")
    shutil.copy2(src, out)
    return refresh_interactive_workbook(out)


# -----------------------------------------------------------------------------
# UI - Editor
# -----------------------------------------------------------------------------


class EditorPlanilhaApp(tk.Toplevel):
    def __init__(self, workbook_path, master=None, initial_patient_name: str | None = None):
        super().__init__(master=master)
        self.workbook_path = Path(workbook_path)
        self.title(f"APS - Editor de Planilha | {self.workbook_path.name}")
        try:
            self.state("zoomed")
        except Exception:
            self.geometry("1200x760")
        self.configure(bg="#EAF2F8")

        self.search_var = tk.StringVar()
        self.filter_prio_var = tk.StringVar(value="Todas")
        self.filter_status_var = tk.StringVar(value="Todos")
        self.sort_var = tk.StringVar(value="Urgencia")
        self.only_dirty_var = tk.BooleanVar(value=False)
        self.selected_bairros: set[str] = set()
        self.all_bairros: list[str] = []
        self.status_var = tk.StringVar(value="Selecione um paciente para editar.")
        self.buffer_var = tk.StringVar(value="0 alteracoes pendentes")
        self.source_dir_var = tk.StringVar(value=str(self.workbook_path.parent))

        self.card_total_var = tk.StringVar(value="0")
        self.card_pend_var = tk.StringVar(value="0")
        self.card_ok_var = tk.StringVar(value="0")
        self.card_media_var = tk.StringVar(value="0.0")

        self.mass_criterio_var = tk.StringVar()
        self.mass_valor_var = tk.StringVar(value="SIM")
        self.mass_scope_var = tk.StringVar(value="Selecionados")
        self.fast_save_var = tk.BooleanVar(value=True)

        self.current_row: int | None = None
        self.criterio_vars: dict[str, tk.StringVar] = {}

        self.base_records: list[dict] = []
        self.records: list[dict] = []
        self.record_map: dict[int, dict] = {}
        self.criteria_info: list[tuple[str, str, int]] = []
        self.criteria_index: dict[str, int] = {}
        self.criteria_weights: dict[str, int] = {}
        self.unified_mode = False
        self.unified_cols: dict[str, int] = {}
        self.general_source_files: list[Path] = []
        self.auto_general_refresh_var = tk.BooleanVar(value=True)
        self.light_mode_enabled = True
        self._general_source_mtimes: dict[Path, int] = {}
        self._auto_general_after_id = None
        self.pending_by_row: dict[int, dict[str, str]] = {}
        self._undo_stack: list[dict[int, dict[str, str]]] = []
        self._chart_regions: list[tuple[int, int, int, int, str]] = []
        self._view_refresh_after_id = None
        self._general_name_index: dict[str, list[dict]] = {}

        self._filter_after_id = None
        self._busy = False
        self._history_path = self.workbook_path.with_name(f"{self.workbook_path.stem}_alteracoes.jsonl")
        self._initial_patient_name = (initial_patient_name or "").strip()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build()
        self.reload_data(async_mode=True)

    def _editor_backup_dir(self) -> Path:
        # Mantem backups do editor separados do arquivo principal para facilitar organizacao.
        folder = self.workbook_path.parent / "BACKUPS_EDITOR"
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def _card(self, parent, title, var, row, col, bg, fg):
        frm = tk.Frame(parent, bg=bg, bd=1, relief="solid")
        frm.grid(row=row, column=col, sticky="nsew", padx=4, pady=4)
        tk.Label(frm, text=title, bg=bg, fg=fg, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=(8, 0))
        tk.Label(frm, textvariable=var, bg=bg, fg=fg, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=10, pady=(0, 8))

    def _build(self):
        top = tk.Frame(self, bg="#EAF2F8")
        top.pack(fill="x", padx=12, pady=12)
        tk.Label(top, text="APS - EDITOR DE PLANILHA INTERATIVA", bg="#1F4E79", fg="white", font=("Segoe UI", 13, "bold"), pady=10).pack(fill="x")
        tk.Label(top, text="Edite normalmente. Suas alteracoes ficam pendentes e sao salvas de uma vez para evitar travamentos.", bg="#EAF2F8", anchor="w", font=("Segoe UI", 9)).pack(fill="x", pady=(8, 0))

        cards = tk.Frame(top, bg="#EAF2F8")
        cards.pack(fill="x", pady=(10, 0))
        for i in range(4):
            cards.columnconfigure(i, weight=1)
        self._card(cards, "Pacientes", self.card_total_var, 0, 0, "#EAF2F8", "#1F4E79")
        self._card(cards, "Pendentes", self.card_pend_var, 0, 1, "#FFF4CC", "#7F6000")
        self._card(cards, "Concluidos", self.card_ok_var, 0, 2, "#EAF7EA", "#2E7D32")
        self._card(cards, "Media", self.card_media_var, 0, 3, "#F3E8FD", "#6A1B9A")

        body = tk.Frame(self, bg="#EAF2F8")
        body.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(1, weight=1)

        filters = tk.Frame(body, bg="#EAF2F8")
        filters.grid(row=0, column=0, sticky="ew", padx=(0, 8), pady=(0, 8))
        tk.Label(filters, text="Buscar:", bg="#EAF2F8").pack(side="left")
        ent = tk.Entry(filters, textvariable=self.search_var)
        ent.pack(side="left", fill="x", expand=True, padx=(6, 6))
        ent.bind("<KeyRelease>", self._schedule_filter)

        tk.Button(filters, text="Filtro Bairro (lista)", command=self.open_bairro_filter).pack(side="left", padx=(0, 8))

        tk.Label(filters, text="Prioridade:", bg="#EAF2F8").pack(side="left")
        cb = ttk.Combobox(filters, textvariable=self.filter_prio_var, values=("Todas", "URGENTE", "ALTA", "MONITORAR", "CONCLUIDO"), state="readonly", width=14)
        cb.pack(side="left")
        cb.bind("<<ComboboxSelected>>", lambda _e: self._apply_filter())

        tk.Label(filters, text="Status:", bg="#EAF2F8").pack(side="left", padx=(8, 0))
        cb_status = ttk.Combobox(filters, textvariable=self.filter_status_var, values=("Todos", "Somente pendentes", "Somente concluidos", "Risco alto"), state="readonly", width=18)
        cb_status.pack(side="left", padx=(6, 0))
        cb_status.bind("<<ComboboxSelected>>", lambda _e: self._apply_filter())

        tk.Label(filters, text="Ordenar:", bg="#EAF2F8").pack(side="left", padx=(8, 0))
        cb_sort = ttk.Combobox(
            filters,
            textvariable=self.sort_var,
            values=("Urgencia", "Pontuacao", "Alfabetica"),
            state="readonly",
            width=12,
        )
        cb_sort.pack(side="left", padx=(6, 0))
        cb_sort.bind("<<ComboboxSelected>>", lambda _e: self._resort_and_refresh())

        ttk.Checkbutton(filters, text="Apenas com alteracao pendente", variable=self.only_dirty_var, command=self._apply_filter).pack(side="left", padx=(8, 0))

        self.tree = ttk.Treeview(body, columns=("prio", "nome", "bairro", "pts"), show="headings", selectmode="extended")
        for key, title, width in [("prio", "Prioridade", 130), ("nome", "Nome", 280), ("bairro", "Bairro", 130), ("pts", "Pontuacao", 90)]:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=width, anchor="w" if key == "nome" else "center")
        self.tree.tag_configure("urgente", background="#FDECEA")
        self.tree.tag_configure("alta", background="#FFF4E5")
        self.tree.tag_configure("monitorar", background="#FFFBE6")
        self.tree.tag_configure("concluido", background="#EAF7EA")
        self.tree.tag_configure("dirty", foreground="#1F4E79")
        self.tree.grid(row=1, column=0, sticky="nsew", padx=(0, 8))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        panel = tk.Frame(body, bg="#FFFFFF", bd=1, relief="solid")
        panel.grid(row=0, column=1, rowspan=2, sticky="nsew")
        panel.columnconfigure(0, weight=1)
        panel.rowconfigure(0, weight=1)

        panel_canvas = tk.Canvas(panel, bg="#FFFFFF", highlightthickness=0)
        panel_scroll = ttk.Scrollbar(panel, orient="vertical", command=panel_canvas.yview)
        panel_canvas.configure(yscrollcommand=panel_scroll.set)
        panel_canvas.grid(row=0, column=0, sticky="nsew")
        panel_scroll.grid(row=0, column=1, sticky="ns")

        panel_inner = tk.Frame(panel_canvas, bg="#FFFFFF")
        panel_window = panel_canvas.create_window((0, 0), window=panel_inner, anchor="nw")
        panel_inner.columnconfigure(1, weight=1)
        panel_inner.rowconfigure(2, weight=1)
        panel_inner.rowconfigure(6, weight=1)

        def _sync_panel_scroll(_ev=None):
            try:
                panel_canvas.configure(scrollregion=panel_canvas.bbox("all"))
                panel_canvas.itemconfigure(panel_window, width=panel_canvas.winfo_width())
            except Exception:
                pass

        panel_inner.bind("<Configure>", _sync_panel_scroll)
        panel_canvas.bind("<Configure>", _sync_panel_scroll)

        self.lbl_nome = tk.Label(panel_inner, text="Paciente: -", bg="#FFFFFF", anchor="w", font=("Segoe UI", 12, "bold"))
        self.lbl_nome.grid(row=0, column=0, columnspan=2, sticky="ew", padx=12, pady=(12, 8))
        self.txt_meta = tk.Text(panel_inner, height=5, wrap="word", bg="#FFFFFF", relief="flat")
        self.txt_meta.grid(row=1, column=0, columnspan=2, sticky="ew", padx=12, pady=(0, 12))
        self.txt_meta.configure(state="disabled")
        self.frm_criterios = tk.Frame(panel_inner, bg="#FFFFFF")
        self.frm_criterios.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=12)

        mass = tk.LabelFrame(panel_inner, text="Alteracao em massa", bg="#FFFFFF", fg="#1F4E79")
        mass.grid(row=3, column=0, columnspan=2, sticky="ew", padx=12, pady=(8, 0))
        mass.columnconfigure(1, weight=1)
        tk.Label(mass, text="Criterio:", bg="#FFFFFF").grid(row=0, column=0, sticky="w", padx=(8, 4), pady=(8, 4))
        self.cb_mass_criterio = ttk.Combobox(mass, textvariable=self.mass_criterio_var, state="readonly")
        self.cb_mass_criterio.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=(8, 4))
        tk.Label(mass, text="Valor:", bg="#FFFFFF").grid(row=1, column=0, sticky="w", padx=(8, 4), pady=(0, 8))
        ttk.Combobox(mass, textvariable=self.mass_valor_var, values=("", "SIM", "NAO", "PENDENTE"), state="readonly", width=12).grid(row=1, column=1, sticky="w", pady=(0, 8))
        ttk.Combobox(mass, textvariable=self.mass_scope_var, values=("Selecionados", "Filtrados", "Todos"), state="readonly", width=16).grid(row=1, column=1, sticky="e", padx=(0, 8), pady=(0, 8))

        actions = tk.Frame(panel_inner, bg="#FFFFFF")
        actions.grid(row=4, column=0, columnspan=2, sticky="ew", padx=12, pady=12)
        for col in range(4):
            actions.columnconfigure(col, weight=1)

        tk.Button(actions, text="Selecionar planilhas (geral)", command=self.load_general_from_files).grid(row=0, column=0, sticky="ew", padx=4, pady=4)
        tk.Button(actions, text="Gerar planilha cruzada", command=self.export_general_workbook).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
        tk.Button(actions, text="Aplicar em massa", command=self.apply_mass_change).grid(row=0, column=2, sticky="ew", padx=4, pady=4)
        tk.Button(actions, text="Salvar", command=self.flush_buffer, bg="#1F4E79", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=3, sticky="ew", padx=4, pady=4)

        tk.Button(actions, text="Desfazer", command=self.undo_last_change).grid(row=1, column=0, sticky="ew", padx=4, pady=4)
        tk.Button(actions, text="Historico de mudancas", command=self.open_change_history).grid(row=1, column=1, sticky="ew", padx=4, pady=4)
        tk.Button(actions, text="Abrir planilha", command=self._open_workbook).grid(row=1, column=2, sticky="ew", padx=4, pady=4)
        tk.Button(actions, text="Atualizar planilha", command=self._manual_refresh_planilha).grid(row=1, column=3, sticky="ew", padx=4, pady=4)

        tk.Label(panel_inner, textvariable=self.buffer_var, bg="#FFFFFF", fg="#1F4E79", anchor="w").grid(row=5, column=0, columnspan=2, sticky="ew", padx=12, pady=(0, 12))
        chart_box = tk.LabelFrame(panel_inner, text="Grafico ao vivo (clique para filtrar prioridade)", bg="#FFFFFF", fg="#1F4E79")
        chart_box.grid(row=6, column=0, columnspan=2, sticky="nsew", padx=12, pady=(0, 12))
        chart_box.columnconfigure(0, weight=1)
        chart_box.rowconfigure(0, weight=1)
        self.chart_canvas = tk.Canvas(chart_box, bg="#FFFFFF", height=170, highlightthickness=0)
        self.chart_canvas.grid(row=0, column=0, sticky="nsew")
        self.chart_canvas.bind("<Button-1>", self._on_chart_click)

        def _on_panel_wheel(event):
            try:
                panel_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        panel_canvas.bind("<MouseWheel>", _on_panel_wheel)

        tk.Label(self, textvariable=self.status_var, anchor="w", relief="sunken").pack(fill="x", side="bottom")

    def _load_context(self):
        wb = load_workbook(self.workbook_path)
        ws = next(
            (
                wb[n]
                for n in wb.sheetnames
                if n.startswith("📋 Dados") or n.startswith("ðŸ“‹ Dados") or n.startswith("Dados")
            ),
            wb[wb.sheetnames[0]],
        )
        header_row = _detect_header(ws)
        cols = _detect_columns(ws, header_row)
        return wb, ws, header_row, cols

    def _detect_unified(self, ws, header_row: int, cols: dict) -> bool:
        headers = {str(v).strip() for v in cols.get("headers", {}).keys()}
        if "Indicadores" in headers and "O que fazer" in headers:
            return True
        if ws.title.startswith("📋 Pacientes") or ws.title.startswith("ðŸ“‹ Pacientes"):
            return True
        return False

    def _unified_columns(self, cols: dict) -> dict[str, int]:
        headers = cols.get("headers", {})
        def _pick(name):
            return headers.get(name)
        return {
            "nome": _pick("Nome"),
            "bairro": _pick("Bairro") or _pick("Bairro/Localidade") or _pick("Localidade"),
            "microarea": _pick("MicroÃ¡rea") or _pick("Microarea"),
            "endereco": _pick("EndereÃ§o") or _pick("Endereco") or _pick("Logradouro"),
            "numero": _pick("Numero") or _pick("NÃºmero") or _pick("Num"),
            "complemento": _pick("Complemento"),
            "cidade": _pick("Cidade") or _pick("MunicÃ­pio") or _pick("Municipio"),
            "uf": _pick("UF") or _pick("Estado"),
            "cep": _pick("CEP"),
            "tel": _pick("Telefone"),
            "indicadores": _pick("Indicadores"),
            "qtd": _pick("Qtd"),
            "pend": _pick("PendÃªncias"),
            "media": _pick("MÃ©dia"),
            "prioridade": _pick("Prioridade"),
            "oqf": _pick("O que fazer"),
        }

    def _set_busy(self, value: bool):
        self._busy = value
        self.configure(cursor="watch" if value else "")

    def _norm_prio(self, value: str) -> str:
        txt = str(value or "").upper()
        if "URGENTE" in txt:
            return "URGENTE"
        if "ALTA" in txt:
            return "ALTA"
        if "MONITORAR" in txt:
            return "MONITORAR"
        if "CONCL" in txt:
            return "CONCLUIDO"
        return txt

    def _pending_text_from_statuses(self, statuses: list[str]) -> str:
        pend = []
        for idx, (_letter, title, _col) in enumerate(self.criteria_info):
            if _normalize_status(statuses[idx]) != "SIM":
                pend.append(title.split("-", 1)[1].strip() if "-" in title else title)
        return "; ".join(pend)

    def _derived_from_statuses(self, statuses: list[str]) -> tuple[int, str, str, str]:
        pts = 0
        for idx, (letter, _title, _col) in enumerate(self.criteria_info):
            if _normalize_status(statuses[idx]) == "SIM":
                pts += self.criteria_weights.get(letter, 0)
        return pts, _classify(pts), _priority(pts), self._pending_text_from_statuses(statuses)

    def _effective_record(self, base: dict) -> dict:
        row = base["row"]
        overrides = self.pending_by_row.get(row, {})
        if self.unified_mode:
            rec = dict(base)
            for key, value in overrides.items():
                rec[key] = value
            rec["dirty"] = bool(overrides)
            return rec
        statuses = list(base["statuses"])
        for title, value in overrides.items():
            idx = self.criteria_index.get(title)
            if idx is not None:
                statuses[idx] = value
        pts, classif, prio, pend = self._derived_from_statuses(statuses)
        rec = dict(base)
        rec["statuses"] = statuses
        rec["pts"] = pts
        rec["classif"] = classif
        rec["prio"] = prio
        rec["pend"] = pend
        rec["dirty"] = bool(overrides)
        return rec

    def _refresh_cards(self):
        total = len(self.records)
        concluidos = sum(1 for p in self.records if self._norm_prio(p.get("prio", "")) == "CONCLUIDO")
        if self.unified_mode:
            medias = []
            for p in self.records:
                try:
                    medias.append(float(str(p.get("media", "")).strip().replace(",", ".")))
                except Exception:
                    pass
            media = sum(medias) / len(medias) if medias else 0
        else:
            media = sum(p["pts"] for p in self.records) / total if total else 0
        self.card_total_var.set(str(total))
        self.card_ok_var.set(str(concluidos))
        self.card_pend_var.set(str(total - concluidos))
        self.card_media_var.set(f"{media:.1f}")
        rows_dirty = len(self.pending_by_row)
        cells_dirty = sum(len(v) for v in self.pending_by_row.values())
        self.buffer_var.set(f"{rows_dirty} linhas com alteracoes pendentes | {cells_dirty} campos pendentes")
        self._update_chart()

    def _build_records_view(self):
        merged = [self._effective_record(rec) for rec in self.base_records]
        merged.sort(key=self._sort_key_for_record)
        self.records = merged
        self.record_map = {p["row"]: p for p in merged}
        self._refresh_cards()

    def _sort_key_for_record(self, rec: dict):
        mode = self.sort_var.get().strip()
        name_key = str(rec.get("nome", "")).lower()
        prio_order = {"URGENTE": 0, "ALTA": 1, "MONITORAR": 2, "CONCLUIDO": 3}
        prio_key = prio_order.get(self._norm_prio(rec.get("prio", "")), 9)

        def _num(val):
            try:
                return float(str(val).replace(",", "."))
            except Exception:
                return 0.0

        score_key = _num(rec.get("media", rec.get("pts", 0)))
        if mode == "Alfabetica":
            return (name_key, prio_key, score_key)
        if mode == "Pontuacao":
            return (score_key, prio_key, name_key)
        return (prio_key, score_key, name_key)

    def _resort_and_refresh(self):
        self._build_records_view()
        self._apply_filter()

    def _configure_tree_columns(self):
        self.tree.heading("pts", text="Media" if self.unified_mode else "Pontuacao")

    def _name_key(self, text: str) -> str:
        s = str(text or "").strip().lower()
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"\s+", " ", s).strip()

    def _rebuild_general_name_index(self):
        idx: dict[str, list[dict]] = {}
        for rec in self.base_records:
            nk = self._name_key(rec.get("nome", ""))
            if not nk:
                continue
            idx.setdefault(nk, []).append(rec)
        self._general_name_index = idx

    def _refresh_bairro_values(self):
        values = sorted({str(r.get("bairro", "")).strip() for r in self.base_records if str(r.get("bairro", "")).strip()})
        self.all_bairros = values
        if not self.selected_bairros:
            self.selected_bairros = set(values)
            return
        self.selected_bairros = {b for b in self.selected_bairros if b in set(values)}
        if not self.selected_bairros:
            self.selected_bairros = set(values)

    def _row_value(self, row, *keys: str) -> str:
        index_map = {_norm_header_text(str(k)): k for k in row.index}
        for key in keys:
            if key in row.index:
                val = row.get(key)
                return "" if val is None else str(val)
            norm = _norm_header_text(key)
            original = index_map.get(norm)
            if original is not None:
                val = row.get(original)
                return "" if val is None else str(val)
        return ""

    def _record_from_unified_row(self, idx: int, row) -> dict:
        bairro = self._row_value(row, "Bairro", "Microárea", "MicroÃ¡rea", "Microarea")
        endereco = self._row_value(row, "Endereco", "Endereço", "EndereÃ§o", "Logradouro")
        ind_scores: dict[str, str] = {}
        for col in row.index:
            col_s = str(col).strip()
            m = re.match(r"^(C\d+)$", col_s, re.I)
            if not m:
                continue
            code = m.group(1).upper()
            v = row.get(col, "")
            ind_scores[code] = "" if v is None else str(v)
        return {
            "row": idx + 1,
            "nome": self._row_value(row, "Nome"),
            "bairro": bairro,
            "microarea": self._row_value(row, "Microárea", "MicroÃ¡rea", "Microarea"),
            "endereco": endereco,
            "endereco_full": endereco,
            "tel": self._row_value(row, "Telefone"),
            "indicadores": self._row_value(row, "Indicadores"),
            "qtd": self._row_value(row, "Qtd"),
            "pend": self._row_value(row, "Pendências", "PendÃªncias"),
            "media": self._row_value(row, "Média", "MÃ©dia"),
            "prio": self._row_value(row, "Prioridade"),
            "oqf": self._row_value(row, "O que fazer"),
            "ind_scores": ind_scores,
        }

    def _records_from_unified_df(self, df) -> list[dict]:
        return [self._record_from_unified_row(idx, row) for idx, row in df.iterrows()]

    def _capture_general_source_mtimes(self):
        mtimes: dict[Path, int] = {}
        for src in self.general_source_files:
            try:
                mtimes[src] = src.stat().st_mtime_ns
            except Exception:
                continue
        self._general_source_mtimes = mtimes

    def _discover_general_source_files(self) -> list[Path]:
        folder = Path(self.source_dir_var.get().strip() or self.workbook_path.parent)
        candidates: list[Path] = []
        if folder.exists():
            for p in folder.glob("*.xlsx"):
                n = p.name.lower()
                if not infer_indicator_code_from_path(p):
                    continue
                if "backup" in n or "interativa" in n or "cruz" in n or "compar" in n or "unificad" in n:
                    continue
                candidates.append(p)
        if not candidates:
            candidates = [p for p in self.general_source_files if p.exists()]
        by_code: dict[str, Path] = {}
        for p in candidates:
            code = infer_indicator_code_from_path(p)
            if not code:
                continue
            old = by_code.get(code)
            if old is None:
                by_code[code] = p
                continue
            try:
                if p.stat().st_mtime_ns >= old.stat().st_mtime_ns:
                    by_code[code] = p
            except Exception:
                by_code[code] = p
        return sorted(by_code.values(), key=lambda x: x.name.lower())

    def _schedule_general_auto_refresh(self):
        if self._auto_general_after_id:
            try:
                self.after_cancel(self._auto_general_after_id)
            except Exception:
                pass
            self._auto_general_after_id = None
        interval_ms = 9000 if self.light_mode_enabled else 3500
        self._auto_general_after_id = self.after(interval_ms, self._check_general_auto_refresh)

    def _schedule_view_refresh(self, delay_ms: int = 80):
        if self._view_refresh_after_id:
            try:
                self.after_cancel(self._view_refresh_after_id)
            except Exception:
                pass
            self._view_refresh_after_id = None

        def _do_refresh():
            self._view_refresh_after_id = None
            selected = self.current_row
            self._build_records_view()
            self._apply_filter()
            if selected is not None and self.tree.exists(str(selected)):
                self.tree.selection_set(str(selected))
                self.tree.focus(str(selected))
                self.tree.see(str(selected))

        self._view_refresh_after_id = self.after(delay_ms, _do_refresh)

    def _check_general_auto_refresh(self):
        self._auto_general_after_id = None
        if not self.winfo_exists():
            return
        if not self.unified_mode or not self.general_source_files:
            return
        if not self.auto_general_refresh_var.get():
            self._schedule_general_auto_refresh()
            return
        if self._busy:
            self._schedule_general_auto_refresh()
            return
        changed = False
        current_mtimes: dict[Path, int] = {}
        for src in self.general_source_files:
            try:
                current_mtimes[src] = src.stat().st_mtime_ns
            except Exception:
                continue
            if self._general_source_mtimes.get(src) != current_mtimes[src]:
                changed = True
        if changed:
            self._refresh_general_from_current_files("Auto atualizacao: detectada mudanca nas planilhas fonte.")
            return
        self._general_source_mtimes = current_mtimes
        self._schedule_general_auto_refresh()

    def _refresh_general_from_current_files(self, status_prefix: str = "Atualizando visao geral..."):
        files = self._discover_general_source_files()
        if not files:
            self.status_var.set("Nao foi possivel atualizar: nenhuma planilha fonte encontrada.")
            self._schedule_general_auto_refresh()
            return
        selected_name = ""
        if self.current_row is not None and self.current_row in self.record_map:
            selected_name = str(self.record_map[self.current_row].get("nome", "")).strip()
        self._set_busy(True)
        self.status_var.set(status_prefix)

        def worker():
            import aps_comparador_paciente as comparador
            df = comparador.build_unified(files)
            if df is None or df.empty:
                raise RuntimeError("Nao foi possivel montar a visao geral com os arquivos atuais.")
            return self._records_from_unified_df(df), files

        def on_ok(result):
            records, selected_files = result
            self._apply_general_records(records, selected_files)
            self._focus_patient_by_name(selected_name)
            self._set_busy(False)
            self.status_var.set(f"Visao geral atualizada automaticamente ({len(records)} pacientes).")

        def on_err(exc):
            self._set_busy(False)
            self.status_var.set(f"Falha na auto atualizacao da geral: {exc}")
            self._schedule_general_auto_refresh()

        def runner():
            try:
                result = worker()
                self.after(0, lambda: on_ok(result))
            except Exception as exc:
                self.after(0, lambda: on_err(exc))

        threading.Thread(target=runner, daemon=True).start()

    def _apply_general_records(self, records: list[dict], source_files: list[Path]):
        self.unified_mode = True
        self.unified_cols = {}
        self.criteria_info = []
        self.criteria_weights = {}
        self.criteria_index = {}
        self.general_source_files = source_files
        self.base_records = records
        self._rebuild_general_name_index()
        if source_files:
            self.source_dir_var.set(str(source_files[0].parent))
        self._refresh_bairro_values()
        self._configure_tree_columns()
        self._build_records_view()
        self._apply_filter()
        self._capture_general_source_mtimes()
        self._schedule_general_auto_refresh()
        self.status_var.set(f"Visao geral carregada com {len(records)} pacientes de {len(source_files)} planilhas.")

    def _to_num(self, value) -> float:
        txt = str(value or "").replace(",", ".")
        m = re.search(r"-?\d+(?:\.\d+)?", txt)
        if not m:
            return 0.0
        try:
            return float(m.group(0))
        except Exception:
            return 0.0

    def _prio_from_media(self, media: float) -> str:
        if media >= 100:
            return "CONCLUIDO"
        if media >= 75:
            return "MONITORAR"
        if media >= 50:
            return "ALTA"
        return "URGENTE"

    def _patch_oqf_indicator(self, old_oqf: str, code: str, pend: str) -> tuple[str, int]:
        lines = [ln.strip() for ln in str(old_oqf or "").splitlines() if ln.strip()]
        lines = [ln for ln in lines if not re.match(rf"^\[{re.escape(code)}\]\s*", ln, re.I)]
        pend_txt = str(pend or "").strip()
        if pend_txt and pend_txt.lower() not in {"none", "nan", "-"}:
            lines.append(f"[{code}] {pend_txt}")
        return ("\n".join(lines) if lines else "Em dia"), len(lines)

    def on_indicator_saved(self, source_file: Path, changed_names: list[str]):
        if not self.unified_mode:
            return
        if not changed_names:
            return
        target_names = {self._name_key(n) for n in changed_names if self._name_key(n)}
        if not target_names:
            return

        self.status_var.set("Aplicando atualizacao rapida na visao geral...")

        def worker():
            import aps_comparador_paciente as comparador
            code, snap = comparador.build_indicator_snapshot(Path(source_file))
            return code, snap

        def on_ok(result):
            code, snap = result
            if not code:
                return
            changed = 0
            for nk in target_names:
                rec_list = self._general_name_index.get(nk, [])
                if not rec_list:
                    continue
                info = snap.get(nk)
                for rec in rec_list:
                    ind = dict(rec.get("ind_scores", {}))
                    if info is None:
                        ind[code] = "â€”"
                        oqf, pend_count = self._patch_oqf_indicator(rec.get("oqf", ""), code, "")
                    else:
                        ind[code] = f"{int(round(info.get('pts', 0), 0))} pts"
                        oqf, pend_count = self._patch_oqf_indicator(rec.get("oqf", ""), code, info.get("pend", ""))
                        if not str(rec.get("bairro", "")).strip() and str(info.get("bairro", "")).strip():
                            rec["bairro"] = info.get("bairro", "")
                        if not str(rec.get("endereco_full", "")).strip() and str(info.get("endereco", "")).strip():
                            rec["endereco_full"] = info.get("endereco", "")
                            rec["endereco"] = info.get("endereco", "")
                        if not str(rec.get("tel", "")).strip() and str(info.get("telefone", "")).strip():
                            rec["tel"] = info.get("telefone", "")
                    rec["ind_scores"] = ind
                    present = []
                    total = 0.0
                    cnt = 0
                    for ckey, v in ind.items():
                        txt = str(v or "").strip()
                        if txt and txt not in {"â€”", "-", "—"}:
                            present.append(ckey)
                            n = self._to_num(txt)
                            total += n
                            cnt += 1
                    media = int(round(total / cnt, 0)) if cnt else 0
                    rec["media"] = str(media)
                    rec["prio"] = self._prio_from_media(media)
                    rec["indicadores"] = " Â· ".join(sorted(present)) if present else "â€”"
                    rec["qtd"] = str(len(present))
                    rec["oqf"] = oqf
                    rec["pend"] = str(pend_count)
                    changed += 1
            if changed:
                self._schedule_view_refresh(30 if self.light_mode_enabled else 0)
                self.status_var.set(f"Visao geral atualizada automaticamente ({changed} registro(s) ajustado(s)).")

        def on_err(exc):
            self.status_var.set(f"Falha na atualizacao rapida da geral: {exc}")

        def runner():
            try:
                result = worker()
                self.after(0, lambda: on_ok(result))
            except Exception as exc:
                self.after(0, lambda: on_err(exc))

        threading.Thread(target=runner, daemon=True).start()

    def load_general_from_files(self):
        paths = filedialog.askopenfilenames(
            title="Selecione as planilhas para montar a visao geral",
            filetypes=[("Excel", "*.xlsx *.xls")],
            initialdir=str(self.source_dir_var.get().strip() or self.workbook_path.parent),
        )
        if not paths:
            return
        files = [Path(p) for p in paths]
        try:
            import aps_comparador_paciente as comparador
            df = comparador.build_unified(files)
        except Exception as exc:
            messagebox.showerror("Erro", f"Falha ao montar visao geral:\n{exc}", parent=self)
            return
        if df is None or df.empty:
            messagebox.showinfo("Sem dados", "Nao foi possivel montar a visao geral com essas planilhas.")
            return
        records = self._records_from_unified_df(df)
        self._apply_general_records(records, files)

    def export_general_workbook(self):
        files = [p for p in self.general_source_files if p.exists()]
        if not files:
            paths = filedialog.askopenfilenames(
                title="Selecione as planilhas para gerar a cruzada",
                filetypes=[("Excel", "*.xlsx *.xls")],
                initialdir=str(self.source_dir_var.get().strip() or self.workbook_path.parent),
            )
            if not paths:
                return
            files = [Path(p) for p in paths]
        try:
            import aps_comparador_paciente as comparador
            df = comparador.build_unified(files)
            if df is None or df.empty:
                messagebox.showwarning("Sem dados", "Nao foi possivel montar a planilha cruzada com esses arquivos.", parent=self)
                return
            default_name = f"CRUZADA_APS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_path = filedialog.asksaveasfilename(
                title="Salvar planilha cruzada",
                defaultextension=".xlsx",
                initialdir=str(self.workbook_path.parent),
                initialfile=default_name,
                filetypes=[("Excel", "*.xlsx")],
            )
            if not out_path:
                return
            comparador.export_excel(df, Path(out_path))
            self.status_var.set(f"Planilha cruzada gerada: {Path(out_path).name}")
            messagebox.showinfo("Concluido", f"Planilha cruzada salva em:\n{out_path}", parent=self)
        except Exception as exc:
            messagebox.showerror("Erro", f"Falha ao gerar planilha cruzada:\n{exc}", parent=self)

    def open_bairro_filter(self):
        self._refresh_bairro_values()
        win = tk.Toplevel(self)
        win.title("Filtro de Bairro")
        win.geometry("420x520")
        win.configure(bg="#F7FAFD")
        vars_map: dict[str, tk.BooleanVar] = {}

        top = tk.Frame(win, bg="#F7FAFD")
        top.pack(fill="x", padx=10, pady=10)
        tk.Label(top, text="Selecione os bairros que devem aparecer:", bg="#F7FAFD", anchor="w").pack(fill="x")

        controls = tk.Frame(top, bg="#F7FAFD")
        controls.pack(fill="x", pady=(8, 4))

        box = tk.Frame(win, bg="#F7FAFD")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        canvas = tk.Canvas(box, bg="#FFFFFF", highlightthickness=1, highlightbackground="#D7E3EE")
        scroll = ttk.Scrollbar(box, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg="#FFFFFF")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        for bairro in self.all_bairros:
            v = tk.BooleanVar(value=bairro in self.selected_bairros)
            vars_map[bairro] = v
            ttk.Checkbutton(inner, text=bairro, variable=v).pack(anchor="w", padx=8, pady=2)

        def mark_all(value: bool):
            for var in vars_map.values():
                var.set(value)

        ttk.Button(controls, text="Marcar todos", command=lambda: mark_all(True)).pack(side="left")
        ttk.Button(controls, text="Limpar", command=lambda: mark_all(False)).pack(side="left", padx=6)

        def apply():
            chosen = {bairro for bairro, var in vars_map.items() if var.get()}
            if not chosen:
                messagebox.showwarning("Filtro de bairro", "Selecione ao menos um bairro.", parent=win)
                return
            self.selected_bairros = chosen
            self._apply_filter()
            win.destroy()

        footer = tk.Frame(win, bg="#F7FAFD")
        footer.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(footer, text="Aplicar", command=apply).pack(side="right")
        ttk.Button(footer, text="Cancelar", command=win.destroy).pack(side="right", padx=(0, 8))

    def reload_data(self, async_mode: bool = False):
        if self._busy:
            return

        def work():
            wb, ws, header_row, cols = self._load_context()
            unified = self._detect_unified(ws, header_row, cols)
            if unified:
                ucols = self._unified_columns(cols)
                records = []
                start = header_row + 1
                for row in range(start, ws.max_row + 1):
                    nome = ws.cell(row, ucols["nome"]).value if ucols.get("nome") else None
                    if not nome:
                        continue
                    oqf = ws.cell(row, ucols["oqf"]).value if ucols.get("oqf") else ""
                    tel = ws.cell(row, ucols["tel"]).value if ucols.get("tel") else ""
                    indicadores = ws.cell(row, ucols["indicadores"]).value if ucols.get("indicadores") else ""
                    qtd = ws.cell(row, ucols["qtd"]).value if ucols.get("qtd") else ""
                    pend = ws.cell(row, ucols["pend"]).value if ucols.get("pend") else ""
                    media = ws.cell(row, ucols["media"]).value if ucols.get("media") else ""
                    prio = ws.cell(row, ucols["prioridade"]).value if ucols.get("prioridade") else ""
                    if not any([oqf, tel, indicadores, qtd, pend, media, prio]):
                        continue
                    rec = {
                        "row": row,
                        "nome": str(nome),
                        "bairro": (ws.cell(row, ucols["bairro"]).value or "") if ucols.get("bairro") else "",
                        "microarea": (ws.cell(row, ucols["microarea"]).value or "") if ucols.get("microarea") else "",
                        "endereco": (ws.cell(row, ucols["endereco"]).value or "") if ucols.get("endereco") else "",
                        "tel": str(tel or ""),
                        "indicadores": str(indicadores or ""),
                        "qtd": str(qtd or ""),
                        "pend": str(pend or ""),
                        "media": str(media or ""),
                        "prio": str(prio or ""),
                        "oqf": str(oqf or ""),
                    }
                    rec["endereco_full"] = _compose_endereco({
                        "endereco": ws.cell(row, ucols["endereco"]).value if ucols.get("endereco") else "",
                        "numero": ws.cell(row, ucols["numero"]).value if ucols.get("numero") else "",
                        "complemento": ws.cell(row, ucols["complemento"]).value if ucols.get("complemento") else "",
                        "bairro": rec["bairro"],
                        "cidade": ws.cell(row, ucols["cidade"]).value if ucols.get("cidade") else "",
                        "uf": ws.cell(row, ucols["uf"]).value if ucols.get("uf") else "",
                        "cep": ws.cell(row, ucols["cep"]).value if ucols.get("cep") else "",
                    })
                    records.append(rec)
                wb.close()
                return records, [], {}, True, ucols

            cols = _ensure_support_columns(ws, header_row, cols)
            patients = _update_data_sheet(ws, header_row, cols)
            wb.save(self.workbook_path)
            wb.close()
            return patients, cols["criterios"], _weights(cols["criterios"]), False, {}

        def done_ok(result):
            patients, criterios, pesos, unified, ucols = result
            self.unified_mode = unified
            self.unified_cols = ucols
            self.base_records = patients
            self._refresh_bairro_values()
            self._configure_tree_columns()
            self.criteria_info = criterios
            self.criteria_weights = pesos
            self.criteria_index = {title: idx for idx, (_l, title, _c) in enumerate(criterios)}
            self.cb_mass_criterio.configure(values=[t for _l, t, _c in criterios])
            if criterios and not self.mass_criterio_var.get():
                self.mass_criterio_var.set(criterios[0][1])
            self._build_records_view()
            self._apply_filter()
            self._focus_patient_by_name(self._initial_patient_name)
            self.status_var.set(f"Planilha carregada: {self.workbook_path.name} | {len(self.base_records)} pacientes")
            self._set_busy(False)

        def done_err(exc):
            self._set_busy(False)
            self.status_var.set(f"Erro ao recarregar: {exc}")
            messagebox.showerror("Erro", str(exc), parent=self)

        if not async_mode:
            self._set_busy(True)
            try:
                done_ok(work())
            except Exception as exc:
                done_err(exc)
            return

        self._set_busy(True)
        self.status_var.set("Recarregando planilha e calculando indicadores...")

        def runner():
            try:
                result = work()
                self.after(0, lambda: done_ok(result))
            except Exception as exc:
                self.after(0, lambda: done_err(exc))

        threading.Thread(target=runner, daemon=True).start()

    def _schedule_filter(self, _evt=None):
        if self._filter_after_id:
            self.after_cancel(self._filter_after_id)
        self._filter_after_id = self.after(180, self._apply_filter)

    def _apply_filter(self):
        term = self.search_var.get().strip().lower()
        prio = self.filter_prio_var.get().strip()
        status = self.filter_status_var.get().strip()
        only_dirty = self.only_dirty_var.get()
        bairro_filter_active = bool(self.all_bairros) and set(self.selected_bairros) != set(self.all_bairros)

        self.tree.delete(*self.tree.get_children())
        for p in self.records:
            nome = str(p["nome"]).lower()
            bairro_raw = str(p.get("bairro", "")).strip()
            bairro_val = bairro_raw.lower()
            endereco_val = str(p.get("endereco_full", p.get("endereco", ""))).lower()
            tel_val = str(p.get("tel", "")).lower()
            prio_norm = self._norm_prio(p["prio"])
            if bairro_filter_active and bairro_raw not in self.selected_bairros:
                continue
            if term and term not in nome and term not in bairro_val and term not in endereco_val:
                continue
            if prio and prio != "Todas" and prio_norm != prio:
                continue
            if status == "Somente pendentes" and prio_norm == "CONCLUIDO":
                continue
            if status == "Somente concluidos" and prio_norm != "CONCLUIDO":
                continue
            if status == "Risco alto" and prio_norm not in {"URGENTE", "ALTA"}:
                continue
            if only_dirty and not p.get("dirty"):
                continue
            tag = {
                "URGENTE": "urgente",
                "ALTA": "alta",
                "MONITORAR": "monitorar",
                "CONCLUIDO": "concluido",
            }.get(prio_norm, "monitorar")
            tags = [tag]
            if p.get("dirty"):
                tags.append("dirty")
            self.tree.insert(
                "",
                "end",
                iid=str(p["row"]),
                values=(prio_norm, p["nome"], p.get("bairro", "") or p.get("microarea", ""), p.get("media", p.get("pts", ""))),
                tags=tuple(tags),
            )
        self._refresh_cards()

    def _focus_patient_by_name(self, patient_name: str):
        name = (patient_name or "").strip().lower()
        if not name:
            return
        for row_id in self.tree.get_children(""):
            rec = self.record_map.get(int(row_id))
            if not rec:
                continue
            rec_name = str(rec.get("nome", "")).strip().lower()
            if rec_name == name or name in rec_name:
                self.tree.selection_set(row_id)
                self.tree.focus(row_id)
                self.tree.see(row_id)
                self._on_select()
                self.status_var.set(f"Paciente localizado automaticamente: {rec.get('nome', '')}")
                return

    def _on_select(self, _evt=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.current_row = int(sel[0])
        p = self.record_map[self.current_row]
        self.lbl_nome.config(text=f"Paciente: {p['nome']}")
        dirty_label = "SIM" if p.get("dirty") else "NAO"
        meta_text = (
            f"Bairro: {p.get('bairro', '')}\n"
            f"Endereco: {p.get('endereco_full', p.get('endereco', ''))}\n"
            f"Telefone: {p.get('tel','')}\n"
            f"Pontuacao: {p.get('pts','')}\n"
            f"Prioridade: {self._norm_prio(p.get('prio',''))}\n"
            f"Alteracao pendente: {dirty_label}"
        )
        self.txt_meta.configure(state="normal")
        self.txt_meta.delete("1.0", "end")
        self.txt_meta.insert("1.0", meta_text)
        self.txt_meta.configure(state="disabled")
        for widget in self.frm_criterios.winfo_children():
            widget.destroy()
        self.criterio_vars = {}

        if self.unified_mode:
            tk.Label(self.frm_criterios, text="Planilha cruzada: edicao de SIM/NAO so nas planilhas originais.", bg="#FFFFFF", anchor="w").grid(
                row=0, column=0, columnspan=2, sticky="w", pady=(2, 8)
            )
            tk.Label(self.frm_criterios, text="Pendencias (geral):", bg="#FFFFFF", anchor="w", font=("Segoe UI", 9, "bold")).grid(
                row=1, column=0, sticky="w", pady=(0, 4)
            )
            tk.Label(self.frm_criterios, text=p.get("oqf", p.get("pend", "")), bg="#FFFFFF", anchor="w", justify="left", wraplength=420).grid(
                row=2, column=0, columnspan=2, sticky="w", pady=(0, 8)
            )
            tk.Label(self.frm_criterios, text="Pasta para buscar planilhas:", bg="#FFFFFF", anchor="w").grid(row=3, column=0, sticky="w")
            tk.Entry(self.frm_criterios, textvariable=self.source_dir_var).grid(row=3, column=1, sticky="ew", pady=(0, 6))
            tk.Button(self.frm_criterios, text="Escolher pasta", command=self.choose_source_folder).grid(
                row=4, column=0, columnspan=2, pady=(0, 8)
            )
            self._unified_indicator_var = tk.StringVar()
            ind_list = []
            if isinstance(p.get("indicadores", ""), str):
                ind_list = [x.strip() for x in p.get("indicadores", "").split("·") if x.strip()]
            if not ind_list:
                ind_list = sorted({code for f in self.general_source_files for code in [infer_indicator_code_from_path(f)] if code})
            cb_ind = ttk.Combobox(self.frm_criterios, textvariable=self._unified_indicator_var, values=ind_list, state="readonly")
            if ind_list:
                self._unified_indicator_var.set(ind_list[0])
            cb_ind.grid(row=5, column=1, sticky="ew", pady=(0, 6))
            tk.Label(self.frm_criterios, text="Indicador:", bg="#FFFFFF", anchor="w").grid(row=5, column=0, sticky="w")
            tk.Button(self.frm_criterios, text="Abrir planilha do indicador", command=self.open_indicator_source).grid(
                row=6, column=0, columnspan=2, pady=(4, 2)
            )
            self.frm_criterios.columnconfigure(1, weight=1)
            return

        for idx, ((_letter, title, _col), value) in enumerate(zip(self.criteria_info, p["statuses"])):
            descr = title.split("-", 1)[1].strip() if "-" in title else title
            tk.Label(self.frm_criterios, text=descr, bg="#FFFFFF", anchor="w").grid(row=idx, column=0, sticky="w", pady=4, padx=(0, 8))
            var = tk.StringVar(value=str(value))
            self.criterio_vars[title] = var
            cb = ttk.Combobox(self.frm_criterios, textvariable=var, values=("", "SIM", "NAO", "PENDENTE"), state="readonly", width=14)
            cb.grid(row=idx, column=1, sticky="ew", pady=4)
            tk.Button(self.frm_criterios, text="SIM", width=6, command=lambda t=title: self._set_criterio_current(t, "SIM")).grid(row=idx, column=2, padx=(6, 2))
            tk.Button(self.frm_criterios, text="NAO", width=6, command=lambda t=title: self._set_criterio_current(t, "NAO")).grid(row=idx, column=3, padx=(0, 2))
            var.trace_add("write", lambda *_args, t=title, v=var: self._on_criterio_changed(t, v.get()))
        pend_row = len(self.criteria_info) + 1
        tk.Label(self.frm_criterios, text="Pendencias:", bg="#FFFFFF", anchor="w", font=("Segoe UI", 9, "bold")).grid(
            row=pend_row, column=0, sticky="w", pady=(10, 4)
        )
        tk.Label(self.frm_criterios, text=p.get("pend", ""), bg="#FFFFFF", anchor="w", justify="left", wraplength=360).grid(
            row=pend_row + 1, column=0, columnspan=4, sticky="w", pady=(0, 6)
        )
        self.frm_criterios.columnconfigure(1, weight=1)
        self.frm_criterios.columnconfigure(2, weight=0)
        self.frm_criterios.columnconfigure(3, weight=0)

    def _append_history(self, payload: dict):
        event = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "user": getpass.getuser(),
            "workbook": str(self.workbook_path),
            **payload,
        }
        with self._history_path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(event, ensure_ascii=False) + "\n")

    def _snapshot_pending(self) -> dict[int, dict[str, str]]:
        return {row: changes.copy() for row, changes in self.pending_by_row.items()}

    def _push_undo(self):
        self._undo_stack.append(self._snapshot_pending())
        if len(self._undo_stack) > 120:
            self._undo_stack = self._undo_stack[-120:]

    def _set_criterio_current(self, title: str, value: str):
        if title not in self.criterio_vars:
            return
        self.criterio_vars[title].set(value)

    def _on_criterio_changed(self, title: str, value: str):
        if self.current_row is None:
            return
        if self.unified_mode:
            return
        self._push_undo()
        self._stage_row_changes(self.current_row, {title: value}, source="inline")

    def stage_unified_current(self):
        if self.current_row is None:
            return
        values = {k: v.get().strip() for k, v in self.criterio_vars.items()}
        if hasattr(self, "_unified_text"):
            values["oqf"] = self._unified_text.get("1.0", "end").strip()
        self._push_undo()
        self._stage_row_changes(self.current_row, values, source="unified")

    def choose_source_folder(self):
        initial = self.source_dir_var.get().strip() or str(self.workbook_path.parent)
        chosen = filedialog.askdirectory(title="Escolha a pasta das planilhas individuais", initialdir=initial)
        if chosen:
            self.source_dir_var.set(chosen)
            self.status_var.set(f"Pasta de busca definida: {chosen}")

    def open_indicator_source(self):
        if not self.unified_mode:
            return
        raw_code = getattr(self, "_unified_indicator_var", tk.StringVar()).get().strip()
        if not raw_code:
            self.status_var.set("Selecione um indicador.")
            return
        m = re.match(r"(C\d+)", raw_code, re.I)
        code = (m.group(1).upper() if m else raw_code.upper())
        candidates = []
        current_sources = self._discover_general_source_files()
        if current_sources:
            for p in current_sources:
                mm = infer_indicator_code_from_path(p)
                if mm and mm.upper() == code:
                    candidates.append(p)
        else:
            folder = Path(self.source_dir_var.get().strip() or self.workbook_path.parent)
            if not folder.exists():
                messagebox.showinfo("Pasta invalida", f"Pasta nao encontrada:\n{folder}")
                return
            candidates = [
                p for p in folder.rglob("*.xlsx")
                if (infer_indicator_code_from_path(p) or "").upper() == code
                if "interativa" not in p.name.lower()
                and "cruz" not in p.name.lower()
                and "compar" not in p.name.lower()
                and "unificad" not in p.name.lower()
            ]
        if not candidates:
            where = "selecao atual" if self.general_source_files else str(Path(self.source_dir_var.get().strip() or self.workbook_path.parent))
            messagebox.showinfo("Nao encontrado", f"Nenhuma planilha do {code} encontrada em:\n{where}")
            return
        latest = max(candidates, key=lambda p: p.stat().st_mtime)
        try:
            patient_name = ""
            if self.current_row is not None and self.current_row in self.record_map:
                patient_name = str(self.record_map[self.current_row].get("nome", "")).strip()
            launch_editor(master=self, workbook_path=latest, initial_patient_name=patient_name)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc), parent=self)

    def undo_last_change(self):
        if not self._undo_stack:
            self.status_var.set("Nada para desfazer.")
            return
        self.pending_by_row = self._undo_stack.pop()
        self._build_records_view()
        self._apply_filter()
        self.status_var.set("Ultima alteracao desfeita.")
        self._append_history({"action": "undo"})

    def _update_chart(self):
        if not hasattr(self, "chart_canvas"):
            return
        cvs = self.chart_canvas
        cvs.delete("all")
        width = max(420, int(cvs.winfo_width() or 420))
        height = max(170, int(cvs.winfo_height() or 170))
        pad = 28
        labels = ["URGENTE", "ALTA", "MONITORAR", "CONCLUIDO"]
        colors = {
            "URGENTE": "#D9534F",
            "ALTA": "#F0AD4E",
            "MONITORAR": "#F7D97B",
            "CONCLUIDO": "#5CB85C",
        }
        counts = {lbl: 0 for lbl in labels}
        for rec in self.records:
            counts[self._norm_prio(rec["prio"])] = counts.get(self._norm_prio(rec["prio"]), 0) + 1
        max_count = max(max(counts.values()), 1)
        col_w = (width - pad * 2) // len(labels)
        self._chart_regions = []
        for i, lbl in enumerate(labels):
            x0 = pad + i * col_w + 10
            x1 = pad + (i + 1) * col_w - 10
            bar_h = int((height - 65) * (counts[lbl] / max_count))
            y1 = height - 30
            y0 = y1 - bar_h
            cvs.create_rectangle(x0, y0, x1, y1, fill=colors[lbl], outline="")
            cvs.create_text((x0 + x1) // 2, y0 - 10, text=str(counts[lbl]), fill="#1F4E79", font=("Segoe UI", 9, "bold"))
            cvs.create_text((x0 + x1) // 2, y1 + 14, text=lbl, fill="#1F4E79", font=("Segoe UI", 9))
            self._chart_regions.append((x0, y0, x1, y1, lbl))

    def _on_chart_click(self, event):
        for x0, y0, x1, y1, lbl in self._chart_regions:
            if x0 <= event.x <= x1 and y0 <= event.y <= y1:
                if self.filter_prio_var.get() == lbl:
                    self.filter_prio_var.set("Todas")
                else:
                    self.filter_prio_var.set(lbl)
                self._apply_filter()
                return

    def _stage_row_changes(
        self,
        row: int,
        values_by_title: dict[str, str],
        source: str,
        *,
        refresh_ui: bool = True,
        log_history: bool = True,
    ):
        base = next((r for r in self.base_records if r["row"] == row), None)
        if not base:
            return
        existing = self.pending_by_row.get(row, {}).copy()
        original = existing.copy()
        diffs = []
        for title, value in values_by_title.items():
            if self.unified_mode:
                old_value = str(existing.get(title, base.get(title, "")))
                if str(value) == str(base.get(title, "")):
                    existing.pop(title, None)
                else:
                    existing[title] = value
                if str(value) != old_value:
                    diffs.append({"criterio": title, "de": old_value, "para": value})
                continue
            idx = self.criteria_index.get(title)
            if idx is None:
                continue
            normalized = _normalize_status(value)
            old_value = _normalize_status(existing.get(title, base["statuses"][idx]))
            base_norm = _normalize_status(base["statuses"][idx])
            if normalized == base_norm:
                existing.pop(title, None)
            else:
                existing[title] = normalized
            if normalized != old_value:
                diffs.append({"criterio": title, "de": old_value, "para": normalized})
        if existing:
            self.pending_by_row[row] = existing
        else:
            self.pending_by_row.pop(row, None)
        if original == existing:
            return
        if log_history:
            self._append_history({
                "action": "stage",
                "source": source,
                "row": row,
                "paciente": str(base.get("nome", "")),
                "changes": diffs,
            })
        if refresh_ui:
            if self.light_mode_enabled:
                self._schedule_view_refresh(80)
            else:
                self._build_records_view()
                self._apply_filter()
                if self.tree.exists(str(row)):
                    self.tree.selection_set(str(row))
                    self.tree.focus(str(row))
            self.status_var.set(f"Alteracao preparada. Linhas com alteracoes pendentes: {len(self.pending_by_row)}")

    def _rows_for_mass_scope(self) -> list[int]:
        scope = self.mass_scope_var.get().strip()
        if scope == "Selecionados":
            return [int(i) for i in self.tree.selection()]
        if scope == "Filtrados":
            return [int(i) for i in self.tree.get_children("")]
        return [p["row"] for p in self.records]

    def apply_mass_change(self):
        criterio = self.mass_criterio_var.get().strip()
        if not criterio:
            self.status_var.set("Escolha um criterio para alteracao em massa.")
            return
        rows = self._rows_for_mass_scope()
        if not rows:
            self.status_var.set("Nenhuma linha encontrada para alteracao em massa.")
            return
        value = self.mass_valor_var.get().strip()
        self._push_undo()
        for row in rows:
            self._stage_row_changes(row, {criterio: value}, source="mass", refresh_ui=False, log_history=False)
        self._schedule_view_refresh(20 if self.light_mode_enabled else 0)
        self._append_history({
            "action": "mass_stage",
            "criterio": criterio,
            "value": value,
            "scope": self.mass_scope_var.get().strip(),
            "rows": len(rows),
        })
        self.status_var.set(f"Alteracao em massa preparada em {len(rows)} linhas.")

    def flush_buffer(self):
        if self._busy:
            return
        if not self.pending_by_row:
            self.status_var.set("Nao ha alteracoes pendentes para salvar.")
            return

        pending = {row: changes.copy() for row, changes in self.pending_by_row.items()}
        self._set_busy(True)
        self.status_var.set("Salvando alteracoes na planilha...")

        def worker():
            fast_mode = bool(self.fast_save_var.get())
            stamp = datetime.now().strftime("%Y-%m-%d_%Hh%Mm%Ss")
            backup = self._editor_backup_dir() / f"{self.workbook_path.stem}__backup__{stamp}{self.workbook_path.suffix}"
            backup_ok = None
            if not fast_mode:
                try:
                    shutil.copy2(self.workbook_path, backup)
                    backup_ok = backup
                except Exception:
                    backup_ok = None

            wb, ws, header_row, cols = self._load_context()
            if self.unified_mode:
                col_by_title = {k: v for k, v in self.unified_cols.items() if v}
            else:
                col_by_title = {title: col for _letter, title, col in cols["criterios"]}
            applied_cells = 0
            for row, changes in pending.items():
                for title, value in changes.items():
                    col = col_by_title.get(title)
                    if col is None:
                        continue
                    ws.cell(row, col).value = value
                    applied_cells += 1
            if not self.unified_mode:
                cols = _ensure_support_columns(ws, header_row, cols)
                _update_data_sheet(ws, header_row, cols)
            wb.save(self.workbook_path)
            wb.close()
            if not self.unified_mode and not fast_mode:
                refresh_interactive_workbook(self.workbook_path)
            return applied_cells, backup_ok, fast_mode

        def on_ok(result):
            applied_cells, backup_ok, fast_mode = result
            base_by_row = {r["row"]: r for r in self.base_records}
            changed_names: list[str] = []
            for row, changes in pending.items():
                base = base_by_row.get(row)
                if not base:
                    continue
                nm = str(base.get("nome", "")).strip()
                if nm:
                    changed_names.append(nm)
                for title, value in changes.items():
                    idx = self.criteria_index.get(title)
                    if idx is None:
                        continue
                    base["statuses"][idx] = value
            self.pending_by_row.clear()
            self._append_history({
                "action": "commit_buffer",
                "rows": len(pending),
                "cells": applied_cells,
                "backup": backup_ok.name if backup_ok else None,
                "save_mode": "rapido" if fast_mode else "completo",
            })
            self._set_busy(False)
            mode_label = "rapido" if fast_mode else "completo"
            self.status_var.set(f"Alteracoes salvas ({mode_label}) ({len(pending)} linhas, {applied_cells} campos).")
            self._build_records_view()
            self._apply_filter()
            parent = self.master if isinstance(self.master, EditorPlanilhaApp) else None
            if parent is not None and not self.unified_mode:
                try:
                    parent.on_indicator_saved(self.workbook_path, changed_names)
                except Exception:
                    pass

        def on_err(exc):
            self._set_busy(False)
            self.status_var.set(f"Erro ao salvar alteracoes: {exc}")
            messagebox.showerror("Erro ao salvar", str(exc), parent=self)

        def runner():
            try:
                result = worker()
                self.after(0, lambda: on_ok(result))
            except Exception as exc:
                self.after(0, lambda: on_err(exc))

        threading.Thread(target=runner, daemon=True).start()

    def open_change_history(self):
        win = tk.Toplevel(self)
        win.title("Historico de alteracoes")
        win.geometry("980x420")
        cols = ("quando", "paciente", "acao", "alteracao")
        tree = ttk.Treeview(win, columns=cols, show="headings")
        for c, w in [("quando", 170), ("paciente", 220), ("acao", 120), ("alteracao", 430)]:
            tree.heading(c, text=c.title())
            tree.column(c, width=w, anchor="w")
        tree.pack(fill="both", expand=True, padx=10, pady=10)
        if not self._history_path.exists():
            tree.insert("", "end", values=(datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "-", "info", "Sem historico salvo ainda."))
            return
        lines = self._history_path.read_text(encoding="utf-8", errors="ignore").splitlines()[-500:]
        for ln in reversed(lines):
            try:
                evt = json.loads(ln)
            except Exception:
                continue
            when = evt.get("ts", "-")
            action = evt.get("action", "-")
            paciente = evt.get("paciente", "-")
            changes = evt.get("changes")
            if isinstance(changes, list):
                detail = "; ".join(
                    f"{d.get('criterio','')}: {d.get('de','')} -> {d.get('para','')}"
                    for d in changes
                ) or "-"
            else:
                detail = json.dumps({k: v for k, v in evt.items() if k not in {"ts", "action", "source", "user", "workbook", "paciente"}}, ensure_ascii=False)
            tree.insert("", "end", values=(when, paciente, action, detail))

    def _open_workbook(self):
        if self.unified_mode:
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="Abrir indicador atual", command=self.open_indicator_source)
            menu.add_command(label="Abrir pasta das fontes", command=self._open_sources_folder)
            try:
                btn = self.focus_get()
                if btn is not None and hasattr(btn, "winfo_rootx"):
                    x = btn.winfo_rootx()
                    y = btn.winfo_rooty() + btn.winfo_height()
                else:
                    x = self.winfo_pointerx()
                    y = self.winfo_pointery()
                menu.tk_popup(x, y)
            finally:
                menu.grab_release()
            return
        try:
            os.startfile(self.workbook_path)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc), parent=self)

    def _manual_refresh_planilha(self):
        if self.unified_mode and self.general_source_files:
            self._refresh_general_from_current_files("Atualizando visao geral...")
            return
        self.reload_data(async_mode=True)

    def _open_sources_folder(self):
        folder = None
        if self.general_source_files:
            folder = self.general_source_files[0].parent
        else:
            folder = Path(self.source_dir_var.get().strip() or self.workbook_path.parent)
        if not folder.exists():
            messagebox.showwarning("Pasta nao encontrada", f"Pasta nao encontrada:\n{folder}", parent=self)
            return
        try:
            os.startfile(folder)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc), parent=self)

    def _on_close(self):
        if self._view_refresh_after_id:
            try:
                self.after_cancel(self._view_refresh_after_id)
            except Exception:
                pass
            self._view_refresh_after_id = None
        if self._auto_general_after_id:
            try:
                self.after_cancel(self._auto_general_after_id)
            except Exception:
                pass
            self._auto_general_after_id = None
        self.destroy()

def _load_criterios(workbook_path: Path):
    wb = load_workbook(workbook_path)
    ws = next((wb[n] for n in wb.sheetnames if n.startswith("ðŸ“‹ Dados") or n.startswith("Dados")), wb[wb.sheetnames[0]])
    header_row = _detect_header(ws)
    cols = _ensure_support_columns(ws, header_row, _detect_columns(ws, header_row))
    criterios = cols["criterios"]
    wb.close()
    return criterios


class ClonadorApp(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__(master=master)
        self.title("APS - Gerar planilha e abrir editor")
        self.geometry("760x280")
        self.configure(bg="#EAF2F8")
        self.path_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Selecione uma planilha base (.xlsx).")
        self._build()

    def _build(self):
        frm = tk.Frame(self, bg="#EAF2F8")
        frm.pack(fill="both", expand=True, padx=20, pady=20)
        tk.Label(frm, text="APS - GERAR PLANILHA + ABRIR EDITOR", bg="#1F4E79", fg="white", font=("Segoe UI", 13, "bold"), pady=10).pack(fill="x")
        tk.Label(frm, text="Novo fluxo: gerar a cÃ³pia interativa e editar no aplicativo interno, sem depender de macro.", bg="#EAF2F8", anchor="w").pack(fill="x", pady=(10, 0))
        line = tk.Frame(frm, bg="#EAF2F8")
        line.pack(fill="x", pady=12)
        tk.Entry(line, textvariable=self.path_var).pack(side="left", fill="x", expand=True)
        tk.Button(line, text="Escolher", command=self.choose).pack(side="left", padx=8)
        tk.Button(frm, text="Gerar e abrir editor", command=self.run, bg="#2E75B6", fg="white", font=("Segoe UI", 11, "bold")).pack(fill="x")
        tk.Label(frm, textvariable=self.status_var, bg="#EAF2F8", fg="#1F4E79").pack(fill="x", pady=(10, 0))

    def choose(self):
        path = filedialog.askopenfilename(title="Selecione a planilha base", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.path_var.set(path)

    def run(self):
        path = self.path_var.get().strip()
        if not path:
            messagebox.showwarning("AtenÃ§Ã£o", "Selecione uma planilha.", parent=self)
            return
        try:
            out = clone_interactive(path)
            self.status_var.set(f"Planilha criada: {out.name}")
            messagebox.showinfo("ConcluÃ­do", f"Planilha criada com sucesso:\n\n{out}", parent=self)
            launch_editor(self, out)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc), parent=self)


# -----------------------------------------------------------------------------
# Launchers
# -----------------------------------------------------------------------------


def launch_editor(master=None, workbook_path=None, initial_patient_name: str | None = None):
    if not workbook_path:
        workbook_path = filedialog.askopenfilename(title="Selecione a planilha APS", filetypes=[("Excel", "*.xlsx")])
        if not workbook_path:
            return None
    return EditorPlanilhaApp(workbook_path, master=master, initial_patient_name=initial_patient_name)



def launch_clonador(master=None):
    return ClonadorApp(master=master)



def main():
    root = tk.Tk()
    root.withdraw()
    app = ClonadorApp(root)
    app.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()


if __name__ == "__main__":
    main()


