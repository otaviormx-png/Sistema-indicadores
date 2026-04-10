"""
aps_historico.py - Tela de historico de execucoes.
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl


def _read_summary(path: Path) -> dict:
    info = {"pacientes": 0, "media_pts": "-", "codigo": "-"}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        info["pacientes"] = sum(1 for r in rows if any(c for c in r))
        pts = []
        headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
        try:
            idx_pts = headers.index("Pontuacao")
            pts = [r[idx_pts] for r in rows if r[idx_pts] is not None]
        except Exception:
            pass
        if pts:
            try:
                info["media_pts"] = f"{sum(float(p) for p in pts) / len(pts):.1f}"
            except Exception:
                pass
        wb.close()
    except Exception:
        pass
    return info


def _parse_filename(name: str) -> tuple[str, str]:
    stem = Path(name).stem
    parts = stem.split("_")
    if len(parts) >= 3:
        code = parts[0].upper()
        try:
            dt = datetime.strptime(f"{parts[1]}_{parts[2]}", "%Y%m%d_%H%M%S")
            return code, dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
    return stem, "-"


class HistoricoWindow(tk.Toplevel):
    def __init__(self, master, out_dir: Path):
        super().__init__(master)
        self.title("Historico de execucoes - APS Suite")
        self.geometry("860x500")
        self.minsize(700, 380)
        self.configure(bg="#EAF2F8")
        self.out_dir = out_dir
        self._build_ui()
        self._load()

    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        top = ttk.Label(
            self,
            text="Historico de execucoes",
            background="#1F4E79",
            foreground="white",
            font=("Segoe UI", 13, "bold"),
            padding=10,
            anchor="center",
        )
        top.pack(fill="x")

        box = ttk.LabelFrame(self, text="Resultados gerados")
        box.pack(fill="both", expand=True, padx=12, pady=12)
        box.rowconfigure(0, weight=1)
        box.columnconfigure(0, weight=1)

        cols = ("indicador", "data", "arquivo", "pacientes", "media_pts")
        self.tree = ttk.Treeview(box, columns=cols, show="headings")
        self.tree.heading("indicador", text="Indicador")
        self.tree.heading("data", text="Data/hora")
        self.tree.heading("arquivo", text="Arquivo")
        self.tree.heading("pacientes", text="Pacientes")
        self.tree.heading("media_pts", text="Media pts")
        self.tree.column("indicador", width=90, anchor="center")
        self.tree.column("data", width=140, anchor="center")
        self.tree.column("arquivo", width=320)
        self.tree.column("pacientes", width=90, anchor="center")
        self.tree.column("media_pts", width=90, anchor="center")
        self.tree.grid(row=0, column=0, sticky="nsew")

        ys = ttk.Scrollbar(box, orient="vertical", command=self.tree.yview)
        ys.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=ys.set)
        self.tree.bind("<Double-1>", self._on_double_click)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))
        ttk.Button(btn_frame, text="Abrir selecionado", command=self._open_selected).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Exportar PDF", command=self._export_pdf).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Reprocessar pendentes", command=self._reprocessar_pendentes).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Atualizar lista", command=self._load).pack(side="left", padx=(0, 8))

    def _load(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        if not self.out_dir.exists():
            return
        files = sorted([f for f in self.out_dir.glob("*.xlsx") if f.is_file()], key=lambda f: f.stat().st_mtime, reverse=True)
        for f in files:
            code, data = _parse_filename(f.name)
            info = _read_summary(f)
            self.tree.insert("", "end", values=(code, data, f.name, info["pacientes"], info["media_pts"]), tags=(str(f),))

    def _open_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Nenhum selecionado", "Clique em um resultado para selecionar.")
            return
        tags = self.tree.item(sel[0], "tags")
        if not tags:
            return
        path = Path(tags[0])
        if not path.exists():
            messagebox.showerror("Nao encontrado", f"Arquivo nao existe mais:\n{path}")
            return
        try:
            os.startfile(path)
        except Exception:
            messagebox.showinfo("Arquivo", str(path))

    def _on_double_click(self, _event):
        self._open_selected()

    def _export_pdf(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Nenhum selecionado", "Clique em um resultado para selecionar.")
            return
        tags = self.tree.item(sel[0], "tags")
        if not tags:
            return
        path = Path(tags[0])
        if not path.exists():
            messagebox.showerror("Nao encontrado", f"Arquivo nao existe mais:\n{path}")
            return
        try:
            import aps_exportar_pdf

            pdf = aps_exportar_pdf.gerar_pdf(path)
            messagebox.showinfo("PDF gerado", f"Resumo exportado para:\n{pdf}")
            try:
                os.startfile(pdf)
            except Exception:
                pass
        except ImportError:
            messagebox.showerror("reportlab nao instalado", "Instale: pip install reportlab")
        except Exception as exc:
            messagebox.showerror("Erro ao gerar PDF", str(exc))

    def _reprocessar_pendentes(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Nenhum selecionado", "Selecione um resultado para reprocessar.")
            return
        tags = self.tree.item(sel[0], "tags")
        if not tags:
            return
        xlsx_anterior = Path(tags[0])
        if not xlsx_anterior.exists():
            messagebox.showerror("Nao encontrado", f"Arquivo nao existe:\n{xlsx_anterior}")
            return

        in_dir = filedialog.askdirectory(title="Escolha a pasta com os arquivos brutos atualizados")
        if not in_dir:
            return
        try:
            import aps_reprocessar

            resultado = aps_reprocessar.reprocessar_pendentes(
                xlsx_anterior=xlsx_anterior,
                in_dir=Path(in_dir),
                out_dir=self.out_dir,
            )
            msg = (
                f"Reprocessamento concluido!\n\n"
                f"Pacientes preservados: {resultado['completos_preservados']}\n"
                f"Reprocessados: {resultado['reprocessados']}\n"
                f"Novos completos: {resultado['novos_completos']}\n\n"
                f"Resultado salvo em:\n{resultado['saida'].name}"
            )
            messagebox.showinfo("Reprocessamento concluido", msg)
            self._load()
            try:
                os.startfile(resultado["saida"])
            except Exception:
                pass
        except Exception as exc:
            messagebox.showerror("Erro no reprocessamento", str(exc))


def launch_historico(master, out_dir: Path) -> HistoricoWindow | None:
    try:
        win = HistoricoWindow(master, out_dir)
        win.lift()
        win.focus_force()
        return win
    except Exception as exc:
        messagebox.showerror("Erro ao abrir historico", str(exc))
        return None
