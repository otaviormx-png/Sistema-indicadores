"""
aps_exportar_pdf.py — Exporta o resumo gerencial de um xlsx gerado
pela APS Suite para um PDF de 1 página.

Usa apenas a biblioteca 'reportlab' (já listada em pyproject.toml).
Instale se necessário:  pip install reportlab
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

import openpyxl


# ------------------------------------------------------------------
# Paleta (mesma do APS Suite)
# ------------------------------------------------------------------
_AZUL  = colors.HexColor("#1F4E79")
_AZUL2 = colors.HexColor("#2E75B6")
_AZUL3 = colors.HexColor("#D6E4F0")
_VERDE = colors.HexColor("#C6EFCE")
_VERDE_TXT = colors.HexColor("#276221")
_VERM  = colors.HexColor("#FFC7CE")
_VERM_TXT = colors.HexColor("#9C0006")
_AMAR  = colors.HexColor("#FFEB9C")
_AMAR_TXT = colors.HexColor("#9C5700")
_ROXO  = colors.HexColor("#7030A0")
_CINZA = colors.HexColor("#F2F2F2")
_BRANCO = colors.white
_PRETO  = colors.black


def _read_resumo_sheet(xlsx_path: Path) -> dict:
    """Lê a aba Resumo e a aba Dados para montar os dados do PDF."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    info: dict = {
        "titulo": "Relatório APS",
        "codigo": "Cx",
        "total": 0,
        "media": 0.0,
        "completos": 0,
        "busca": 0,
        "classes": {},
        "criterios": [],
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }

    # Aba de dados (primeira)
    ws_dados = wb.worksheets[0]
    titulo_cell = ws_dados.cell(1, 1).value or ""
    info["titulo"] = str(titulo_cell).split("|")[0].strip()

    headers = [ws_dados.cell(3, c).value for c in range(1, ws_dados.max_column + 1)]
    rows_data = list(ws_dados.iter_rows(min_row=4, values_only=True))

    try:
        idx_pts = headers.index("Pontuação")
        idx_cls = headers.index("Classificação")
        pontuacoes = [r[idx_pts] for r in rows_data if r[idx_pts] is not None]
        info["total"] = len(pontuacoes)
        info["media"] = round(sum(float(p) for p in pontuacoes) / len(pontuacoes), 1) if pontuacoes else 0
        info["completos"] = sum(1 for p in pontuacoes if float(p) >= 100)
        info["busca"] = info["total"] - info["completos"]
        classes_raw = [r[idx_cls] for r in rows_data if r[idx_cls]]
        for cls in ["Ótimo", "Bom", "Suficiente", "Regular"]:
            info["classes"][cls] = classes_raw.count(cls)
    except (ValueError, TypeError):
        pass

    # Critérios (colunas entre clínicas e Pontuação)
    criterio_cols = [h for h in headers if h and " - " in str(h)
                     and h not in ("Pontuação", "Classificação", "Prioridade", "Pendências")]
    for col in criterio_cols:
        try:
            idx = headers.index(col)
            vals = [r[idx] for r in rows_data if r[idx] is not None]
            sim = sum(1 for v in vals if str(v).upper() == "SIM")
            total = len(vals)
            pct = round(sim / total * 100, 1) if total else 0
            info["criterios"].append({"label": col, "sim": sim, "total": total, "pct": pct})
        except Exception:
            pass

    wb.close()
    return info


def gerar_pdf(xlsx_path: Path, pdf_path: Path | None = None) -> Path:
    """Gera o PDF de resumo a partir do xlsx. Retorna o caminho do PDF."""
    if not REPORTLAB_OK:
        raise ImportError(
            "reportlab não está instalado.\n"
            "Execute:  pip install reportlab"
        )

    if pdf_path is None:
        pdf_path = xlsx_path.with_suffix(".pdf")

    info = _read_resumo_sheet(xlsx_path)

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    s_titulo = ParagraphStyle("titulo", parent=styles["Normal"],
                              fontSize=16, textColor=_BRANCO,
                              alignment=TA_CENTER, spaceAfter=4, fontName="Helvetica-Bold")
    s_sub = ParagraphStyle("sub", parent=styles["Normal"],
                           fontSize=9, textColor=_AZUL2,
                           alignment=TA_CENTER, spaceAfter=2)
    s_sec = ParagraphStyle("sec", parent=styles["Normal"],
                           fontSize=11, textColor=_BRANCO,
                           fontName="Helvetica-Bold", alignment=TA_CENTER)
    s_body = ParagraphStyle("body", parent=styles["Normal"],
                            fontSize=9, textColor=_PRETO, alignment=TA_LEFT)

    W = A4[0] - 3.6 * cm  # largura útil

    story = []

    # Cabeçalho
    tbl_header = Table(
        [[Paragraph(info["titulo"], s_titulo)],
         [Paragraph(f"Gerado em: {info['gerado_em']}  |  Fonte: bruto e-SUS", s_sub)]],
        colWidths=[W]
    )
    tbl_header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), _AZUL),
        ("BACKGROUND", (0, 1), (0, 1), _AZUL3),
        ("BOX", (0, 0), (-1, -1), 0.5, _AZUL),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(tbl_header)
    story.append(Spacer(1, 0.4 * cm))

    # Cards de totais
    card_data = [
        ["TOTAL\nPESSOAS", "PONTUAÇÃO\nMÉDIA", "COMPLETOS\n(100 pts)", "EM BUSCA\nATIVA"],
        [str(info["total"]), str(info["media"]), str(info["completos"]), str(info["busca"])],
    ]
    card_w = W / 4
    tbl_cards = Table(card_data, colWidths=[card_w] * 4, rowHeights=[20, 28])
    tbl_cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (1, 1), _AZUL3),
        ("BACKGROUND", (2, 0), (2, 1), _VERDE),
        ("BACKGROUND", (3, 0), (3, 1), _VERM),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("TEXTCOLOR", (0, 0), (1, 1), _AZUL),
        ("TEXTCOLOR", (2, 0), (2, 1), _VERDE_TXT),
        ("TEXTCOLOR", (3, 0), (3, 1), _VERM_TXT),
        ("BOX", (0, 0), (-1, -1), 0.5, _AZUL2),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, _AZUL2),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl_cards)
    story.append(Spacer(1, 0.4 * cm))

    # Seção: Distribuição por classificação
    tbl_sec1 = Table([[Paragraph("Distribuição por classificação", s_sec)]], colWidths=[W])
    tbl_sec1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _ROXO),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tbl_sec1)
    story.append(Spacer(1, 0.15 * cm))

    cls_colors = {
        "Ótimo": (_VERDE, _VERDE_TXT),
        "Bom": (colors.HexColor("#E2EFDA"), _VERDE_TXT),
        "Suficiente": (_AMAR, _AMAR_TXT),
        "Regular": (_VERM, _VERM_TXT),
    }
    cls_rows = [["Classificação", "Qtd", "Percentual"]]
    total = info["total"] or 1
    for cls in ["Ótimo", "Bom", "Suficiente", "Regular"]:
        qtd = info["classes"].get(cls, 0)
        pct = f"{round(qtd / total * 100, 1)}%"
        cls_rows.append([cls, str(qtd), pct])

    tbl_cls = Table(cls_rows, colWidths=[W * 0.5, W * 0.25, W * 0.25])
    cls_style = [
        ("BACKGROUND", (0, 0), (-1, 0), _AZUL3),
        ("TEXTCOLOR", (0, 0), (-1, 0), _AZUL),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("BOX", (0, 0), (-1, -1), 0.5, _AZUL2),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, cls in enumerate(["Ótimo", "Bom", "Suficiente", "Regular"], start=1):
        bg, fg = cls_colors[cls]
        cls_style.append(("BACKGROUND", (0, i), (0, i), bg))
        cls_style.append(("TEXTCOLOR", (0, i), (0, i), fg))
    tbl_cls.setStyle(TableStyle(cls_style))
    story.append(tbl_cls)
    story.append(Spacer(1, 0.4 * cm))

    # Seção: Adesão por critério
    if info["criterios"]:
        tbl_sec2 = Table([[Paragraph("Adesão por critério", s_sec)]], colWidths=[W])
        tbl_sec2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _AZUL),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl_sec2)
        story.append(Spacer(1, 0.15 * cm))

        crit_rows = [["Critério", "Atingiram", "Total", "Adesão"]]
        for c in info["criterios"]:
            label = c["label"].split(" - ", 1)[-1] if " - " in c["label"] else c["label"]
            pct_val = c["pct"]
            crit_rows.append([label, str(c["sim"]), str(c["total"]), f"{pct_val}%"])

        tbl_crit = Table(crit_rows,
                         colWidths=[W * 0.55, W * 0.15, W * 0.15, W * 0.15])
        crit_style = [
            ("BACKGROUND", (0, 0), (-1, 0), _AZUL3),
            ("TEXTCOLOR", (0, 0), (-1, 0), _AZUL),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("BOX", (0, 0), (-1, -1), 0.5, _AZUL2),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_BRANCO, _CINZA]),
        ]
        for i, c in enumerate(info["criterios"], start=1):
            pct = c["pct"]
            bg = _VERDE if pct > 75 else _AMAR if pct > 50 else _VERM
            fg = _VERDE_TXT if pct > 75 else _AMAR_TXT if pct > 50 else _VERM_TXT
            crit_style.append(("BACKGROUND", (3, i), (3, i), bg))
            crit_style.append(("TEXTCOLOR", (3, i), (3, i), fg))
        tbl_crit.setStyle(TableStyle(crit_style))
        story.append(tbl_crit)

    # Rodapé
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        f"APS Suite  •  {info['gerado_em']}  •  Dados extraídos do e-SUS",
        ParagraphStyle("rodape", parent=styles["Normal"],
                       fontSize=7, textColor=colors.HexColor("#888888"),
                       alignment=TA_CENTER)
    ))

    doc.build(story)
    return pdf_path
